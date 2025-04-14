import shutil
from typing import Dict

from PT import *
from openpyxl import load_workbook
from publicFun import *
import json
import sys
import os
import time
import zipfile
import csv
import subprocess
import threading
import multiprocessing
import chardet
import codecs
import io
random.seed(1)


def copyFile(fileDir, tarDir, picknumber):
    pathDir = os.listdir(fileDir)  # 取图片的原始路径
    sample = random.sample(pathDir, picknumber)  # 随机选取picknumber数量的样本图片
    for i in range(len(sample)):
        shutil.copy(fileDir + sample[i], tarDir + "input{}.txt".format(i))
    return


def movefile(tarpath, tardir, oridir):
    if os.path.exists(tardir):
        if os.path.exists(tardir + '/' + tarpath):
            # os.remove(tarpath)
            pass
        else:
            shutil.move(oridir + tarpath, tardir)
    else:
        os.makedirs(tardir)
        shutil.move(oridir + tarpath, tardir)


def zip_file(src_dir):

    zip_name = src_dir + '.zip'

    z = zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED)

    for dirpath, dirnames, filenames in os.walk(src_dir):

        fpath = dirpath.replace(src_dir, '')

        fpath = fpath and fpath + os.sep or ''

        for filename in filenames:

            z.write(os.path.join(dirpath, filename), fpath+filename)

    z.close()


def un_zip(file_name):
    """unzip zip file"""
    zip_file = zipfile.ZipFile(file_name)
    tardir = '/Applications/work/data/MT/FS/PT/'
    if os.path.exists(file_name[:-4]):
        return
    for names in zip_file.namelist():
        zip_file.extract(names, tardir)
    zip_file.close()


def getMG(mu, mr_list, num_of_samples, originaloutput, ts):
    outputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/PT/RandomOutput{}.csv'.format(mu)
    csvFile = open(outputdir, "r")
    reader = csv.reader(csvFile)
    mutateoutput = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            mutateoutput[item[0]] = item[1].split("\n")
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    MGS = []
    for i in range(num_of_samples):
        MGs = []
        MG = [0] * len(mr_list)
        ts.setInputOutput("input{}".format(i), "output{}".format(i))
        original_output = MR().getResults(ts, mutateoutput)
        followup_ts = copy.copy(ts)
        for j in range(len(mr_list)):
            followup_ts.setInputOutput("input{}_{}".format(i, j), "output{}_{}".format(i, j))
            mr = mr_list[j]
            followup_output = mr.getResults(followup_ts, mutateoutput)
            expected_output = mr.getExpectedOutput(ts, original_output)
            isViolate = mr.assertViolation(expected_output, followup_output)
            if isViolate:
                MG[j] = 1
            elif (mutateoutput["output{}".format(i)] != originaloutput["output{}".format(i)] or
                  mutateoutput["output{}_{}".format(i, j)] != originaloutput["output{}_{}".format(i, j)]):
                MG[j] = 3
        MGs.append(MG)
        for m in range(len(mr_list)):
            MG = [0] * len(mr_list)
            ts.setInputOutput("input{}_{}".format(i, m), "output{}_{}".format(i, m))
            original_output = MR().getResults(ts, mutateoutput)
            followup_ts = copy.copy(ts)
            for n in range(len(mr_list)):
                followup_ts.setInputOutput("input{}_{}_{}".format(i, m, n), "output{}_{}_{}".format(i, m, n))
                mr = mr_list[n]
                followup_output = mr.getResults(followup_ts, mutateoutput)
                expected_output = mr.getExpectedOutput(ts, original_output)
                isViolate = mr.assertViolation(expected_output, followup_output)
                if isViolate:
                    MG[n] = 1
                elif (mutateoutput["output{}_{}".format(i, m)] != originaloutput["output{}_{}".format(i, m)] or
                      mutateoutput["output{}_{}_{}".format(i, m, n)] != originaloutput["output{}_{}_{}".format(i, m, n)]):
                    MG[n] = 3
            MGs.append(MG)
        MGS.append(MGs)
    return MGS


def getpf(mu, ts, num_of_samples, mr_list, outputdata):
    # pf = []
    # for i in range(num_of_samples):
    #     result = [0] * (len(mr_list) * (len(mr_list) + 1) + 1)
    #     ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(0, i))
    #     original_output = MR().getResults(ts)
    #     program_ts = ts
    #     program_ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(mu, i))
    #     program_output = MR().getResults(program_ts)
    #     isViolate = MR().assertViolation(original_output, program_output)
    #     if isViolate:
    #         result[0] = 1
    #
    #     for j in range(len(mr_list)):
    #         ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}_{}.txt".format(0, i, j))
    #         original_output = MR().getResults(ts)
    #         program_ts = ts
    #         program_ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}_{}.txt".format(mu, i, j))
    #         program_output = MR().getResults(program_ts)
    #         isViolate = MR().assertViolation(original_output, program_output)
    #         if isViolate:
    #             result[j + 1] = 1
    #
    #     for m in range(len(mr_list)):
    #         for n in range(len(mr_list)):
    #             ts.setInputOutput("input{}_{}_{}.txt".format(i, m, n),
    #                                        "output{}_{}_{}_{}.txt".format(0, i, m, n))
    #             original_output = MR().getResults(ts)
    #             program_ts = ts
    #             program_ts.setInputOutput("input{}_{}_{}.txt".format(i, m, n), "output{}_{}_{}_{}.txt".format(mu, i, m, n))
    #             program_output = MR().getResults(program_ts)
    #             isViolate = MR().assertViolation(original_output, program_output)
    #             if isViolate:
    #                 result[len(mr_list) + 1 + m * len(mr_list) + n] = 1
    #     pf.append(result)
    pf = []
    for i in range(num_of_samples):
        result1 = [0] * (len(mr_list) * (len(mr_list) + 1) + 1)  # t1,t2,...,t13
        ts.setInputOutput("input{}".format(i), "output{}_{}".format(0, i))
        original_output = MR().getResults(ts, outputdata)
        program_ts = copy.copy(ts)
        program_ts.setInputOutput("input{}".format(i), "output{}_{}".format(mu, i))
        program_output = MR().getResults(program_ts, outputdata)
        isViolate = MR().assertViolation(original_output, program_output)
        if isViolate:
            result1[0] = 1

        for j in range(len(mr_list)):
            ts.setInputOutput("input{}_{}".format(i, j), "output{}_{}_{}".format(0, i, j))
            original_output = MR().getResults(ts, outputdata)
            program_ts = copy.copy(ts)
            program_ts.setInputOutput("input{}_{}".format(i, j), "output{}_{}_{}".format(mu, i, j))
            program_output = MR().getResults(program_ts, outputdata)
            isViolate = MR().assertViolation(original_output, program_output)
            if isViolate:
                result1[j + 1] = 1

        for m in range(len(mr_list)):
            for n in range(len(mr_list)):
                ts.setInputOutput("input{}_{}_{}".format(i, m, n),
                                  "output{}_{}_{}_{}".format(0, i, m, n))
                original_output = MR().getResults(ts, outputdata)
                program_ts = copy.copy(ts)
                program_ts.setInputOutput("input{}_{}_{}".format(i, m, n), "output{}_{}_{}_{}".format(mu, i, m, n))
                program_output = MR().getResults(program_ts, outputdata)
                isViolate = MR().assertViolation(original_output, program_output)
                if isViolate:
                    result1[len(mr_list) + 1 + m * len(mr_list) + n] = 1

        # t14,t15,...
        result2 = []
        for j in range(len(mr_list)):
            r1 = []
            for k in range(len(mr_list)):
                r2 = []
                ts.setInputOutput("input{}_{}_{}".format(i, j, k),
                                  "output{}_{}_{}_{}".format(0, i, j, k))
                original_output = MR().getResults(ts, outputdata)
                program_ts = copy.copy(ts)
                program_ts.setInputOutput("input{}_{}_{}".format(i, j, k), "output{}_{}_{}_{}".format(mu, i, j, k))
                program_output = MR().getResults(program_ts, outputdata)
                isViolate = MR().assertViolation(original_output, program_output)
                if isViolate:
                    result = 1
                else:
                    result = 0
                r2.append(result)
                for m in range(len(mr_list)):
                    ts.setInputOutput("input{}_{}_{}_{}".format(i, j, k, m),
                                      "output{}_{}_{}_{}_{}".format(0, i, j, k, m))
                    original_output = MR().getResults(ts, outputdata)
                    program_ts = copy.copy(ts)
                    program_ts.setInputOutput("input{}_{}_{}_{}".format(i, j, k, m),
                                              "output{}_{}_{}_{}_{}".format(mu, i, j, k, m))
                    program_output = MR().getResults(program_ts, outputdata)
                    isViolate = MR().assertViolation(original_output, program_output)
                    if isViolate:
                        result = 1
                    else:
                        result = 0
                    r2.append(result)
                for m in range(len(mr_list)):
                    for n in range(len(mr_list)):
                        ts.setInputOutput("input{}_{}_{}_{}_{}".format(i, j, k, m, n),
                                          "output{}_{}_{}_{}_{}_{}".format(0, i, j, k, m, n))
                        original_output = MR().getResults(ts, outputdata)
                        program_ts = copy.copy(ts)
                        program_ts.setInputOutput("input{}_{}_{}_{}_{}".format(i, j, k, m, n),
                                                  "output{}_{}_{}_{}_{}_{}".format(mu, i, j, k, m, n))
                        program_output = MR().getResults(program_ts, outputdata)
                        isViolate = MR().assertViolation(original_output, program_output)
                        if isViolate:
                            result = 1
                        else:
                            result = 0
                        r2.append(result)
                r1.append(r2)
            result2.append(r1)
        pf.append([result1, result2])

    return pf


def get_input():
    inputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/PT/RandomInput.csv'
    # inputdir = '/home/rdx/data/MT/STVR/PT/RandomInput.csv'
    # 建立空字典
    csvFile = open(inputdir, "r")
    # reader = csv.reader(csvFile)
    reader = csv.reader((line.replace('\0', '') for line in csvFile))
    # 建立空字典
    inputcase = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            inputcase[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()
    return inputcase


def dataTrans(mr, num_of_samples):
    inputs_write = '/Users/rendaixu/OneDrive/data/MT/STVR/PT/RandomInput.csv'
    # outputs_write = '/Users/rendaixu/OneDrive/data/MT/FS/PT/RandomOutput'
    inputs_read = '/Applications/work/data/MT/MFT/PT/RandomInput'
    # outputs_read = '/Applications/work/data/MT/FS/PT/RandomOutput'
    fileHeader = ["name", "value"]
    inputdata = [fileHeader]

    for i in range(num_of_samples):
        f = open(inputs_read+'/input'+str(i)+'.txt')
        data = ['input{}'.format(i), f.read()]
        inputdata.append(data)
        for j in range(mr):
            f = open(inputs_read + '/input{}_{}.txt'.format(i, j))
            data = ['input{}_{}'.format(i, j), f.read()]
            inputdata.append(data)
            for k in range(mr):
                f = open(inputs_read + '/input{}_{}_{}.txt'.format(i, j, k))
                data = ['input{}_{}_{}'.format(i, j, k), f.read()]
                inputdata.append(data)
    # 将数据存下来
    with open(inputs_write, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(inputdata)


def verify(mr, num_of_samples):
    inputs_read = '/Applications/work/data/MT/MFT/PT/RandomOutput'
    fileHeader = ["name", "value"]
    inputdata = [fileHeader]
    for i in range(num_of_samples):
        f = open(inputs_read+'/output0_'+str(i)+'.txt')
        data = ['output0_{}'.format(i), f.read()]
        inputdata.append(data)
        for j in range(mr):
            f = open(inputs_read + '/output0_{}_{}.txt'.format(i, j))
            data = ['output0_{}_{}'.format(i, j), f.read()]
            inputdata.append(data)
            for k in range(mr):
                f = open(inputs_read + '/output0_{}_{}_{}.txt'.format(i, j, k))
                data = ['output0_{}_{}_{}'.format(i, j, k), f.read()]
                inputdata.append(data)
    inputcase = get_input()
    lst1 = list(inputcase.items())
    lst2 = list(inputcase.items())
    for i in range(len(lst1)):
        if not lst1[i][1] == lst2[i][1]:
            print(i)


def getOutput(mu, inputcase):
    fileHeader = ["name", "value"]
    outputdata = [fileHeader]
    command = './Mutants/printtokens_v{}/print_tokens'.format(mu)
    lst = list(inputcase.items())
    for i in range(len(lst)):
        # 将输入字符串写入临时文件
        try:
            output = subprocess.run(command, input=lst[i][1].encode(), stdout=subprocess.PIPE,
                                    stderr=subprocess.STDOUT, shell=True).stdout.decode()
        except UnicodeDecodeError:
            print("Mutant{} 编码格式问题，跳过".format(mu))
            return 0
        data = ['output'+lst[i][0][5:], output]
        outputdata.append(data)
    outputs_write = '/Users/rendaixu/OneDrive/data/MT/STVR/PT/RandomOutput{}.csv'.format(mu)
    # outputs_write = '/home/rdx/data/MT/STVR/PT/RandomOutput{}.csv'.format(mu)
    with open(outputs_write, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(outputdata)
    print('第{}个已完成'.format(mu))


def mutaterate(mu):
    outputdir1 = '/Users/rendaixu/OneDrive/data/MT/STVR/PT/RandomOutput0.csv'
    outputdir2 = '/Users/rendaixu/OneDrive/data/MT/STVR/PT/RandomOutput{}.csv'.format(mu)
    csvFile = open(outputdir1, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    outputcase1 = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase1[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    csvFile = open(outputdir2, "r")
    reader = csv.reader(csvFile)
    outputcase2 = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase2[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    num = 0
    outputcase1 = list(outputcase1.values())
    outputcase2 = list(outputcase2.values())
    index = []
    for i in range(len(outputcase1)):
        if not outputcase1[i] == outputcase2[i]:
            num += 1
            index.append(i)
    rate = num / (len(outputcase1))
    print(mu, round(rate*100, 2))
    return index


def run_getOutput(i, inputcase, timeout):
    p = multiprocessing.Process(target=getOutput, args=(i, inputcase))
    p.start()
    p.join(timeout)
    if p.is_alive():
        p.terminate()
        print("Mutant{} 超时，跳过".format(i))


def rank_of_element(lst, element):
    sorted_list = sorted(lst, reverse=True)
    return sorted_list.index(element) + 1


if __name__ == "__main__":
    num_case = 100
    num_mr = 11
    num_mu = 17
    # dataTrans(mr, num_of_samples)
    # inputcase = get_input()
    # timeout = 600  # 设定阈值为**秒
    # for i in range(38, 39):
    #     run_getOutput(i, inputcase, timeout)
    # for mu in range(num_mu+1):
    #     index = mutaterate(mu)
    #
    # outputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/PT/RandomOutput0.csv'
    # csvFile = open(outputdir, "r")
    # reader = csv.reader(csvFile)
    # # # 建立空字典
    # originaloutput = {}
    # for item in reader:
    #     # 忽略第一行
    #     if reader.line_num == 1:
    #         continue
    #     try:
    #         originaloutput[item[0]] = item[1].split("\n")
    #     except:
    #         print(item)
    #         print(reader.line_num)
    #         break
    # csvFile.close()
    # mr_list = [MR1(), MR2(), MR3(), MR4(), MR5(), MR6(), MR7(), MR8(), MR9(), MR10(), MR11()]
    # ts = TestCase()
    # MGS_set = []
    # SMGS_set = []
    # for mu in range(1, 24):
    #     MGS, SMGS = getMG(mu, mr_list, num_of_samples, originaloutput, ts)
    #     MGS_set.append(MGS)
    #     SMGS_set.append(SMGS)

    # verify(mr, num_of_samples)
    # # myenv = MyEnv()
    # # myenv.CreateWorkingDirs()
    project = 'STVR'
    # project2 = 'FS'
    row = 1
    string = 'PT'
    path = '/Applications/work/data/MT/' + project + '/Result/result25_one.xlsx'  # '+sys.argv[1][:-1]+'
    wb = load_workbook(path)
    if string not in wb.sheetnames:
        ws = wb.create_sheet(string)
    del wb[string]
    ws = wb.create_sheet(string)
    MG_set = []
    Sus_set = []
    Metric_set = []
    Union_set = []
    Unique_set = []
    Flag_set = []
    # inters = set(ExelinesS[0][0])
    # # 逐个计算交集
    # for i in range(len(ExelinesS)):
    #     for j in range(len(ExelinesS[0])):
    #         inters = inters.intersection(set(ExelinesS[i][j]))
    # # 将结果转换为列表
    # inters = list(inters)
    # inters.sort()
    removemu = [3, 4, 5, 6, 7, 8, 9, 10, 12, 13, 14, 15, 16, 17]
    # for mu in range(1, num_mu+1):  # 18
    #     if mu == 5:
    #         continue
    #     with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
    #               'r') as load_f:
    #         data = json.load(load_f)
    #     data = json.loads(data)
    #     Flag = data['Flag']
    #     fault = Executable[Flag.index(1)]
    #     if fault in inters:
    #         removemu.append(mu)
    # for i in range(len(ExelinesS)):
    #     for j in range(len(ExelinesS[0])):
    #         ExelinesS[i][j] = [x for x in ExelinesS[i][j] if x not in inters]
    # removemg = []
    # for i in range(len(ExelinesS)):
    #     a = 0
    #     for j in range(len(ExelinesS[i])):
    #         if len(ExelinesS[i][j]) == 0:
    #             a += 1
    #     if a == len(ExelinesS[i]):
    #        removemg.append(i)
    # for index in sorted(removemg, reverse=True):
    #     del ExelinesS[index]
    for mu in range(1, num_mu+1):  # 18
        if mu in removemu:
            continue
        # datadir = '/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/statementResult{}.csv'.format(mu)
        # csvFile = open(datadir, "r")
        # reader = csv.reader(csvFile)
        # # 建立空字典
        # originaldata = {}
        # result = {}
        # for item in reader:
        #     originaldata[item[0]] = item[1:-1]
        #     if reader.line_num == 1:
        #         continue
        #     result[item[0]] = item[-1:]
        # csvFile.close()
        # ExecutableS = originaldata.get('inputs')
        # ExecutableS = [int(x) for x in ExecutableS]
        # ExelineS = []
        # Result = []
        # for i in range(num_case):
        #     Exelines = []
        #     r = []
        #     candidate = originaldata.get('input{}'.format(i))
        #     candidate = [int(x) for x in candidate]
        #     indices = [i for i in range(len(candidate)) if candidate[i] == 1]
        #     Exelines.append([ExecutableS[i] for i in indices])
        #     a = result.get('input{}'.format(i))
        #     a = [int(x) for x in a]
        #     r.append(a)
        #     for j in range(num_mr):
        #         candidate = originaldata.get('input{}_{}'.format(i, j))
        #         candidate = [int(x) for x in candidate]
        #         indices = [i for i in range(len(candidate)) if candidate[i] == 1]
        #         Exelines.append([ExecutableS[i] for i in indices])
        #         a = result.get('input{}_{}'.format(i, j))
        #         a = [int(x) for x in a]
        #         r.append(a)
        #     for m in range(num_mr):
        #         for n in range(num_mr):
        #             candidate = originaldata.get('input{}_{}_{}'.format(i, m, n))
        #             candidate = [int(x) for x in candidate]
        #             indices = [i for i in range(len(candidate)) if candidate[i] == 1]
        #             Exelines.append([ExecutableS[i] for i in indices])
        #             a = result.get('input{}_{}_{}'.format(i, m, n))
        #             a = [int(x) for x in a]
        #             r.append(a)
        #     Result.append(r)
        #     ExelineS.append(Exelines)
        # data = {
        #     'Exec': ExecutableS, 'Exel': ExelineS, 'Result': Result
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/statements{}.json'.format(mu), 'w') as f:
        #     json.dump(json_str, f)
        with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/statements{}.json'.format(mu), 'r') as f:
            data = json.load(f)
        data = json.loads(data)
        ExelineS = data['Exel']
        Executable = data['Exec']
        # Result = data['Result']
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                  'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        MGS = data['MGS']
        # ExecutableS = data['Exec']
        # ExelinesS = data['Exel']
        Flag = data['Flag']
        # MGS = getMG(mu, mr_list, num_case, originaloutput, ts)
        # MG_set.append(SMGS)
        # ind = []
        # for i in range(len(SMGS)):
        #     if 1 in SMGS[i][0]:
        #         ind.append(i)
        # Ind.append(ind)

        # FaSus, percent = FAILTIMSlice(SMGS, ExecutableS, ExelinesS, Flag)
        # mSus = data['Sus']
        # MG_set.append(SMGS)
        # MG_set2.append(SMGS2)
        # Flag = [0] * len(Executable)
        # if mu == 1:
        #     Flag[161] = 1
        # elif mu == 2:
        #     Flag[162] = 1
        # elif mu == 3:
        #     Flag[67] = 1
        # elif mu == 4:
        #     Flag[90] = 1
        # elif mu == 5:
        #     Flag[161] = 1
        # elif mu == 6:
        #     Flag[91] = 1
        # elif mu == 7:
        #     Flag[161] = 1
        # elif mu == 8:
        #     Flag[202] = 1
        # elif mu == 9:
        #     Flag[207] = 1
        # elif mu == 10:
        #     Flag[207] = 1
        # elif mu == 11:
        #     Flag[46] = 1
        # elif mu == 12:
        #     Flag[204] = 1
        # elif mu == 13:
        #     Flag[46] = 1
        # elif mu == 14:
        #     Flag[88] = 1
        # elif mu == 15:
        #     Flag[148] = 1
        # elif mu == 16:
        #     Flag[155] = 1
        # elif mu == 17:
        #     Flag[202] = 1
        # data['Flag'] = Flag
        # mSus, metric = MSlice(MGS, Executable, ExelineS)
        # mSus2, AllF, metric2 = MSlice2(MGS, ExecutableS, ExelinesS, Flag)
    #     with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
    #               'r') as load_f:
    #         data = json.load(load_f)
    #     data = json.loads(data)
    #     mSus = data['mSus']
    #     # mSus2 = data['mSus2']
    #     # mSus3 = data['mSus3']
    #     mSus4 = data['mSus4']
    #     # metric = data['metric']
    #     # metric2 = data['metric2']
    #     # metric3 = data['metric3']
    #     # metric4 = data['metric4']
    #     Flag = data['Flag']
    #     MGS = data['MGS']
        Result = data['result']
        mSus4 = data['mSus4_one']
        mSus = data['mSus_one']
        Para = data['Para']
        # AllF = data['AllF']
        # Para = data['Para']
        # mSus4_all = data['mSus4_all']
        # mSus_all = data['mSus_all']
        # mSus4_one = data['mSus4_one']
        # mSus_one = data['mSus_one']
    #     mSus4_nofs = data['mSus4_nofs']
    #     mSus_nofs = data['mSus_nofs']
    #     # Sus = []
    #     # Metric = []
    #     # for i in range(len(mSus)):
    #     #     a = [round(mSus[i], 4)]
    #     #     b = [metric[i]]
    #     #     for j in range(len(mSus3)):
    #     #         a.append(mSus3[j][5][i])
    #     #         b.append(metric3[j][i])
    #     #     Sus.append(a)
    #     #     Metric.append(b)
    #     # Sus_set.append(Sus)
    #     # Metric_set.append(Metric)
    #     mSus_nofs, metric_nofs = MSlice(MGS, Executable, ExelineS)
    #     data['mSus_nofs'] = mSus_nofs
    #     data['metric_nofs'] = metric_nofs
    #     # mSus3, AllF, _ = MSlice3(MGS, Executable, ExelinesS, Flag)
    #     # Union_set.append(Union)
    #     # Unique_set.append(Unique)
    #     result = []
    #     for i in range(len(Result)):
    #         a = []
    #         for j in range(len(Result[i])):
    #             a.append(Result[i][j][0])
    #         result.append(a)
    #     data['result'] = result
    #     mSus4_nofs, Para, metric4_nofs = MSlice4(MGS, Executable, ExelineS, Flag, result)
        # data['Para'] = Para
        # data['mSus4_nofs'] = mSus4_nofs
        # data['metric4_nofs'] = metric4_nofs
        # mSus, metric = MSlice_one(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus_one'] = mSus
        # data['metric_one'] = metric
        # mSus, metric = MSlice_all(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus_all'] = mSus
        # data['metric_all'] = metric
        # mSus4, _, metric4 = MSlice4_one(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus4_one'] = mSus4
        # data['metric4_one'] = metric4
        # data['Para'] = Para
        # mSus4, _, metric4 = MSlice4_all(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus4_all'] = mSus4
        # data['metric4_all'] = metric4
    #     # data = {
    #     #          'mSus': mSus, 'mSus4': mSus4, 'metric': metric,
    #     #     'metric4': metric4, 'Para': Para, 'MGS': MGS, 'Flag': Flag,
    #     # }
    #     json_str = json.dumps(data)
    #     with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
    #               'w') as f:
    #         json.dump(json_str, f)
        row = eval('getMetrics_5')(row, ws, mu, MGS, mSus, mSus4, Para[-1], Flag, Para, ExelineS)
    wb.save(path)

