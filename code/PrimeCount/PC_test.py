import subprocess
from openpyxl import load_workbook
from publicFun import *
import csv
import json
import os

# 自动切换工作目录为脚本所在目录
os.chdir(os.path.dirname(os.path.abspath(__file__)))

def run_primecount(mu, x: int) -> int:
    result = subprocess.run(
        ['Mutants/PrimeCount_v{}/PrimeCount'.format(mu), str(x)],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True  # 以字符串方式读取输出
    )

    if result.returncode != 0:
        raise RuntimeError(f"运行失败：{result.stderr.strip()}")

    output = result.stdout.strip()
    if '=' in output:
        y_str = output.split('=')[-1].strip()
        return int(y_str)
    else:
        raise ValueError(f"无法解析输出：{output}")

# project = 'EMS'
# string = 'PrimeCount'
# path = '../../data/MT/'+project+'/' + string + '/followup'
# input_path = path + "input"
# inputs = []
# with open(input_path, "r") as fin:
#     for line in fin:
#         line = line.strip()
#         if not line:
#             continue
#         parts = line.split()
#         if len(parts) != 3:
#             print(f"格式错误：{line}")
#             continue
#         _, id, x_str = parts
#         x = int(x_str)
#         inputs.append((line, x))  # 原始行 + 解析后的x值
#
# # Step 2: 对每个 mutant 分别运行
# for mu in range(30):
#     output_path = path + f"output_v{mu}"
#     with open(output_path, "w") as fout:
#         for line, x in inputs:
#             try:
#                 y = run_primecount(mu, x)
#                 fout.write(f"{y}\n")
#             except Exception as e:
#                 fout.write(f"{line} => 错误: {e}\n")

if __name__ == "__main__":
    num_case = 100
    num_mr = 3
    num_mu = 17
    project = 'EMS'
    string = 'PrimeCount'
    # title = [30, 31, 32, 33]
    # N = [1, 3, 5, 10]
    for i in range(1):
        row = 1
        path = '../../data/results/resultEXAM.xlsx'  # '+sys.argv[1][:-1]+'
        wb = load_workbook(path)
        if string in wb.sheetnames:
            del wb[string]
        ws = wb.create_sheet(string)
        # sheet = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'dice',
        #            'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
        #            'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3',
        #            'Arithmetic Mean', 'Cohen', 'Fleiss', 'Dstar', 'GP13', 'size', 'Dice', 'grep2', 'KNN']
        # for f in sheet:
        #     if f in wb.sheetnames:
        #         del wb[f]
        for mu in range(1, num_mu+1):
            # datadir = '../../data/MT/' + project + '/' + string + '/statementsource_v{}.csv'.format(mu)
            # csvFile = open(datadir, "r")
            # reader = csv.reader(csvFile)
            # # 建立空字典
            # sourcecov = {}
            # sourceresult = {}
            # for item in reader:
            #     sourcecov[item[0]] = item[1:-1]
            #     if reader.line_num == 1:
            #         continue
            #     sourceresult[item[0]] = item[-1:]
            # csvFile.close()
            # datadir = '../../data/MT/' + project + '/' + string + '/statementfollow_v{}.csv'.format(mu)
            # csvFile = open(datadir, "r")
            # reader = csv.reader(csvFile)
            # # 建立空字典
            # followcov = {}
            # followresult = {}
            # for item in reader:
            #     followcov[item[0]] = item[1:-1]
            #     if reader.line_num == 1:
            #         continue
            #     followresult[item[0]] = item[-1:]
            # csvFile.close()
            # ExecutableS = sourcecov.get('inputs')
            # ExecutableS = [int(x) for x in ExecutableS]
            # ExelineS = []
            # Result = []
            # for i in range(num_case):
            #     Exelines = []
            #     r = []
            #     candidate = sourcecov.get('input{}'.format(i))
            #     candidate = [int(x) for x in candidate]
            #     indices = [i for i in range(len(candidate)) if candidate[i] == 1]
            #     e = [ExecutableS[i] for i in indices]
            #     Exelines.append([e])
            #     a = sourceresult.get('input{}'.format(i))
            #     a = [int(x) for x in a]
            #     r.append([a[0]])
            #     for j in range(num_mr):
            #         if j < 2:
            #             candidate = followcov.get('input{}'.format(i*4+j))
            #             candidate = [int(x) for x in candidate]
            #             indices = [i for i in range(len(candidate)) if candidate[i] == 1]
            #             e = [ExecutableS[i] for i in indices]
            #             Exelines.append([e])
            #             a = followresult.get('input{}'.format(i*4+j))
            #             a = [int(x) for x in a]
            #             r.append([a[0]])
            #         else:
            #             # MR3包含两个衍生测试用例
            #             candidate = followcov.get('input{}'.format(i*4+j))
            #             candidate = [int(x) for x in candidate]
            #             indices = [i for i in range(len(candidate)) if candidate[i] == 1]
            #             e1 = [ExecutableS[i] for i in indices]
            #             a = followresult.get('input{}'.format(i*4+j))
            #             a1 = [int(x) for x in a]
            #             candidate = followcov.get('input{}'.format(i*4+j+1))
            #             candidate = [int(x) for x in candidate]
            #             indices = [i for i in range(len(candidate)) if candidate[i] == 1]
            #             e2 = [ExecutableS[i] for i in indices]
            #             Exelines.append([e1, e2])
            #             a = followresult.get('input{}'.format(i*4+j+1))
            #             a2 = [int(x) for x in a]
            #             r.append([a1[0],a2[0]])
            #     Result.append(r)
            #     ExelineS.append(Exelines)
            # data = {
            #     'Exec': ExecutableS, 'Exel': ExelineS, 'Result': Result
            # }
            # json_str = json.dumps(data)
            # with open('../../data/MT/' + project + '/' + string + '/statements{}.json'.format(mu), 'w') as f:
            #     json.dump(json_str, f)
            # with open('../../data/MT/'+project+'/' + string + '/statements{}.json'.format(mu), 'r') as f:
            #     data = json.load(f)
            # data = json.loads(data)
            # ExelineS = data['Exel']
            # ExecutableS = data['Exec']
            # Result = data['Result']
            with open('../../data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                      'r') as load_f:
                data = json.load(load_f)
            data = json.loads(data)
            # MGS = data['MGS']
            # ExecutableS = data['Exec']
            # ExelinesS = data['Exel']
            # FaSus, percent = FAILTIMSlice(SMGS, ExecutableS, ExelinesS, Flag)
            # mSus = data['Sus']
            # Flag = [0] * len(ExecutableS)
            # raw_indices = {1: 615,2: 1277,3: 1284,4: 1383,5: 1384,6: 1384,7: 1384,8: 1384,9: 1384,10: 1384,
            #                     11: 1384,12: 1384,13: 1451,14: 456,15: 591,16: 1260, 17: 1451}
            # mu_to_flag_index = {k: ExecutableS.index(v) for k, v in raw_indices.items()}
            # if mu in mu_to_flag_index:
            #     Flag[mu_to_flag_index[mu]] = 1
            # data['Flag'] = Flag
            # mSus, metric = MSlice(MGS, Executable, ExelineS)
            # mSus2, AllF, metric2 = MSlice2(MGS, ExecutableS, ExelinesS, Flag)
        #     with open('../../data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #               'r') as load_f:
        #         data = json.load(load_f)
        #     data = json.loads(data)
        #     mSus = data['mSus']
        #     # mSus2 = data['mSus2']
        #     # mSus3 = data['mSus3']
        #     mSus4 = data['mSus4']
        #     # metric = data['metric']
        #     # metric2 = data['metric2']
        #     # metric3 = data['metric3']
        #     # metric4 = data['metric4']
            Flag = data['Flag']
            # MGS = data['MGS']
            # Result = data['result']
            # mSus4 = data['mSus4_nofs']
            # mSus = data['mSus_nofs']
            sbflsus = data['sbflsus']
            # Para = data['Para']
            # if i == 0:
            #     mSus, metric = MSlice_PC(MGS, ExecutableS, ExelineS)
            #     mSus4, _, metric4 = MSlice4_PC_all(MGS, ExecutableS, ExelineS, Flag, Result)
            #     # sbflsus, sbflmetric = SBFL(MGS, Result, Executable, ExelineS)
            #     data['mSus_all'] = mSus
            #     data['metric_all'] = metric
            #     data['mSus4_all'] = mSus4
            #     data['metric4_all'] = metric4
            #     # data['sbflsus'] = sbflsus
            #     # data['sbflmetric'] = sbflmetric
            #     # Union_set.append(Union)
            #     # Unique_set.append(Unique)
            #     # mSus4, _, metric4 = MSlice4_one(MGS, Executable, ExelineS, Flag, Result)
            #     # data['mSus4_one'] = mSus4
            #     # data['metric4_one'] = metric4
            #     # data['Para'] = Para
            #     # mSus4, _, metric4 = MSlice4_all(MGS, Executable, ExelineS, Flag, Result)
            #     # data['mSus4_all'] = mSus4
            #     # data['metric4_all'] = metric4
            #     # data = {
            #     #          'mSus': mSus, 'mSus4': mSus4, 'metric': metric,
            #     #     'metric4': metric4, 'Para': Para, 'MGS': MGS, 'Flag': Flag,
            #     # }
            #     json_str = json.dumps(data)
            #     with open('../../data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
            #               'w') as f:
            #         json.dump(json_str, f)
            # else:
            mSus4 = data['mSus4']
            mSus = data['mSus']
            # row = eval('getMetrics_5')(row, ws, mu, MGS, mSus, mSus4,sbflsus, Para[-1], Flag, Para, ExelineS, N[i])
            row = eval('getMetrics_6')(row, ws, mu, mSus, mSus4, sbflsus, Flag)
        wb.save(path)

