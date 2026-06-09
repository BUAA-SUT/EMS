from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
# import matplotlib.pyplot as plt
# from matplotlib import ticker
import scipy.stats as stats
from statsmodels.stats.multitest import multipletests
from publicFun import *
import json


def are_sublists_identical(lst):
    return all(x == lst[0] for x in lst)


def min_max_normalize(data):
    min_val = min(data)
    max_val = max(data)
    normalized_data = [(x - min_val) / (max_val - min_val) for x in data]
    return normalized_data


def find_max_min(dictionary):
    """
    字典
    """
    if len(dictionary) == 0:
        return None, None

    max_key = max(dictionary, key=dictionary.get)
    min_key = min(dictionary, key=dictionary.get)
    max_value = dictionary[max_key]
    min_value = dictionary[min_key]

    return max_key, max_value, min_key, min_value

def findmaxmin(lst):
    """
    列表
    """
    if len(lst) == 0:
        return None, None  # 如果列表为空，返回None
    lst[0] = lst[4] = max(lst)-0.1*max(lst)  # 为了忽略naish1和binary
    max_value = max(lst)  # 获取最大值
    min_value = min(lst)  # 获取最小值

    max_indices = [i for i, x in enumerate(lst) if x == max_value]  # 找出所有最大值的索引
    min_indices = [i for i, x in enumerate(lst) if x == min_value]  # 找出所有最小值的索引

    return max_indices, min_indices


def merge_keys_with_equal_values(d):
    # 创建一个新的字典来存储合并后的结果
    merged_dict = {}

    # 遍历字典中的键和值
    for key, value in d.items():
        # 查找当前值是否已经存在于新的字典中
        found = False
        for merged_keys in list(merged_dict.keys()):
            if merged_dict[merged_keys] == value:
                # 将 tuple 转换为 list，追加键，然后重新转换为 tuple
                new_keys = list(merged_keys) + [key]
                merged_dict.pop(merged_keys)  # 删除旧的键
                merged_dict[tuple(new_keys)] = value  # 添加更新后的键值对
                found = True
                break
        # 如果值还不存在，创建新的键列表
        if not found:
            merged_dict[tuple([key])] = value

    return merged_dict


def find_min_max_indices_in_dict(d):
    min_value = float('inf')  # 初始设置一个非常大的最小值
    max_value = float('-inf')  # 初始设置一个非常小的最大值
    min_key = None  # 存储最小值所在的键
    min_index = None  # 存储最小值在列表中的索引
    max_key = None  # 存储最大值所在的键
    max_index = None  # 存储最大值在列表中的索引

    # 遍历字典
    for key, value_list in d.items():
        # 遍历每个键对应的列表
        for index, value in enumerate(value_list):
            if value < min_value:
                min_value = value
                min_key = key
                min_index = index
            if value > max_value:
                max_value = value
                max_key = key
                max_index = index

    return min_key, min_index, max_key, max_index


# Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'dice',
#            'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
#            'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3',
#            'Arithmetic Mean', 'Cohen', 'Fleiss', 'Dstar', 'GP13']
#
# project = 'STVR'
# path = '../../data/results/result21.xlsx'  # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# program = ['TSQ', 'DM', 'Tcas', 'SMM', 'grep', 'PT', 'PT2']  # 'TSQ', 'DM', 'Tcas', 'SMM', 'KNN', 'grep', 'PT', 'PT2'
# sheet = Formula
# for s in sheet:
#     if s == 'dice':
#         s = 'Dice'
#     Value = []
#     datadist = {}
#     title = ['oms']
#     t = 1
#     for i in range(25):
#         title.append("i1d{}".format(100-t*2))
#         t += 1
#     t = 1
#     for i in range(25):
#         title.append("i{}d1".format(100-t*2))
#         t += 1
#     # title.append('ensemble1')
#     # title.append('ensemble2')
#     # title.append('imse1')
#     # title.append('imse2')
#     # title.append('dmse1')
#     # title.append('dmse2')
#     title.append('SMG(%)')
#     title.append('FS(%)')
#     title.append('AllFofV(%)')
#     title.append('union')
#     title.append('unique')
#     title.append('unionvmg')
#     title.append('uniquevmg')
#     title.append('unionsmg')
#     title.append('uniquesmg')
#     tablelist = {"Mutants": title}
#     # title = ['MS', 'MS1', 'MS2', 'MS3', 'MS4', 'MS5', 'MS6', 'MS7', 'MS8', 'MS9', 'MS10',
#     #                                   'MaximalM', 'MaximalM1', 'MaximalM2', 'MaximalM3', 'MaximalM4', 'MaximalM5',
#     #                                   'MaximalM6', 'MaximalM7', 'MaximalM8', 'MaximalM9', 'MaximalM10',
#     #                                   'SMG(%)', 'FS(%)', 'AllFofV(%)']
#     # tablelist = {"Mutants": title}
#     datadist.update(tablelist)
#     for k in program:
#         ws = wb[k]
#         max_row = ws.max_row
#         t = 0
#         n = 2
#         x = 2
#         while True:
#             # if ws.cell(x, 59).value == 100 or ws.cell(x, 59).value == 0:
#             #     x += 1
#             #     n += 1
#             #     if n == 32:
#             #         n = 2
#             #         x += 3
#             #     if x > max_row:
#             #         break
#             #     continue
#             a = []
#             for i in range(2, len(title)+2):
#                 if ws.cell(x, 1).value not in [s]:
#                     break
#                 value = ws.cell(x, i).value
#                 a.append(value)
#             Value.append(a)
#             x += 1
#             n += 1
#             if n == 32:
#                 n = 2
#                 x += 3
#             if x > max_row:
#                 break
#     t = 0
#     x = 0
#     while True:
#         ave = []
#         x = t * 30
#         for i in range(len(title)):
#             n = 0
#             value = []
#             while True:
#                 if len(Value[x]) == 0:
#                     n += 1
#                     x += 1
#                     if n == 30:
#                         x = t * 30
#                         break
#                     continue
#                 value.append(Value[x][i])
#                 n += 1
#                 x += 1
#                 if n == 30:
#                     x = t * 30
#                     break
#             ave.append(value[0])
#         t += 1
#         # for k in range(11):
#         #     ave[k] = ave[k] - ave[k+11]
#         data = {
#             'Mutant'+str(t): ave
#         }
#         datadist.update(data)
#         if t >= len(Value) / 30:
#             break
#     string = s
#     if string not in wb.sheetnames:
#         ws = wb.create_sheet(string)
#     del wb[string]
#     ws = wb.create_sheet(string)
#     row = 1
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     wb.save(path)
#
# top-n
# RQ1
# project = 'EMS'
# Datadict = {"weights": ["TOP-1", "TOP-3", "TOP-5", "TOP-10"], "oms": [], "SBFL": []}
# t = 1
# for i in range(25):
#     Datadict["i1d{}".format(100 - t * 2)] = []
#     t += 1
# t = 1
# for i in range(25):
#     Datadict["i{}d1".format(100 - t * 2)] = []
#     t += 1
# for kk in range(26, 30):
#     path = '../../data/results/result{}.xlsx'.format(kk)  # '+sys.argv[1][:-1]+'
#     wb = load_workbook(path)
#     program = ['TSQ', 'DM', 'Tcas', 'SMM', 'grep', 'PT', 'PT2', 'PrimeCount', 'SeqMap']  # 'TSQ', 'DM', 'Tcas', 'SMM', 'grep', 'PT', 'PT2', 'PrimeCount', 'SeqMap'
#     sheet = program
#     datadict = {"oms": [], "SBFL": []}
#     t = 1
#     for i in range(25):
#         datadict["i1d{}".format(100-t*2)] = []
#         t += 1
#     t = 1
#     for i in range(25):
#         datadict["i{}d1".format(100-t*2)] = []
#         t += 1
#     for k in sheet:
#         ws = wb[k]
#         max_row = ws.max_row
#         max_col = ws.max_column
#         n = 2
#         x = 2
#         a = []
#         while True:
#             # if ws.cell(x, max_col-7).value == 0:
#             #     x += 33
#             #     continue
#             b = []
#             for j in range(2, 54): # oms+sbfl+50权重
#                 b.append(ws.cell(x, j).value)
#             a.append(b)
#             x += 1
#             n += 1
#             if n == 34: # 32个公式
#                 for q in range(len(a[0])):
#                     c = []
#                     aaa = [0, 4]
#                     for p in range(len(a)):
#                         if p in aaa:  # p == 0 or p == 4
#                             continue
#                         c.append(a[p][q])
#                     datadict[list(datadict.keys())[q]].append(np.mean(c))
#                 a = []
#                 n = 2
#                 x += 3
#             if x > max_row:
#                 break
#
#     for i in range(len(list(datadict.keys()))):
#         key = list(datadict.keys())[i]
#         if isinstance(datadict[key], list):  # 确保它是列表
#             Datadict[key].append(round(sum(datadict[key]), 2))
#         else:  # 如果不是列表，可能需要转换为列表
#             Datadict[key] = [datadict[key]]
#             Datadict[key].append(round(datadict[key][0], 2))
#         # datadict[list(datadict.keys())[i]] = round(sum(datadict[list(datadict.keys())[i]]), 2)

# path = '../../data/results/29.xlsx'  # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# string = "Sheet1"
# if string not in wb.sheetnames:
#     ws = wb.create_sheet(string)
# del wb[string]
# ws = wb.create_sheet(string)
# row = 1
# for i, j in Datadict.items():  # i--公式名称, j--指标值
#     ws.cell(row, 1).value = i  # 添加第 1 列的数据
#     for col in range(2, len(j) + 2):  # values列表中索引
#         ws.cell(row, col).value = j[col - 2]
#     row += 1  # 行数
# wb.save(path)


# EXAM
# RQ1
# project = 'EMS'
# Datadict = {"weights": ["EXAM"], "oms": [], "SBFL": []}
# t = 1
# for i in range(25):
#     Datadict["i1d{}".format(100 - t * 2)] = []
#     t += 1
# t = 1
# for i in range(25):
#     Datadict["i{}d1".format(100 - t * 2)] = []
#     t += 1
# path = '../../data/results/resultEXAM.xlsx' # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# program = ['TSQ', 'DM', 'Tcas', 'SMM', 'grep', 'PT', 'PT2', 'PrimeCount', 'SeqMap']  # 'TSQ', 'DM', 'Tcas', 'SMM', 'grep', 'PT', 'PT2', 'PrimeCount', 'SeqMap'
# sheet = program
# datadict = {"oms": [], "SBFL": []}
# t = 1
# for i in range(25):
#     datadict["i1d{}".format(100-t*2)] = []
#     t += 1
# t = 1
# for i in range(25):
#     datadict["i{}d1".format(100-t*2)] = []
#     t += 1
# for k in sheet:
#     ws = wb[k]
#     max_row = ws.max_row
#     max_col = ws.max_column
#     n = 2
#     x = 2
#     a = []
#     while True:
#         # if ws.cell(x, max_col-7).value == 0:
#         #     x += 33
#         #     continue
#         b = []
#         for j in range(2, 54): # oms+sbfl+50权重
#             b.append(ws.cell(x, j).value)
#         a.append(b)
#         x += 1
#         n += 1
#         if n == 34: # 32个公式
#             for q in range(len(a[0])):
#                 c = []
#                 aaa = [0, 4]
#                 for p in range(len(a)):
#                     if p in aaa:  # p == 0 or p == 4
#                         continue
#                     c.append(a[p][q])
#                 datadict[list(datadict.keys())[q]].append(np.mean(c))
#             a = []
#             n = 2
#             x += 3
#         if x > max_row:
#             break
#
# for i in range(len(list(datadict.keys()))):
#     key = list(datadict.keys())[i]
#     if isinstance(datadict[key], list):  # 确保它是列表
#         Datadict[key].append(round(sum(datadict[key])/len(datadict[key]), 2))
#     else:  # 如果不是列表，可能需要转换为列表
#         Datadict[key] = [datadict[key]]
#         Datadict[key].append(round(datadict[key][0], 2))
#     # datadict[list(datadict.keys())[i]] = round(sum(datadict[list(datadict.keys())[i]]), 2)
#
# path = '../../data/results/resultEXAM.xlsx'  # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# string = "Sheet1"
# if string not in wb.sheetnames:
#     ws = wb.create_sheet(string)
# del wb[string]
# ws = wb.create_sheet(string)
# row = 1
# for i, j in Datadict.items():  # i--公式名称, j--指标值
#     ws.cell(row, 1).value = i  # 添加第 1 列的数据
#     for col in range(2, len(j) + 2):  # values列表中索引
#         ws.cell(row, col).value = j[col - 2]
#     row += 1  # 行数
# wb.save(path)


# # 统计每个权重的最大最小公式，RQ2
# project = 'EMS'
# path = '../../data/results/result28.xlsx'  # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'dice',
#            'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
#            'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3',
#            'Arithmetic Mean', 'Cohen', 'Fleiss', 'Dstar', 'GP13']
# Labels = ['oms']
# t = 1
# for i in range(25):
#     Labels.append("i{}".format(100 - t * 2))
#     t += 1
# t = 1
# for i in range(25):
#     Labels.append("d{}".format(100 - t * 2))
#     t += 1
#
# statistics = {}
# ws = wb["Sum"]
# max_row = ws.max_row
# max_col = ws.max_column
# value = {}
# n = 2
# for i in range(2, len(Labels) + 3):
#     if i < 29:  # SBFL
#         continue
#     v = []
#     for c in range(2, max_row + 1):
#         v.append(ws.cell(c, i).value)
#     if i == 2:
#         value[Labels[i - 2]] = v
#     else:
#         value[Labels[i - 3]] = v
#
# # 对全部组合排序
# # 创建一个列表存储键、索引和值的组合
# result = []
# # 遍历字典，获取键和值
# for key, values in value.items():
#     # 遍历每个列表的索引和值
#     for idx, val in enumerate(values):
#         # 将键、索引和值组合为元组并添加到结果中
#         if idx == 0 or idx == 4:
#             continue
#         result.append((key, Formula[idx], val))
# # 按照值从大到小排序
# sorted_result = sorted(result, key=lambda x: x[2], reverse=True)
# max_val = sorted_result[0][2]
# min_val = sorted_result[-1][2]
# # 筛选出所有等于最大值和最小值的项
# max_items = [item for item in sorted_result if item[2] == max_val]
# min_items = [item for item in sorted_result if item[2] == min_val]
#
# # 统计每个权重的最大最小公式
# min_key, min_index, max_key, max_index = find_min_max_indices_in_dict(value)
# maxformula = Formula[max_index]
# minformula = Formula[min_index]
# # 遍历键和值
# for key, val in value.items():
#     max_index, min_index = findmaxmin(val)
#     statistics[key] = {"max": [Formula[i] for i in max_index], "min": [Formula[i] for i in min_index]}
# result = merge_keys_with_equal_values(statistics)


# Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'dice',
#            'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
#            'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3',
#            'Arithmetic Mean', 'Cohen', 'Fleiss', 'Dstar', 'GP13']
#
#假设检验-topn
V = []
V2 = []
for i in range(26, 30):
    project = 'EMS'
    path = '../../data/results/result{}.xlsx'.format(i)  # '+sys.argv[1][:-1]+'
    wb = load_workbook(path)
    program = ['TSQ', 'DM', 'Tcas', 'SMM', 'grep', 'PT', 'PT2', 'PrimeCount', 'SeqMap']  # 'TSQ', 'DM', 'Tcas', 'SMM', 'KNN', 'grep', 'PT', 'PT2'
    sheet = program
    zong = [[0] * 52 for _ in range(32)]
    for k in sheet:
        ws = wb[k]
        max_row = ws.max_row
        n = 2
        x = 2
        a = []
        while True:
            b = []
            for j in range(2, 54):
                b.append(ws.cell(x, j).value)
            # b.append(ws.cell(x, 55).value)
            a.append(b)
            x += 1
            n += 1
            if n == 34:
                for p in range(len(a)):
                    for q in range(len(a[0])):
                        zong[p][q] += a[p][q]
                a = []
                n = 2
                x += 3
            if x > max_row:
                break

    # string = "Sum"
    # if string not in wb.sheetnames:
    #     ws = wb.create_sheet(string)
    # del wb[string]
    # ws = wb.create_sheet(string)
    # datadict = {}
    # title = ['oms', 'SBFL']
    # t = 1
    # for i in range(25):
    #     title.append("i1d{}".format(100-t*2))
    #     t += 1
    # t = 1
    # for i in range(25):
    #     title.append("i{}d1".format(100-t*2))
    #     t += 1
    # title.append("FS")
    # tablelist = {"Sum": title}
    # datadict.update(tablelist)
    # for i in range(len(Formula)):
    #     data = {
    #         Formula[i]: zong[i]
    #     }
    #     datadict.update(data)
    # row = 1
    # for i, j in datadict.items():  # i--公式名称, j--指标值
    #     ws.cell(row, 1).value = i  # 添加第 1 列的数据
    #     for col in range(2, len(j) + 2):  # values列表中索引
    #         ws.cell(row, col).value = j[col - 2]
    #     row += 1  # 行数
    # wb.save(path)

    # 假设检验
    # 新权重和oms， mannwhitneyu
    vv = []
    x1 = [] #oms
    for i in range(len(zong)):
        if i == 0 or i == 4: # Naish1 和Binary
            continue
        x1.append(zong[i][0])
    for i in range(2, len(zong[0])):
        x2 = []
        for j in range(len(zong)):
            if j == 0 or j == 4:
                continue
            x2.append(zong[j][i])
        statistic, pvalue = stats.wilcoxon(x1, x2, alternative='greater')
        # 找到显著差异的位置
        # significant_positions = [i + 1 for i in range(len(pvalue)) if pvalue[i] < 0.05]
        # if len(significant_positions) == 0:
        #     significant_positions = [-1]
        # value = [statistic, pvalue, significant_positions[0]]
        vv.append(round(pvalue, 4))
    # 进行 Benjamini-Hochberg 校正
    p_adjusted = multipletests(vv, method='fdr_bh')[1]
    V2.append(np.round(p_adjusted, 4).tolist())
    V.append(vv)

#假设检验-exam
# project = 'EMS'
# path = '../../data/results/resultEXAM.xlsx' # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# program = ['TSQ', 'DM', 'Tcas', 'SMM', 'grep', 'PT', 'PT2', 'PrimeCount', 'SeqMap']  # 'TSQ', 'DM', 'Tcas', 'SMM', 'KNN', 'grep', 'PT', 'PT2'
# sheet = program
# zong = [[0] * 52 for _ in range(32)]
# mutant_num = 0
# for k in sheet:
#     ws = wb[k]
#     max_row = ws.max_row
#     n = 2
#     x = 2
#     a = []
#     while True:
#         b = []
#         for j in range(2, 54):
#             b.append(ws.cell(x, j).value)
#         # b.append(ws.cell(x, 55).value)
#         a.append(b)
#         x += 1
#         n += 1
#         if n == 34:
#             mutant_num += 1
#             for p in range(len(a)):
#                 for q in range(len(a[0])):
#                     zong[p][q] += a[p][q]
#             a = []
#             n = 2
#             x += 3
#         if x > max_row:
#             break
#
# for i in range(len(zong)):
#     for j in range(len(zong[0])):
#         zong[i][j] /= mutant_num
#
# # string = "Sum"
# # if string not in wb.sheetnames:
# #     ws = wb.create_sheet(string)
# # del wb[string]
# # ws = wb.create_sheet(string)
# # datadict = {}
# # title = ['oms', 'SBFL']
# # t = 1
# # for i in range(25):
# #     title.append("i1d{}".format(100-t*2))
# #     t += 1
# # t = 1
# # for i in range(25):
# #     title.append("i{}d1".format(100-t*2))
# #     t += 1
# # title.append("FS")
# # tablelist = {"Sum": title}
# # datadict.update(tablelist)
# # for i in range(len(Formula)):
# #     data = {
# #         Formula[i]: zong[i]
# #     }
# #     datadict.update(data)
# # row = 1
# # for i, j in datadict.items():  # i--公式名称, j--指标值
# #     ws.cell(row, 1).value = i  # 添加第 1 列的数据
# #     for col in range(2, len(j) + 2):  # values列表中索引
# #         ws.cell(row, col).value = j[col - 2]
# #     row += 1  # 行数
# # wb.save(path)
#
# # 假设检验
# # 新权重和oms， mannwhitneyu
# vv = []
# x1 = [] #oms
# for i in range(len(zong)):
#     if i == 0 or i == 4: # Naish1 和Binary
#         continue
#     x1.append(zong[i][0])
# for i in range(2, len(zong[0])):
#     x2 = []
#     for j in range(len(zong)):
#         if j == 0 or j == 4:
#             continue
#         x2.append(zong[j][i])
#     statistic, pvalue = stats.wilcoxon(x2, x1, alternative='less')
#     # 找到显著差异的位置
#     # significant_positions = [i + 1 for i in range(len(pvalue)) if pvalue[i] < 0.05]
#     # if len(significant_positions) == 0:
#     #     significant_positions = [-1]
#     # value = [statistic, pvalue, significant_positions[0]]
#     vv.append(round(pvalue, 4))
# p_adjusted = multipletests(vv, method='fdr_bh')[1]
# p = np.round(p_adjusted, 4).tolist()


# Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'dice',
#            'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
#            'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3',
#            'Arithmetic Mean', 'Cohen', 'Fleiss']
#
# project = 'STVR'
# path = '../../data/results/result18.xlsx'  # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# ws = wb["Wong3"]
# max_row = ws.max_row
# x = 2
# a = []
# while True:
#     a.append(ws.cell(x, 28).value)
#     x += 1
#     if x > max_row:
#         break
#
# ws = wb["Kulczynski2"]
# max_row = ws.max_row
# x = 2
# b = []
# while True:
#     b.append(ws.cell(x, 3).value)
#     x += 1
#     if x > max_row:
#         break
#
# # 假设检验
# # 权重-公式， mannwhitneyu
# statistic, pvalue = stats.mannwhitneyu(a, b, alternative='two-sided')
# # 进行 Benjamini-Hochberg 校正
# p_adjusted = multipletests(pvalue, method='fdr_bh')[1]
# # 找到显著差异的位置
# significant_positions = [i + 1 for i in range(len(p_adjusted)) if p_adjusted[i] < 0.05]
# if len(significant_positions) == 0:
#     significant_positions = [-1]
# value = [statistic, pvalue, significant_positions[0]]


#
# # 统计每个权重下公式排名
# project = 'STVR'
# path = '../../data/results/result13.xlsx'  # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# sheet = Formula
# title = ['oms']
# t = 1
# for i in range(25):
#     title.append("i1d{}".format(100-t*2))
#     t += 1
# t = 1
# for i in range(25):
#     title.append("i{}d1".format(100-t*2))
#     t += 1
# fr = {}
# for k in range(len(title)):
#     Value = {}
#     for s in sheet:
#         if s == 'dice':
#             s = 'Dice'
#         ws = wb[s]
#         max_row = ws.max_row
#         value = []
#         for i in range(2, max_row+1):
#             value.append(ws.cell(i, k+2).value)
#         Value[s] = np.mean(value)
#     sorted_dict = dict(sorted(Value.items(), key=lambda x: x[1]))
#     fr[title[k]] = sorted_dict
#
# fr2 = {}
# for i in fr.keys():
#     a = []
#     for j in fr[i].keys():
#         a.append(j)
#     fr2[i] = a
#
# remove = ['Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman', 'qe', 'CBI Inc.', 'Hamann', 'Simple Matching',
#           'Sokal', 'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Rogot1']
# b = {}
# b1 = []
# for i in fr2.keys():
#     for k in remove:
#         if k in fr2[i]:
#             fr2[i].remove(k)
#     if fr2[i] not in b1:
#         b1.append(fr2[i])
#         b[i] = fr2[i]
#     else:
#         for j in b.keys():
#             if fr2[i] == b[j]:
#                 b[j+i] = b[j]
#                 break

# 方案之间， mannwhitneyu
# project = 'STVR'
# path = '../../data/results/result13.xlsx'
# path2 = '../../data/results/statistic3.xlsx'
# wb = load_workbook(path)
# wb2 = load_workbook(path2)
# sheet = Formula
# # scheme = ['MS', 'MS1', 'MS2', 'MS3', 'MS4', 'MS5', 'MS6', 'MS7', 'MS8', 'MS9', 'MS10']
# scheme = ['oms']
# t = 1
# for i in range(25):
#     scheme.append("i1d{}".format(100-t*2))
#     t += 1
# t = 1
# for i in range(25):
#     scheme.append("i{}d1".format(100-t*2))
#     t += 1
# scheme.append('ensemble1')
# scheme.append('ensemble2')
# scheme.append('imse1')
# scheme.append('imse2')
# scheme.append('dmse1')
# scheme.append('dmse2')
# # scheme = Labels
# remove = [19, 21,23,24,25,26,27,28,29,30,33,34,35,36,37,39,41,42,43,44,47,49,
#           50,51,57,58,59,61,62,63,65,66,67,69,70,72,74,75,76,79,80,81]
# for string in sheet:
#     if string == 'dice':
#         string = 'Dice'
#     ws = wb[string]
#     max_row = ws.max_row - 1
#     if string not in wb2.sheetnames:
#         ws2 = wb2.create_sheet(string)
#     del wb2[string]
#     ws2 = wb2.create_sheet(string)
#     row = 1
#     tablelist = {'MS pairs': ['statistic', 'pvalue', 'position']}
#     datadist = {}
#     datadist.update(tablelist)
#     fp = []
#     for m in range(1):
#         for n in range(1, len(scheme)):
#             i = 2
#             x1 = []
#             x2 = []
#             s = 0
#             if scheme[m] == scheme[n]:
#                 continue
#             # if not fp:
#             #     pass
#             # else:
#             #     for k in fp:
#             #         if scheme[m] + '-' + scheme[n] == k or scheme[n] + '-' + scheme[m] == k:
#             #             s = 1
#             #             break
#             # if s == 1:
#             #     continue
#             string2 = scheme[m] + '-' + scheme[n]
#             fp.append(string2)
#             while True:
#                 if i in remove:
#                     i += 1
#                     continue
#                 value1 = ws.cell(i, m+2).value
#                 value2 = ws.cell(i, n+2).value
#                 # if value1 == value2:
#                 #     i += 1
#                 #     if i > max_row:
#                 #         break
#                 #     continue
#                 x1.append(value1)
#                 x2.append(value2)
#                 i += 1
#                 if i > max_row:
#                     break
#             if len(x1) == 0 or len(x2) == 0:
#                 statistic = 1
#                 pvalue = 1
#                 significant_positions = [-1]
#             else:
#                 statistic, pvalue = stats.mannwhitneyu(x1, x2, alternative='two-sided')
#                 # 进行 Benjamini-Hochberg 校正
#                 p_adjusted = multipletests(pvalue, method='fdr_bh')[1]
#
#                 # 找到显著差异的位置
#                 significant_positions = [i + 1 for i in range(len(p_adjusted)) if p_adjusted[i] < 0.05]
#                 if len(significant_positions) == 0:
#                     significant_positions = [-1]
#             value = [statistic, pvalue, significant_positions[0]]
#             data = {
#                 # Formula[k]: value
#                 string2: value
#             }
#             datadist.update(data)
#
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws2.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws2.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     wb2.save(path2)


# 公式之间， mannwhitneyu
# project = 'STVR'
# path = '../../data/results/result5.xlsx'
# path2 = '../../data/results/statistic.xlsx'
# wb = load_workbook(path)
# wb2 = load_workbook(path2)
# sheet = ['MS0', 'MS1', 'MS2', 'MS3', 'MS4', 'MS5', 'MS6', 'MS7', 'MS8', 'MS9', 'MS10']
# # Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'dice',
# #            'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
# #            'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3',
# #            'Arithmetic Mean', 'Cohen', 'Fleiss']
# Formula = ['Ochiai', 'Jaccard', 'Tarantula', 'Naish1', 'Wong1', 'Scott', 'Wong2']
# for string in sheet:
#     ws = wb[string]
#     max_column = ws.max_column - 1
#     max_row = ws.max_row
#     if string not in wb2.sheetnames:
#         ws2 = wb2.create_sheet(string)
#     del wb2[string]
#     ws2 = wb2.create_sheet(string)
#     row = 1
#     tablelist = {'Formula pairs': ['statistic', 'pvalue']}
#     datadist = {}
#     datadist.update(tablelist)
#     fp = []
#     sort = []
#     for i in range(2, max_row + 1):
#         sort.append(ws.cell(i, 1).value)
#     for m in range(len(Formula)):
#         for n in range(len(Formula)):
#             i = 2
#             x1 = []
#             x2 = []
#             s = 0
#             if Formula[m] == Formula[n]:
#                 continue
#             if not fp:
#                 pass
#             else:
#                 for k in fp:
#                     if Formula[m] in k and Formula[n] in k:
#                         s = 1
#                         break
#             if s == 1:
#                 continue
#             string2 = Formula[m] + '-' + Formula[n]
#             fp.append(string2)
#             row1 = sort.index(Formula[m]) + 2
#             row2 = sort.index(Formula[n]) + 2
#             while True:
#                 value1 = ws.cell(row1, i).value
#                 value2 = ws.cell(row2, i).value
#                 # if value1 == value2:
#                 #     i += 1
#                 #     if i > max_column:
#                 #         break
#                 #     continue
#                 x1.append(value1)
#                 x2.append(value2)
#                 i += 1
#                 if i > max_column:
#                     break
#             if len(x1) == 0 or len(x2) == 0:
#                 statistic = 1
#                 pvalue = 1
#             else:
#                 statistic, pvalue = stats.mannwhitneyu(x1, x2, alternative='two-sided')
#             value = [statistic, pvalue]
#             data = {
#                 # Formula[k]: value
#                 string2: value
#             }
#             datadist.update(data)
#
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws2.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws2.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     wb2.save(path2)


# 统计FS前后对比，topn
# V = []
# V2 = []
# im_stats = []
# for N in range(26, 30):
#     project = 'EMS'
#     path = '../../data/results/_nofs.xlsx'.format(N)  # '+sys.argv[1][:-1]+'
#     wb = load_workbook(path)
#     path2 = '../../data/results/result{}.xlsx'.format(N)  # '+sys.argv[1][:-1]+'
#     wb2 = load_workbook(path2)
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
#                'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3',
#                'Arithmetic Mean', 'Cohen', 'Fleiss', 'Dstar', 'GP13']
#     Labels = ['oms']
#     t = 1
#     for i in range(25):
#         Labels.append("d{}".format(100-t*2))
#         t += 1
#     t = 1
#     for i in range(25):
#         Labels.append("i{}".format(100-t*2))
#         t += 1
#     ws = wb["Sum"]
#     max_row = ws.max_row
#     x = 2
#     v = []
#     vv = []
#     Value = []
#     for s in Formula:
#         if s == 'Naish1' or s == 'Binary':
#         # if not s == 'Kulczynski2':
#             x += 1
#             continue
#         a = []
#         for i in range(2, len(Labels) + 3):
#             if i == 3:
#                 continue
#             value = ws.cell(x, i).value
#             a.append(value)
#         # a.append(ws.cell(x, 54).value)  # fs, exam 60, topn 54
#         Value.append(a)
#         x += 1
#
#     result = []
#     for row in Value:
#         oms = row[0]
#         d_mean = np.mean(row[1:26])
#         i_mean = np.mean(row[26:51])
#         result.append([oms, d_mean, i_mean])
#
#     value = []
#     for i in range(len(Value[0])):
#         a = []
#         for j in range(len(Value)):
#             a.append(Value[j][i])
#         value.append(np.mean(a))
#     v.append(value)
#
#     for i in range(len(v[0])):
#         a = []
#         for j in range(len(v)):
#             a.append(v[j][i])
#         vv.append(round(sum(a)/len(a), 2))
#     average_1_25 = round(sum(vv[1:26]) / len(vv[1:26]), 2)
#     average_26_50 = round(sum(vv[26:51]) / len(vv[26:51]), 2)
#     del vv[26:51]
#     vv.insert(26, average_26_50)
#     del vv[1:26]
#     vv.insert(1, average_1_25)
#     V.append(vv)
#
#     ws = wb2["Sum"]
#     x = 2
#     v = []
#     vv = []
#
#     Value = []
#     for s in Formula:
#         if s == 'Naish1' or s == 'Binary':
#         # if not s == 'Kulczynski2':
#             x += 1
#             continue
#         a = []
#         for i in range(2, len(Labels) + 3):
#             if i == 3:
#                 continue
#             value = ws.cell(x, i).value
#             a.append(value)
#         # a.append(ws.cell(x, 54).value)  # fs, exam 60, topn 54
#         Value.append(a)
#         x += 1
#
#     result2 = []
#     for row in Value:
#         oms = row[0]
#         d_mean = np.mean(row[1:26])
#         i_mean = np.mean(row[26:51])
#         result2.append([oms, d_mean, i_mean])
#
#     value = []
#     for i in range(len(Value[0])):
#         a = []
#         for j in range(len(Value)):
#             a.append(Value[j][i])
#         value.append(np.mean(a))
#     v.append(value)
#
#     for i in range(len(v[0])):
#         a = []
#         for j in range(len(v)):
#             a.append(v[j][i])
#         vv.append(round(sum(a) / len(a), 2))
#     average_1_25 = round(sum(vv[1:26]) / len(vv[1:26]), 2)
#     average_26_50 = round(sum(vv[26:51]) / len(vv[26:51]), 2)
#     del vv[26:51]
#     vv.insert(26, average_26_50)
#     del vv[1:26]
#     vv.insert(1, average_1_25)
#     V2.append(vv)
#
#     improvement_stats = []
#
#     for j in range(3):  # 遍历三种策略
#         up = down = same = 0
#         max_increase = float('-inf')
#         min_improve = float('inf')
#         max_inc_index = -1
#         min_improve_index = -1
#
#         for i in range(30):  # 遍历30个公式
#             diff = result[i][j] - result2[i][j]
#             # 提升/降低/不变分类统计
#             if diff > 0:
#                 up += 1
#             elif diff < 0:
#                 down += 1
#             else:
#                 same += 1
#
#             # 最大提升值
#             if diff > max_increase:
#                 max_increase = diff
#                 max_inc_index = i
#
#             # 最小提升值（包含负数和0）
#             if diff < min_improve:
#                 min_improve = diff
#                 min_improve_index = i
#
#         improvement_stats.append({
#             'strategy_index': j,
#             '提升': up,
#             '不变': same,
#             '降低': down,
#             '最大提升值': round(max_increase, 2),
#             '最大提升公式索引': max_inc_index,
#             '最小提升值': round(min_improve, 2),
#             '最小提升公式索引': min_improve_index
#         })
#     im_stats.append(improvement_stats)
