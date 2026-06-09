import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import seaborn as sns
from matplotlib import ticker
from scipy.interpolate import make_interp_spline
import json
import matplotlib
from matplotlib.colors import LinearSegmentedColormap
import matplotlib.colors as mcolors
from matplotlib.lines import Line2D
from matplotlib.font_manager import FontProperties
# matplotlib.use('TkAgg')

# 热力图-TOP-N
# project = 'STVR'
# path = '../../data/results/result25.xlsx'  # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# sheet = ['Naish2', 'Wong1', 'Russel&Rao', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#            'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
#            'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3',
#            'Arithmetic Mean', 'Cohen', 'Fleiss']
# Labels = ['oms']
# t = 1
# for i in range(25):
#     Labels.append("i{}".format(100-t*2))
#     t += 1
# t = 1
# for i in range(25):
#     Labels.append("d{}".format(100-t*2))
#     t += 1
# Value = []
# for s in sheet:
#     ws = wb[s]
#     max_row = ws.max_row
#     value = []
#     prop = []
#     n = 2
#     for i in range(2, len(Labels)+2):
#         v = []
#         for c in range(2, max_row + 1):
#             v.append(ws.cell(c, i).value)
#         value.append(sum(v)/len(v))
#     Value.append(value)
#
# # 平均方式1
# # avg = []
# # for i in range(len(Value[0])):
# #     a = []
# #     for j in range(len(Value)):
# #         if j == 0 or j == 4:
# #             continue
# #         a.append(Value[j][i])
# #     avg.append(round(sum(a)/len(a), 2))
# # Value.append(avg)
#
# # 平均方式2
# Value1 = []
# for s in sheet:
#     if s == 'Naish1' or s == 'Binary':
#         continue
#     ws = wb[s]
#     max_row = ws.max_row
#     max_col = ws.max_column
#     value = []
#     n = 2
#     for i in range(2, len(Labels)+2):
#         v = []
#         for c in range(2, max_row + 1):
#             v.append(ws.cell(c, i).value)
#         value.append(np.mean(v))
#     Value1.append(value)
# value = []
# for i in range(len(Value1[0])):
#     a = []
#     for j in range(len(Value1)):
#         a.append(Value1[j][i])
#     value.append(np.mean(a))
# Value.append(value)
# transposed_Value = list(zip(*Value))
# # 将数据转换为NumPy数组
# data = np.array(transposed_Value)
# new_array_2d = []
# for i in range(1, 27):  # 遍历前26行
#     new_array_2d.append(data[26 - i])
# # 将剩余的行直接复制到新的二维数组中
# new_array_2d.extend(data[26:])
# data = np.array(new_array_2d)
# # 设置格子之间的间隔
# dx = dy = 2.5
# # 绘制热力图
# plt.figure(figsize=(25, 30))  # 设置图形大小（可选）
# heatmap = plt.imshow(data, cmap='GnBu', extent=[0, dx * data.shape[1], 0, dy * data.shape[0]])
# # 添加颜色条
# plt.colorbar(heatmap, aspect=30, pad=0.02)
#
# # 计算每一列的最大值和最小值
# max_values = np.max(data, axis=0)
# min_values = np.min(data, axis=0)
# dataf = np.flip(data, axis=0)
# # 添加标记
# for i, value in enumerate(min_values):
#     min_indices = np.where(dataf[:, i] == value)[0]
#     for index in min_indices:
#         plt.text((i + 0.5) * dx, (index + 0.5) * dy, u'\u2718', fontname='DejaVu Sans',
#                  fontsize="20", ha='center', va='center')
#
# for i, value in enumerate(max_values):
#     max_indices = np.where(dataf[:, i] == value)[0]
#     for index in max_indices:
#         plt.text((i + 0.5) * dx, (index + 0.5) * dy, u'\u2713', fontname='DejaVu Sans',
#                  fontsize="20",  ha='center', va='center')
#
# # 标记最大值和最小值
# for i, row in enumerate(dataf):
#     max_value = np.max(row)
#     min_value = np.min(row)
#     max_indices = np.where(row == max_value)[0]
#     min_indices = np.where(row == min_value)[0]
#     # 标记所有最大值（方框）
#     for min_index in min_indices:
#         plt.scatter((min_index + 0.5) * dx, (i + 0.5) * dy, marker='s',
#                     edgecolors='red', s=300, facecolors='none', linewidths=2)
#
#     # 标记所有最小值（圆圈）
#     for max_index in max_indices:
#         plt.scatter((max_index + 0.5) * dx, (i + 0.5) * dy, marker='o',
#                     edgecolors='green', s=300, facecolors='none', linewidths=2)
#     # # 标记最大值（方框）
#     # plt.scatter((max_index + 0.5) * dx, (i + 0.5) * dy, marker='s',
#     #             edgecolors='red', s=300, facecolors='none', linewidths=2)
#     #
#     # # 标记最小值（圆圈）
#     # plt.scatter((min_index + 0.5) * dx, (i + 0.5) * dy, marker='o',
#     #             edgecolors='green', s=300, facecolors='none', linewidths=2)
#
# # 获取最大值和最小值的索引
# max_index = np.unravel_index(np.argmax(dataf), dataf.shape)
# min_index = np.unravel_index(np.argmin(dataf), dataf.shape)
#
# # 设置标记符号的大小和颜色
# marker_size = 100
# marker_color = 'red'
#
# # 标记最大值和最小值
# plt.scatter((max_index[1]+0.5)*dx, (max_index[0]+0.5)*dy, marker='v', color=marker_color, s=marker_size)
# plt.scatter((min_index[1]+0.5)*dx, (min_index[0]+0.5)*dy, marker='^', color=marker_color, s=marker_size)
#
#
# # 设置横轴刻度标签
# x_ticks = np.arange(0, dx * data.shape[1], dx) + dx/2
# sheet.append('Avg.')
# plt.xticks(x_ticks, labels=sheet, rotation=45, ha='center')
# plt.tight_layout()  # 调整布局
# # 设置纵轴刻度标签
# # # 设置横轴刻度标签
# reversed_first_25 = Labels[:26][::-1]
# rest_of_list = Labels[26:]
# # 重新组合列表
# Labels3 = reversed_first_25 + rest_of_list
# y_ticks = np.arange(0, dy * data.shape[0], dy) + dy/2
# plt.yticks(y_ticks, labels=Labels3[::-1], va='center')
# # 添加标题和轴标签
# # plt.title('Heatmap')
# # plt.xlabel('X-axis')
# # plt.ylabel('Y-axis')
# # 显示图形
# plt.show()


# 画热力图
# 相比oms提升百分比
# project = 'STVR'
# path = '../../data/results/result16.xlsx'  # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# sheet = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#            'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
#            'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3',
#            'Arithmetic Mean', 'Cohen', 'Fleiss']
# # Labels = ['oms', 'dms9', 'dms8', 'dms7', 'dms6', 'dms5', 'ims9', 'ims8', 'ims7', 'ims6', 'ims5']
# Labels = ['oms']
# t = 1
# for i in range(25):
#     Labels.append("d{}".format(100-t*2))
#     t += 1
# t = 1
# for i in range(25):
#     Labels.append("i{}".format(100-t*2))
#     t += 1
# # Labels.append('e1')
# # Labels.append('e2')
# # Labels.append('ie1')
# # Labels.append('ie2')
# # Labels.append('de1')
# # Labels.append('de2')
# Value = []
# for s in sheet:
#     ws = wb[s]
#     max_row = ws.max_row
#     value = []
#     prop = []
#     n = 2
#     for i in range(2, len(Labels)+2):
#         v = []
#         for c in range(2, max_row + 1):
#             v.append(ws.cell(c, i).value)
#         value.append(sum(v)/len(v))
#     for i in range(1, len(value)):
#         prop.append(round((value[0] - value[i])/value[0]*100, 2))
#     Value.append(prop)
#
# # 平均方式1
# # avg = []
# # for i in range(len(Value[0])):
# #     a = []
# #     for j in range(len(Value)):
# #         if j == 0 or j == 4:
# #             continue
# #         a.append(Value[j][i])
# #     avg.append(round(sum(a)/len(a), 2))
# # Value.append(avg)
#
# # 平均方式2
# Value1 = []
# for s in sheet:
#     if s == 'Naish1' or s == 'Binary':
#         continue
#     ws = wb[s]
#     max_row = ws.max_row
#     max_col = ws.max_column
#     value = []
#     n = 2
#     for i in range(2, len(Labels)+2):
#         v = []
#         for c in range(2, max_row + 1):
#             v.append(ws.cell(c, i).value)
#         value.append(np.mean(v))
#     Value1.append(value)
# value = []
# for i in range(len(Value1[0])):
#     a = []
#     for j in range(len(Value1)):
#         a.append(Value1[j][i])
#     value.append(np.mean(a))
# prop = []
# for i in range(1, len(value)):
#     prop.append(round((value[0] - value[i])/value[0]*100, 2))
# Value.append(prop)
#
# transposed_Value = list(zip(*Value))
# # 将数据转换为NumPy数组
# data = np.array(transposed_Value)
# # data = np.random.rand(5, 5)  # 生成一个5x5的随机数据矩阵
# colors = ['red', 'white', 'green']
# cmap = mcolors.LinearSegmentedColormap.from_list('custom_cmap', colors)
# # 定义归一化器
# vmin = -abs(np.min(data))  # 白色所代表的值
# # vmax = max(abs(np.min(data)), abs(np.max(data)))  # 数据的最大绝对值
# vmax = np.max(data)  # 数据的最大绝对值
# norm = mcolors.TwoSlopeNorm(vmin=vmin, vcenter=0, vmax=vmax)
# # colormap = LinearSegmentedColormap.from_list('CustomColors', colors)
# # 设置格子之间的间隔
# dx = dy = 2.5
# # 绘制热力图
# plt.figure(figsize=(30, 30))  # 设置图形大小（可选）
# heatmap = plt.imshow(data, cmap=cmap, norm=norm, extent=[0, dx * data.shape[1], 0, dy * data.shape[0]])
# # 添加颜色条
# plt.colorbar(heatmap, aspect=30, pad=0.02)
#
# # 计算每一列的最大值和最小值
# max_values = np.max(data, axis=0)
# min_values = np.min(data, axis=0)
# dataf = np.flip(data, axis=0)
# # 添加标记
# for i, value in enumerate(max_values):
#     max_indices = np.where(dataf[:, i] == value)[0]
#     for index in max_indices:
#         plt.text((i + 0.5) * dx, (index + 0.5) * dy, u'\u2713', fontname='DejaVu Sans', fontsize="20",
#                  ha='center', va='center')
#
# for i, value in enumerate(min_values):
#     min_indices = np.where(dataf[:, i] == value)[0]
#     for index in min_indices:
#         plt.text((i + 0.5) * dx, (index + 0.5) * dy, u'\u2718', fontname='DejaVu Sans', fontsize="20",
#                  ha='center', va='center')
#
# # 标记最大值和最小值
# for i, row in enumerate(dataf):
#     max_value = np.max(row)
#     min_value = np.min(row)
#     max_index = np.where(row == max_value)[0][0]
#     min_index = np.where(row == min_value)[0][0]
#
#     # 标记最大值（方框）
#     plt.scatter((max_index + 0.5) * dx, (i + 0.5) * dy, marker='o',
#                 edgecolors='blue', s=300, facecolors='none', linewidths=2)
#
#     # 标记最小值（圆圈）
#     plt.scatter((min_index + 0.5) * dx, (i + 0.5) * dy, marker='s',
#                 edgecolors='blue', s=300, facecolors='none', linewidths=2)
#
# # 获取最大值和最小值的索引
# max_index = np.unravel_index(np.argmax(dataf), dataf.shape)
# min_index = np.unravel_index(np.argmin(dataf), dataf.shape)
#
# # 设置标记符号的大小和颜色
# marker_size = 100
# marker_color = 'black'
#
# # 标记最大值和最小值
# plt.scatter((max_index[1]+0.5)*dx, (max_index[0]+0.5)*dy, marker='^', color=marker_color, s=marker_size)
# plt.scatter((min_index[1]+0.5)*dx, (min_index[0]+0.5)*dy, marker='v', color=marker_color, s=marker_size)
#
#
# # 设置横轴刻度标签
# x_ticks = np.arange(0, dx * data.shape[1], dx) + dx/2
# sheet.append('Avg.')
# plt.xticks(x_ticks, labels=sheet, rotation=45, ha='center')
# plt.tight_layout()  # 调整布局
# # 设置纵轴刻度标签
# y_ticks = np.arange(0, dy * data.shape[0], dy) + dy/2
# plt.yticks(y_ticks, labels=Labels[::-1][:-1], va='center')
# # 添加标题和轴标签
# # plt.title('Heatmap')
# # plt.xlabel('X-axis')
# # plt.ylabel('Y-axis')
#
# # 显示图形
# plt.show()

# 画热力图
# 按program
# project = 'STVR'
# path = '../../data/results/result16.xlsx'  # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#            'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
#            'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3',
#            'Arithmetic Mean', 'Cohen', 'Fleiss']
# # Formula2 = ['Naish2', 'Kulczynski2', 'Ochiai', 'M2']
# program = ['TSQ', 'DM', 'Tcas', 'SMM', 'grep', 'PT', 'PT2']  # 'TSQ', 'DM', 'Tcas', 'SMM', 'grep', 'PT', 'PT2'
# Labels = ['oms']
# t = 1
# for i in range(25):
#     Labels.append("d{}".format(100-t*2))
#     t += 1
# t = 1
# for i in range(25):
#     Labels.append("i{}".format(100-t*2))
#     t += 1
# # Labels.append('e1')
# # Labels.append('e2')
# # Labels.append('ie1')
# # Labels.append('ie2')
# # Labels.append('de1')
# # Labels.append('de2')
# Value = []
# for k in program:
#     v = []
#     ws = wb[k]
#     for s in Formula:
#         V = []
#         max_row = ws.max_row
#         t = 0
#         n = 2
#         x = 2
#         while True:
#             if ws.cell(x, 59).value == 100 or ws.cell(x, 59).value == 0:
#                 x += 1
#                 n += 1
#                 if n == 32:
#                     n = 2
#                     x += 3
#                 if x > max_row:
#                     break
#                 continue
#             if s == 'Naish1' or s == 'Binary':
#             # if s not in Formula2:
#                 x += 1
#                 n += 1
#                 if n == 32:
#                     n = 2
#                     x += 3
#                 if x > max_row:
#                     break
#                 continue
#             a = []
#             for i in range(2, len(Labels) + 2):
#                 if ws.cell(x, 1).value not in [s]:
#                     break
#                 value = ws.cell(x, i).value
#                 a.append(value)
#             if len(a) == 0:
#                 pass
#             else:
#                 V.append(a)
#             x += 1
#             n += 1
#             if n == 32:
#                 n = 2
#                 x += 3
#             if x > max_row:
#                 break
#         if len(V) == 0:
#             pass
#         else:
#             v.append(V)
#     c = []
#     for i in range(len(v)):
#         a = []
#         for j in range(len(v[0][0])):
#             b = []
#             for m in range(len(v[0])):
#                 b.append(v[i][m][j])
#             a.append(np.mean(b))
#         c.append(a)
#     f = []
#     for i in range(len(c[0])):
#         d = []
#         for j in range(len(c)):
#             d.append(c[j][i])
#         f.append(round(sum(d)/len(d), 2))
#     Value.append(f)
#
# Value1 = []
# for s in Formula:
#     if s == 'Naish1' or s == 'Binary':
#     # if s not in Formula2:
#         continue
#     ws = wb[s]
#     max_row = ws.max_row
#     max_col = ws.max_column
#     value = []
#     n = 2
#     for i in range(2, len(Labels)+2):
#         v = []
#         for c in range(2, max_row + 1):
#             v.append(ws.cell(c, i).value)
#         value.append(np.mean(v))
#     Value1.append(value)
# value = []
# for i in range(len(Value1[0])):
#     a = []
#     for j in range(len(Value1)):
#         a.append(Value1[j][i])
#     value.append(np.mean(a))
# Value.append(value)
#
# transposed_Value = list(zip(*Value))
# # 将数据转换为NumPy数组
# data = np.array(Value)
# for row in data:
#     row[:26] = row[:26][::-1]
# # 设置格子之间的间隔
# dx = dy = 2.5
# # 绘制热力图
# plt.figure(figsize=(30, 10))  # 设置图形大小（可选）
# heatmap = plt.imshow(data, cmap='GnBu', extent=[0, dx * data.shape[1], 0, dy * data.shape[0]])
# # 添加颜色条
# plt.colorbar(heatmap, aspect=60, pad=0.05, orientation='horizontal')
#
# max_values = np.max(data, axis=1)
# min_values = np.min(data, axis=1)
# dataf = np.flip(data, axis=0)
# # 添加标记
# for i, value in enumerate(max_values[::-1]):
#     max_indices = np.where(dataf[i, :] == value)[0]
#     for index in max_indices:
#         plt.scatter((index + 0.5) * dx, (i + 0.5) * dy, marker='^', color='black', s=300, linewidths=2)
#
# for i, value in enumerate(min_values[::-1]):
#     min_indices = np.where(dataf[i, :] == value)[0]
#     for index in min_indices:
#         plt.scatter((index + 0.5) * dx, (i + 0.5) * dy, marker='v', color='blue', s=300, linewidths=2)
#
# # # 按行标记最大值和最小值
# # for i, row in enumerate(data):
# #     max_value = np.max(row)
# #     min_value = np.min(row)
# #     max_index = np.where(row == max_value)[0][0]
# #     min_index = np.where(row == min_value)[0][0]
# #     # 标记最大值（方框）
# #     plt.scatter((max_index + 0.5) * dx, (i + 0.5) * dy, marker='v',
# #                 edgecolors='red', s=300, facecolors='none', linewidths=2)
# #
# #     # 标记最小值（圆圈）
# #     plt.scatter((min_index + 0.5) * dx, (i + 0.5) * dy, marker='^',
# #                 edgecolors='green', s=300, facecolors='none', linewidths=2)
#
# # 设置横轴刻度标签
# x_ticks = np.arange(0, dx * data.shape[1], dx) + dx/2
# # # 设置横轴刻度标签
# reversed_first_25 = Labels[:26][::-1]
# rest_of_list = Labels[26:]
# # # 重新组合列表
# Labels3 = reversed_first_25 + rest_of_list
# program.append('Avg.')
# plt.xticks(x_ticks, labels=Labels3, rotation=45, ha='center')
# plt.tight_layout()  # 调整布局
# # 设置纵轴刻度标签
# y_ticks = np.arange(0, dy * data.shape[0], dy) + dy/2
# plt.yticks(y_ticks, labels=program[::-1], va='center')
# # 添加标题和轴标签
# # plt.title('Heatmap')
# # plt.xlabel('X-axis')
# # plt.ylabel('Y-axis')
#
# # 显示图形
# plt.show()


# 画热力图
# 按program，提升百分比
# project = 'STVR'
# path = '../../data/results/result16.xlsx'  # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#            'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
#            'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3',
#            'Arithmetic Mean', 'Cohen', 'Fleiss']
# # Formula2 = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
# #            'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
# #            'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3',
# #            'Arithmetic Mean', 'Cohen']
# program = ['TSQ', 'DM', 'Tcas', 'SMM', 'grep', 'PT', 'PT2']  # 'TSQ', 'DM', 'Tcas', 'SMM', 'grep', 'PT', 'PT2'
# Labels = ['oms']
# t = 1
# for i in range(25):
#     Labels.append("d{}".format(100-t*2))
#     t += 1
# t = 1
# for i in range(25):
#     Labels.append("i{}".format(100-t*2))
#     t += 1
# # Labels.append('e1')
# # Labels.append('e2')
# # Labels.append('ie1')
# # Labels.append('ie2')
# # Labels.append('de1')
# # Labels.append('de2')
# Value = []
# for k in program:
#     v = []
#     ws = wb[k]
#     for s in Formula:
#         V = []
#         max_row = ws.max_row
#         t = 0
#         n = 2
#         x = 2
#         while True:
#             if ws.cell(x, 59).value == 100 or ws.cell(x, 59).value == 0:
#                 x += 1
#                 n += 1
#                 if n == 32:
#                     n = 2
#                     x += 3
#                 if x > max_row:
#                     break
#                 continue
#             if s == 'Naish1' or s == 'Binary':
#             # if s not in Formula2:
#                 x += 1
#                 n += 1
#                 if n == 32:
#                     n = 2
#                     x += 3
#                 if x > max_row:
#                     break
#                 continue
#             a = []
#             for i in range(2, len(Labels) + 2):
#                 if ws.cell(x, 1).value not in [s]:
#                     break
#                 value = ws.cell(x, i).value
#                 a.append(value)
#             if len(a) == 0:
#                 pass
#             else:
#                 V.append(a)
#             x += 1
#             n += 1
#             if n == 32:
#                 n = 2
#                 x += 3
#             if x > max_row:
#                 break
#         if len(V) == 0:
#             pass
#         else:
#             v.append(V)
#     c = []
#     for i in range(len(v)):
#         a = []
#         for j in range(len(v[0][0])):
#             b = []
#             for m in range(len(v[0])):
#                 b.append(v[i][m][j])
#             a.append(np.mean(b))
#         c.append(a)
#     f = []
#     for i in range(len(c[0])):
#         d = []
#         for j in range(len(c)):
#             d.append(c[j][i])
#         f.append(round(sum(d)/len(d), 2))
#     prop = []
#     for i in range(1, len(f)):
#         prop.append(round((f[0] - f[i]) / f[0] * 100, 2))
#     Value.append(prop)
#
# Value1 = []
# for s in Formula:
#     if s == 'Naish1' or s == 'Binary':
#     # if s not in Formula2:
#         continue
#     ws = wb[s]
#     max_row = ws.max_row
#     max_col = ws.max_column
#     value = []
#     n = 2
#     for i in range(2, len(Labels)+2):
#         v = []
#         for c in range(2, max_row + 1):
#             v.append(ws.cell(c, i).value)
#         value.append(np.mean(v))
#     Value1.append(value)
# value = []
# for i in range(len(Value1[0])):
#     a = []
#     for j in range(len(Value1)):
#         a.append(Value1[j][i])
#     value.append(np.mean(a))
# prop = []
# for i in range(1, len(value)):
#     prop.append(round((value[0] - value[i])/value[0]*100, 2))
# Value.append(prop)
#
# transposed_Value = list(zip(*Value))
# # 将数据转换为NumPy数组
# data = np.array(Value)
# for row in data:
#     row[:25] = row[:25][::-1]
# # 设置格子之间的间隔
# dx = dy = 2.5
# colors = ['red', 'white', 'green']
# cmap = mcolors.LinearSegmentedColormap.from_list('custom_cmap', colors)
# # 定义归一化器
# vmin = -abs(np.min(data))  # 白色所代表的值
# # vmax = max(abs(np.min(data)), abs(np.max(data)))  # 数据的最大绝对值
# vmax = np.max(data)  # 数据的最大绝对值
# norm = mcolors.TwoSlopeNorm(vmin=vmin, vcenter=0, vmax=vmax)
# # 绘制热力图
# plt.figure(figsize=(30, 10))  # 设置图形大小（可选）
# heatmap = plt.imshow(data, cmap=cmap, norm=norm, extent=[0, dx * data.shape[1], 0, dy * data.shape[0]])
# # 添加颜色条
# plt.colorbar(heatmap, aspect=60, pad=0.05, orientation='horizontal')
#
# max_values = np.max(data, axis=1)
# min_values = np.min(data, axis=1)
# dataf = np.flip(data, axis=0)
# # 添加标记
# for i, value in enumerate(max_values[::-1]):
#     max_indices = np.where(dataf[i, :] == value)[0]
#     if value < 0:
#         for index in max_indices:
#             plt.scatter((index + 0.5) * dx, (i + 0.5) * dy, marker='v', color='black', s=300, linewidths=2)
#     else:
#         for index in max_indices:
#             plt.scatter((index + 0.5) * dx, (i + 0.5) * dy, marker='v', color='blue', s=300, linewidths=2)
#
# for i, value in enumerate(min_values[::-1]):
#     min_indices = np.where(dataf[i, :] == value)[0]
#     for index in min_indices:
#         plt.scatter((index + 0.5) * dx, (i + 0.5) * dy, marker='^', color='black', s=300, linewidths=2)
#
# # 设置横轴刻度标签
# Labels2 = Labels[1:]
# reversed_first_25 = Labels2[:25][::-1]
# rest_of_list = Labels2[25:]
# # 重新组合列表
# Labels3 = reversed_first_25 + rest_of_list
#
# x_ticks = np.arange(0, dx * data.shape[1], dx) + dx/2
# program.append('Avg.')
# plt.xticks(x_ticks, labels=Labels3, rotation=45, ha='center')
# plt.tight_layout()  # 调整布局
# # 设置纵轴刻度标签
# y_ticks = np.arange(0, dy * data.shape[0], dy) + dy/2
# plt.yticks(y_ticks, labels=program[::-1], va='center')
# # 添加标题和轴标签
# # plt.title('Heatmap')
# # plt.xlabel('X-axis')
# # plt.ylabel('Y-axis')
#
# # 显示图形
# plt.show()


# 画箱线图，对比ims和dms
# project = 'EMS'
# path = '../../data/results/29.xlsx'  # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# ws = wb['Sheet1']
# max_row = ws.max_row
# Value = []
# for i in range(2, 6):
#     x = 2
#     a = []
#     while True:
#         value = ws.cell(x, i).value
#         a.append(value)
#         x += 1
#         if x > max_row:
#             break
#     Value.append(a)
# data = {
#     'Category': ['TOP-1'] * 50 + ['TOP-3'] * 50 + ['TOP-5'] * 50 + ['TOP-10'] * 50,
#     'Group': ['ims'] * 25 + ['dms'] * 25 + ['ims'] * 25 + ['dms'] * 25 + ['ims'] * 25 + ['dms'] * 25 + ['ims'] * 25 + ['dms'] * 25,
#     'Values': np.concatenate([Value[0][1:26], Value[0][26:],
#                               Value[1][1:26], Value[1][26:],
#                               Value[2][1:26], Value[2][26:],
#                               Value[3][1:26], Value[3][26:]])
# }
# matplotlib.rcParams['font.family'] = ["Hiragino Sans GB", "DejaVu Sans"]  # 中文字体
# # 创建DataFrame
# df = pd.DataFrame(data)
# # custom_palette = {'ims': '#1f77b4', 'dms': '#2ca02c'}
# custom_palette = {'ims': '#8e44ad', 'dms': '#27ae60'}
# # 定义额外的数据点
# extra_data = {'TOP-1': Value[0][0], 'TOP-3': Value[1][0], 'TOP-5': Value[2][0], 'TOP-10': Value[3][0]}
# # # 创建图形
# # plt.figure()
# #
# # # 绘制箱线图
# # ax = sns.boxplot(x='Category', y='Values', hue='Group', data=df, palette=custom_palette)
# #
# # # 设置标题和标签
# # # plt.title('Boxplot of Values by Category and Group')
# # # plt.xlabel('Category')
# # plt.ylabel('Values')
# #
# # # 显示图例
# # plt.legend(title='')
# # # 隐藏横轴标签
# # ax.set_xlabel('')
# #
# # # 显示图形
# # plt.show()
#
# # 创建图形
#
# font_chinese = FontProperties(fname='./fonts/SimSun.ttf')
# font_english = FontProperties(fname='/System/Library/Fonts/Supplemental/Times New Roman.ttf')
# font_chinese.set_size(15)  # 这里设置字体大小
# font_english.set_size(15)  # 这里设置字体大小
#
# fig, axes = plt.subplots(1, 4, sharex=True)
#
# # 定义每个子图的数据范围
# categories = ['TOP-1', 'TOP-3', 'TOP-5', 'TOP-10']
#
# for i, (ax, category) in enumerate(zip(axes, categories)):
#     subset = df[df['Category'] == category]
#
#     if subset.empty:
#         print(f"Warning: The subset for category {category} is empty. Skipping this plot.")
#         continue
#
#     sns.boxplot(x='Group', y='Values', hue='Group', data=df[df['Category'] == category],
#                 palette=custom_palette, ax=ax, width=0.4,
#                 flierprops={'marker': 'd', 'markerfacecolor': 'black', 'markeredgecolor': 'black'})
#     ax.set_title(category, fontproperties=font_english)
#     ax.set_xlabel('')  # 隐藏横轴标签
#
#     # 获取当前的纵坐标范围
#     current_ylim = ax.get_ylim()
#     # 计算新的上限
#     current_range = current_ylim[1] - current_ylim[0]
#     new_ylim_upper = current_ylim[1] + 0.06 * current_range
#     # 设置新的纵坐标范围
#     ax.set_ylim(current_ylim[0], new_ylim_upper)
#
#     if i == 0:
#         # ax.set_ylabel('Values')  # 仅在第一个子图显示纵轴标签
#         ax.set_ylabel('指标值', fontproperties=font_chinese)  # 仅在第一个子图显示纵轴标签
#     # elif i == 3:
#     #     ax.set_ylabel('')  # 其他子图不显示纵轴标签
#     #     ax.set_ylim(current_ylim[0]-0.2 * current_range, current_ylim[1] + 0.08 * current_range)
#     else:
#         ax.set_ylabel('')  # 其他子图不显示纵轴标签
#         # ax.yaxis.set_ticklabels([])  # 其他子图不显示纵轴刻度
#     # 添加额外的数据点
#     extra_value = extra_data[category]
#     ax.plot([0.5], [extra_value], 'ro', lw=0)  # 使用短线表示额外数据点
#     line = Line2D([0], [0], color='r', lw=0, marker='o', label='oms')
#     # 添加图例
#     # legend = ax.legend(handles=[line], loc='upper left', prop=font_english)
#     legend = ax.legend(
#         handles=[line],
#         loc='upper left',
#         prop=font_english,
#         handlelength=1.0,  # 控制图例前图标的长度，默认是2.0
#         handletextpad=0.4,  # 控制图标和文字之间的间距，默认是0.8
#         borderaxespad=0.3,  # 控制 legend 和图的位置的距离，默认0.5
#         borderpad=0.2 # 默认是 0.4，减小它可以让框更紧凑
#     )
#     # 调整图例标签颜色和位置
#     for text in legend.get_texts():
#         text.set_color('red')  # 设置图例标签的颜色为红色
#         # text.set_verticalalignment('bottom')  # 设置标签位置为线条下方
#     # 设置 ax1 的 x 和 y 轴刻度标签字体
#     for label in ax.get_xticklabels():
#         label.set_fontproperties(font_english)
#     for label in ax.get_yticklabels():
#         label.set_fontproperties(font_english)
#
# # 调整布局
# plt.tight_layout()
# plt.show()
# plt.savefig("./figures/EMS-比较结果one.pdf", format="pdf", bbox_inches='tight')


# 画箱线图，对比ims和dms
# project = 'EMS'
# path = '../../data/results/29.xlsx'  # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# ws = wb['Sheet1']
# max_row = ws.max_row
# Value = []
# for i in range(2, 6):
#     x = 2
#     a = []
#     while True:
#         if x == 3:
#             x += 1
#             continue
#         value = ws.cell(x, i).value
#         a.append(value)
#         x += 1
#         if x > max_row:
#             break
#     Value.append(a)
# data = {
#     'Category': ['TOP-1'] * 50 + ['TOP-3'] * 50 + ['TOP-5'] * 50 + ['TOP-10'] * 50,
#     'Group': ['ims'] * 25 + ['dms'] * 25 + ['ims'] * 25 + ['dms'] * 25 + ['ims'] * 25 + ['dms'] * 25 + ['ims'] * 25 + ['dms'] * 25,
#     'Values': np.concatenate([Value[0][1:26], Value[0][26:],
#                               Value[1][1:26], Value[1][26:],
#                               Value[2][1:26], Value[2][26:],
#                               Value[3][1:26], Value[3][26:]])
# }
# matplotlib.rcParams['font.family'] = ["Hiragino Sans GB", "DejaVu Sans"]  # 中文字体
# # 创建DataFrame
# df = pd.DataFrame(data)
# # custom_palette = {'ims': '#1f77b4', 'dms': '#2ca02c'}
# custom_palette = {'ims': '#8e44ad', 'dms': '#27ae60'}
# # 定义额外的数据点
# extra_data = {'TOP-1': Value[0][0], 'TOP-3': Value[1][0], 'TOP-5': Value[2][0], 'TOP-10': Value[3][0]}
# # # 创建图形
# # plt.figure()
# #
# # # 绘制箱线图
# # ax = sns.boxplot(x='Category', y='Values', hue='Group', data=df, palette=custom_palette)
# #
# # # 设置标题和标签
# # # plt.title('Boxplot of Values by Category and Group')
# # # plt.xlabel('Category')
# # plt.ylabel('Values')
# #
# # # 显示图例
# # plt.legend(title='')
# # # 隐藏横轴标签
# # ax.set_xlabel('')
# #
# # # 显示图形
# # plt.show()
#
# # 创建图形
#
# # font_chinese = FontProperties(fname='./fonts/SimSun.ttf')
# font_english = FontProperties(fname='/System/Library/Fonts/Supplemental/Times New Roman.ttf')
# # font_chinese.set_size(15)  # 这里设置字体大小
# font_english.set_size(13)  # 这里设置字体大小
#
# fig, axes = plt.subplots(1, 4, sharex=True)
#
# # 定义每个子图的数据范围
# categories = ['TOP-1', 'TOP-3', 'TOP-5', 'TOP-10']
#
# for i, (ax, category) in enumerate(zip(axes, categories)):
#     subset = df[df['Category'] == category]
#
#     if subset.empty:
#         print(f"Warning: The subset for category {category} is empty. Skipping this plot.")
#         continue
#
#     sns.boxplot(x='Group', y='Values', hue='Group', data=df[df['Category'] == category],
#                 palette=custom_palette, ax=ax, width=0.4,
#                 flierprops={'marker': 'd', 'markerfacecolor': 'black', 'markeredgecolor': 'black'})
#     ax.set_title(category, fontproperties=font_english)
#     ax.set_xlabel('')  # 隐藏横轴标签
#
#     # 获取当前的纵坐标范围
#     current_ylim = ax.get_ylim()
#     # 计算新的上限
#     current_range = current_ylim[1] - current_ylim[0]
#     new_ylim_upper = current_ylim[1] + 0.06 * current_range
#     # 设置新的纵坐标范围
#     ax.set_ylim(current_ylim[0], new_ylim_upper)
#
#     if i == 0:
#         # ax.set_ylabel('Values')  # 仅在第一个子图显示纵轴标签
#         # ax.set_ylabel('指标值', fontproperties=font_chinese)  # 仅在第一个子图显示纵轴标签
#         ax.set_ylabel('Values', fontproperties=font_english)  # 仅在第一个子图显示纵轴标签
#     # elif i == 3:
#     #     ax.set_ylabel('')  # 其他子图不显示纵轴标签
#     #     ax.set_ylim(current_ylim[0]-0.2 * current_range, current_ylim[1] + 0.08 * current_range)
#     else:
#         ax.set_ylabel('')  # 其他子图不显示纵轴标签
#         # ax.yaxis.set_ticklabels([])  # 其他子图不显示纵轴刻度
#     # 添加额外的数据点
#     extra_value = extra_data[category]
#     ax.plot([0.5], [extra_value], 'ro', lw=0)  # 使用短线表示额外数据点
#     line = Line2D([0], [0], color='r', lw=0, marker='o', label='oms')
#     # 添加图例
#     # legend = ax.legend(handles=[line], loc='upper left', prop=font_english)
#     legend = ax.legend(
#         handles=[line],
#         loc='upper left',
#         prop=font_english,
#         handlelength=1.0,  # 控制图例前图标的长度，默认是2.0
#         handletextpad=0.4,  # 控制图标和文字之间的间距，默认是0.8
#         borderaxespad=0.3,  # 控制 legend 和图的位置的距离，默认0.5
#         borderpad=0.2 # 默认是 0.4，减小它可以让框更紧凑
#     )
#     # 调整图例标签颜色和位置
#     for text in legend.get_texts():
#         text.set_color('red')  # 设置图例标签的颜色为红色
#         # text.set_verticalalignment('bottom')  # 设置标签位置为线条下方
#     # 设置 ax1 的 x 和 y 轴刻度标签字体
#     for label in ax.get_xticklabels():
#         label.set_fontproperties(font_english)
#     for label in ax.get_yticklabels():
#         label.set_fontproperties(font_english)
#
# # 调整布局
# plt.tight_layout()
# plt.show()
# plt.savefig('./figures/output.pdf'
#             "提交/修改/TOSEM/提交/修改/RQ3-all.pdf", format="pdf", bbox_inches='tight')

# 画箱线图，对比ims和dms，统一纵轴刻度
# project = 'EMS'
# path = '../../data/results/29.xlsx'  # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# ws = wb['Sheet1']
# max_row = ws.max_row
# Value = []
# for i in range(2, 6):
#     x = 2
#     a = []
#     while True:
#         if x == 3:
#             x += 1
#             continue
#         value = ws.cell(x, i).value
#         a.append(value)
#         x += 1
#         if x > max_row:
#             break
#     Value.append(a)
# Value = np.array(Value)
# data = {
#     'Category': ['TOP-1'] * 50 + ['TOP-3'] * 50 + ['TOP-5'] * 50 + ['TOP-10'] * 50,
#     'Group': ['ims'] * 25 + ['dms'] * 25 + ['ims'] * 25 + ['dms'] * 25 + ['ims'] * 25 + ['dms'] * 25 + ['ims'] * 25 + ['dms'] * 25,
#     'Values': np.concatenate([Value[0][1:26], Value[0][26:],
#                               Value[1][1:26], Value[1][26:],
#                               Value[2][1:26], Value[2][26:],
#                               Value[3][1:26], Value[3][26:]])
# }
# matplotlib.rcParams['font.family'] = ["Hiragino Sans GB", "DejaVu Sans"]  # 中文字体
# # 创建DataFrame
# df = pd.DataFrame(data)
# # custom_palette = {'ims': '#1f77b4', 'dms': '#2ca02c'}
# custom_palette = {'ims': '#8e44ad', 'dms': '#27ae60'}
# # 定义额外的数据点
# extra_data = {'TOP-1': Value[0][0], 'TOP-3': Value[1][0], 'TOP-5': Value[2][0], 'TOP-10': Value[3][0]}
# # # 创建图形
# # font_chinese = FontProperties(fname='./fonts/SimSun.ttf')
# font_english = FontProperties(fname='/System/Library/Fonts/Supplemental/Times New Roman.ttf')
# # font_chinese.set_size(15)  # 这里设置字体大小
# font_english.set_size(13)  # 这里设置字体大小
#
# fig, axes = plt.subplots(1, 4, sharex=True)
#
# # 定义每个子图的数据范围
# categories = ['TOP-1', 'TOP-3', 'TOP-5', 'TOP-10']
# # 找出所有 Values 中的最大值和最小值
# # 获取每个子图中的主要数据（排除第0个值，即额外点）
# all_main_values = np.concatenate([
#     Value[0][1:], Value[1][1:], Value[2][1:], Value[3][1:]
# ])
# ymin = all_main_values.min()
# ymax = all_main_values.max()
# yrange = ymax - ymin
#
# # 略微扩展上下界
# ymin -= 0.05 * yrange
# ymax += 0.05 * yrange
# for i, (ax, category) in enumerate(zip(axes, categories)):
#     subset = df[df['Category'] == category]
#
#     if subset.empty:
#         print(f"Warning: The subset for category {category} is empty. Skipping this plot.")
#         continue
#
#     sns.boxplot(x='Group', y='Values', hue='Group', data=df[df['Category'] == category],
#                 palette=custom_palette, ax=ax, width=0.4,
#                 flierprops={'marker': 'd', 'markerfacecolor': 'black', 'markeredgecolor': 'black'})
#     ax.set_title(category, fontproperties=font_english)
#     ax.set_xlabel('')  # 隐藏横轴标签
#
#     ax.set_ylim(ymin, ymax)
#
#     if i == 0:
#         # ax.set_ylabel('Values')  # 仅在第一个子图显示纵轴标签
#         # ax.set_ylabel('指标值', fontproperties=font_chinese)  # 仅在第一个子图显示纵轴标签
#         ax.set_ylabel('Values', fontproperties=font_english)  # 仅在第一个子图显示纵轴标签
#     # elif i == 3:
#     #     ax.set_ylabel('')  # 其他子图不显示纵轴标签
#     #     ax.set_ylim(current_ylim[0]-0.2 * current_range, current_ylim[1] + 0.08 * current_range)
#     else:
#         ax.set_ylabel('')  # 其他子图不显示纵轴标签
#         # ax.yaxis.set_ticklabels([])  # 其他子图不显示纵轴刻度
#     # 添加额外的数据点
#     extra_value = extra_data[category]
#     ax.plot([0.5], [extra_value], 'ro', lw=0)  # 使用短线表示额外数据点
#     line = Line2D([0], [0], color='r', lw=0, marker='o', label='oms')
#     # 添加图例
#     # legend = ax.legend(handles=[line], loc='upper left', prop=font_english)
#     legend = ax.legend(
#         handles=[line],
#         loc='upper left',
#         prop=font_english,
#         handlelength=1.0,  # 控制图例前图标的长度，默认是2.0
#         handletextpad=0.4,  # 控制图标和文字之间的间距，默认是0.8
#         borderaxespad=0.3,  # 控制 legend 和图的位置的距离，默认0.5
#         borderpad=0.2 # 默认是 0.4，减小它可以让框更紧凑
#     )
#     # 调整图例标签颜色和位置
#     for text in legend.get_texts():
#         text.set_color('red')  # 设置图例标签的颜色为红色
#         # text.set_verticalalignment('bottom')  # 设置标签位置为线条下方
#     # 设置 ax1 的 x 和 y 轴刻度标签字体
#     for label in ax.get_xticklabels():
#         label.set_fontproperties(font_english)
#     for label in ax.get_yticklabels():
#         label.set_fontproperties(font_english)
#
# # 调整布局
# plt.tight_layout()
# plt.show()
# plt.savefig('./figures/output.pdf'
#             "提交/修改/TOSEM/提交/修改/RQ1.pdf", format="pdf", bbox_inches='tight')

# 折线图，对比ims和dms
# project = 'EMS'
# path = '../../data/results/29.xlsx'  # '+sys.argv[1][:-1]+'
# wb = load_workbook(path)
# ws = wb['Sheet1']
# max_row = ws.max_row
# Value = []
# for i in range(2, 6):
#     x = 2
#     a = []
#     while True:
#         if x == 3:
#             x += 1
#             continue
#         value = ws.cell(x, i).value
#         a.append(value)
#         x += 1
#         if x > max_row:
#             break
#     Value.append(a)
# Value = np.array(Value)
# data = {
#     'Category': ['TOP-1'] * 50 + ['TOP-3'] * 50 + ['TOP-5'] * 50 + ['TOP-10'] * 50,
#     'Group': ['ims'] * 25 + ['dms'] * 25 + ['ims'] * 25 + ['dms'] * 25 + ['ims'] * 25 + ['dms'] * 25 + ['ims'] * 25 + ['dms'] * 25,
#     'Values': np.concatenate([Value[0][1:26], Value[0][26:],
#                               Value[1][1:26], Value[1][26:],
#                               Value[2][1:26], Value[2][26:],
#                               Value[3][1:26], Value[3][26:]])
# }
# matplotlib.rcParams['font.family'] = ["Hiragino Sans GB", "DejaVu Sans"]  # 中文字体
# # 创建DataFrame
# df = pd.DataFrame(data)
# # custom_palette = {'ims': '#1f77b4', 'dms': '#2ca02c'}
# custom_palette = {'ims': '#8e44ad', 'dms': '#27ae60'}
# # 定义额外的数据点
# extra_data = {'TOP-1': Value[0][0], 'TOP-3': Value[1][0], 'TOP-5': Value[2][0], 'TOP-10': Value[3][0]}
# # # 创建图形
# # font_chinese = FontProperties(fname='./fonts/SimSun.ttf')
# font_english = FontProperties(fname='/System/Library/Fonts/Supplemental/Times New Roman.ttf')
# italic_font = FontProperties(fname='/System/Library/Fonts/Supplemental/Times New Roman Italic.ttf')
# # font_chinese.set_size(15)  # 这里设置字体大小
# font_english.set_size(13)  # 这里设置字体大小
# italic_font.set_size(13)  # 这里设置字体大小
#
# # 横坐标：从 100 到 50 共 26 个点
# x_weight = np.linspace(1, 0.5, 26)
#
# # 颜色与样式
# color_ims = '#8e44ad'
# color_dms = '#27ae60'
# ims_marker_style = dict(marker='^', markersize=5)
# dms_marker_style = dict(marker='s', markersize=5)
#
# # 标签
# top_labels = ['TOP-1', 'TOP-3', 'TOP-5', 'TOP-10']
#
# # 创建 2x2 子图
# fig, axes = plt.subplots(2, 2, figsize=(10, 8))
#
# for idx in range(4):
#     row = idx // 2
#     col = idx % 2
#     ax = axes[row][col]
#
#     # 构造曲线数据（第一个点为 oms）
#     y_ims = [Value[idx][0]] + list(Value[idx][1:26])   # oms + ims
#     y_dms = [Value[idx][0]] + list(Value[idx][26:])    # oms + dms
#
#     # 画 ims 和 dms 曲线
#     ax.plot(x_weight, y_ims, label='ims', color=color_ims, **ims_marker_style)
#     ax.plot(x_weight, y_dms, label='dms', color=color_dms, **dms_marker_style)
#
#     # 红点：oms，位于 x=100 处
#     ax.plot([1], [Value[idx][0]], 'ro', label='oms', markersize=7)
#
#     # 设置图标题和轴
#     ax.set_title(top_labels[idx], fontproperties=font_english)
#     ax.set_xlabel("Weight", fontproperties=font_english)
#     ax.set_ylabel("Value", fontproperties=font_english)
#
#     # 设置图例，单独调整 TOP-3 的图例位置
#     # if idx == 1:  # TOP-3
#     #     ax.legend(
#     #         prop=italic_font,
#     #         loc='upper left',
#     #         bbox_to_anchor=(0, 0.55),  # 调低一些
#     #     )
#     # else:
#     #     ax.legend(prop=italic_font)
#     ax.legend(prop=italic_font)
#     # 设置坐标轴刻度字体
#     for label in ax.get_xticklabels():
#         label.set_fontproperties(font_english)
#     for label in ax.get_yticklabels():
#         label.set_fontproperties(font_english)
#
# # 优化布局
# plt.tight_layout()
# plt.show()
# plt.savefig('./figures/output.pdf'
#             "提交/修改/TOSEM/提交/修改/RQ3曲线-one.pdf", format="pdf", bbox_inches='tight')


# 折线图，对比ims和dms
project = 'EMS'
path = '../../data/results/resultEXAM.xlsx'  # '+sys.argv[1][:-1]+'
wb = load_workbook(path)
ws = wb['Sheet1']
max_row = ws.max_row
Value = []
for i in range(2, 3):
    x = 2
    a = []
    while True:
        if x == 3:
            x += 1
            continue
        value = ws.cell(x, i).value
        a.append(value)
        x += 1
        if x > max_row:
            break
    Value.append(a)
Value = np.array(Value)
data = {
    'Category': ['EXAM'] * 50,
    'Group': ['ims'] * 25 + ['dms'] * 25,
    'Values': np.concatenate([Value[0][1:26], Value[0][26:]])
}
# matplotlib.rcParams['font.family'] = ["Hiragino Sans GB", "DejaVu Sans"]  # 中文字体
# 创建DataFrame
df = pd.DataFrame(data)
# custom_palette = {'ims': '#1f77b4', 'dms': '#2ca02c'}
custom_palette = {'ims': '#8e44ad', 'dms': '#27ae60'}
# 定义额外的数据点
extra_data = {'EXAM': Value[0][0]}
# # 创建图形
# font_chinese = FontProperties(fname='./fonts/SimSun.ttf')
font_english = FontProperties(fname='/System/Library/Fonts/Supplemental/Times New Roman.ttf')
italic_font = FontProperties(fname='/System/Library/Fonts/Supplemental/Times New Roman Italic.ttf')
# font_chinese.set_size(15)  # 这里设置字体大小
font_english.set_size(13)  # 这里设置字体大小
italic_font.set_size(13)  # 这里设置字体大小

# 横坐标：从 100 到 50 共 26 个点
x_weight = np.linspace(1, 0.5, 26)

# 颜色与样式
color_ims = '#8e44ad'
color_dms = '#27ae60'
ims_marker_style = dict(marker='^', markersize=5)
dms_marker_style = dict(marker='s', markersize=5)

# 标签
top_labels = ['EXAM']

# 创建 2x2 子图
fig, ax = plt.subplots(figsize=(10, 8))

# 构造曲线数据（第一个点为 oms）
y_ims = [Value[0][0]] + list(Value[0][1:26])   # oms + ims
y_dms = [Value[0][0]] + list(Value[0][26:])    # oms + dms
# 画 ims 和 dms 曲线
ax.plot(x_weight, y_ims,

        label='ims',

        color=color_ims,

        **ims_marker_style)

ax.plot(x_weight, y_dms,

        label='dms',

        color=color_dms,

        **dms_marker_style)

# 红点：oms

ax.plot([1], [Value[0][0]],

        'ro',

        label='oms',

        markersize=7)

# 标题和坐标轴

ax.set_title('EXAM', fontproperties=font_english)

ax.set_xlabel("Weight", fontproperties=font_english)

ax.set_ylabel("Value", fontproperties=font_english)

# 图例

ax.legend(prop=italic_font)

# 坐标轴字体

for label in ax.get_xticklabels():

    label.set_fontproperties(font_english)

for label in ax.get_yticklabels():

    label.set_fontproperties(font_english)

# 优化布局
plt.tight_layout()
plt.show()
plt.savefig('./figures/output.pdf'
            "TOSEM/提交/修改/IST/提交/修改/EXAM.pdf", format="pdf", bbox_inches='tight')

