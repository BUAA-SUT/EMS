import os
import json
import coverage
from DM import *
from publicFun import *
from openpyxl import load_workbook
from Original import *
from Mutant1 import *
from Mutant2 import *
from Mutant3 import *
from Mutant4 import *
from Mutant5 import *
from Mutant6 import *
from Mutant7 import *
random.seed(1)


# def getOriginalInput():
#     source_case_set = []
#     ty = random.randint(1, 5)
#     t = 10
#     while 1:
#         matrix = []
#         if ty == 1:
#             while len(matrix) < 9:
#                 matrix.append(random.randint(1, t))  # 主对角线非0
#                 matrix.append(random.randint(0, t))
#                 matrix.append(random.randint(0, t))
#                 matrix.append(0)
#                 matrix.append(random.randint(0, t))
#                 matrix.append(random.randint(0, t))
#                 matrix.append(0)
#                 matrix.append(0)
#                 matrix.append(random.randint(0, t))
#             ty = random.randint(1, 5)
#         elif ty == 2:
#             while len(matrix) < 9:
#                 matrix.append(random.randint(0, t))
#                 matrix.append(random.randint(0, t))
#                 matrix.append(random.randint(1, t))  # 次对角线非0
#                 matrix.append(random.randint(0, t))
#                 matrix.append(random.randint(0, t))
#                 matrix.append(0)
#                 matrix.append(random.randint(0, t))
#                 matrix.append(0)
#                 matrix.append(0)
#             ty = random.randint(1, 5)
#         else:
#             while len(matrix) < 9:
#                 matrix.append(random.randint(0, t))
#             ty = random.randint(1, 5)
#         source_case_set.append(matrix)
#         if len(source_case_set) >= 1000:
#             break
#
#     # 随机取100个测试用例
#     random_input = random.sample(source_case_set, 100)
#     data = {
#             'source_case_set': source_case_set,
#             'random_input': random_input
#     }
#     json_str = json.dumps(data)
#     with open('/Applications/work/data/MT/FS/' + string + '/OriginalInputNew.json', 'w') as f:
#         json.dump(json_str, f)
#     return source_case_set, random_input
#
#
# def FailureRate(dynamic):
#     # 把SourceCases读出来
#     Result = []
#     with open('/Applications/work/data/MT/MFT/'+string+'/OriginalInputNew.json', 'r') as load_f:
#         data = json.load(load_f)
#     data = json.loads(data)
#     source_case_set = data['source_case_set']
#     for i in range(len(source_case_set)):
#         result_s_a = DeterMinant().Determinant(source_case_set[i], n)  # oracle
#         result_s_m = dynamic.Determinant(source_case_set[i], n)
#         if result_s_a[0] == result_s_m[0]:
#             Result.append(0)
#         else:
#             Result.append(1)
#     FR = round(Result.count(1) / len(Result) * 100, 2)
#     return FR


# def riskIndex(argv, dynamic):
#
#     MGS1 = []  # 原始
#     MGS2 = []  # 新增
#     Result1 = []
#     Result2 = []
#     testcase1 = []  # 原始
#     testcase2 = []  # 新增
#     follow1 = []  # 数字代表层数的大小
#     source_case = argv.copy()
#     testcase1.append(source_case)
#     MG, follow_case1 = MTG(source_case, dynamic)  # t1t2t3t4
#     for i in range(len(follow_case1)):
#         testcase1.append(follow_case1[i])
#     MGS1.append(MG)
#     for i in range(len(follow_case1)):  # t2t3t4
#         MG, follow_case2 = MTG(follow_case1[i], dynamic)  # t2t5...
#         follow1.append(follow_case2)  # [t5t6t7], [t8t9t10],...
#         MGS1.append(MG)
#         for j in range(len(follow_case2)):  # t5t6t7, t8t9t10,...
#             testcase1.append(follow_case2[j])  # t1t2t3t4t5t6t7t8...
#
#     for i in range(len(follow1)):
#         mgs1 = []
#         ts1 = []
#         for j in range(len(follow1[i])):
#             mgs2 = []
#             ts2 = [follow1[i][j]]
#             MG, follow_case3 = MTG(follow1[i][j], dynamic)  # t5t14...
#             mgs2.append(MG)
#             follow2 = []
#             for k in range(len(follow_case3)):  # t14t15t16
#                 MG, follow_case4 = MTG(follow_case3[k], dynamic)  # t14t17...
#                 follow2.append(follow_case4)
#                 mgs2.append(MG)
#                 ts2.append(follow_case3[k])  # t14t15t16
#
#             mgs1.append(mgs2)
#
#             for n in range(len(follow2)):
#                 for m in range(len(follow2[n])):
#                     ts2.append(follow2[n][m])  # # t17t18t19...
#             ts1.append(ts2)
#
#         MGS2.append(mgs1)
#         testcase2.append(ts1)
#
#     # MG统计完, testcase统计完
#     n = 3
#     for i in range(len(testcase1)):
#         result_s_a = DeterMinant().Determinant(testcase1[i], n)  # oracle
#         result_s_m = dynamic.Determinant(testcase1[i], n)
#         if result_s_a[0] == result_s_m[0]:
#             Result1.append(0)
#         else:
#             Result1.append(1)
#
#     # 去掉巧合满足性
#     for i in range(len(MGS1)):
#         for j in range(len(MGS1[i])):
#             if MGS1[i][j] == 0 and (Result1[i] or Result1[i * len(MGS1[0]) + j + 1]):  # 如果satisfied
#                 MGS1[i][j] = 3
#
#     # MG统计完, testcase统计完
#     for i in range(len(testcase2)):
#         result1 = []
#         for j in range(len(testcase2[i])):
#             result2 = []
#             for k in range(len(testcase2[i][j])):
#                 result_s_a = DeterMinant().Determinant(testcase2[i][j][k], n)  # oracle
#                 result_s_m = dynamic.Determinant(testcase2[i][j][k], n)
#                 if result_s_a[0] == result_s_m[0]:
#                     result2.append(0)
#                 else:
#                     result2.append(1)
#             result1.append(result2)
#         Result2.append(result1)
#
#     # 去掉巧合满足性
#     for i in range(len(MGS2)):
#         for j in range(len(MGS2[i])):
#             for k in range(len(MGS2[i][j])):
#                 for m in range(len(MGS2[i][j][k])):
#                     if MGS2[i][j][k][m] == 0 and (Result2[i][j][k] or Result2[i][j][k * len(MGS1[0]) + m + 1]):  # 如果satisfied
#                         MGS2[i][j][k][m] = 3
#
#     MGS = [MGS1, MGS2]
#     Result = [Result1, Result2]
#
#     SMGS = copy.deepcopy(MGS)
#
#     # 随机去除一些MG
#     for i in range(1, len(MGS1)):  # 第一组不变
#         t = random.randint(1, len(MGS1[i])-1)  # 去几个
#         a = [n for n in range(len(MGS1[i]))]
#         random.shuffle(a)
#         b = a[:t]
#         for j in b:
#             MGS1[i][j] = 4
#
#     return MGS, Result, SMGS

def getInput(argv, dynamic):
    testcase = []  # 原始
    follow1 = []  # 数字代表层数的大小
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case1 = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case1)):
        testcase.append(follow_case1[i])
    for i in range(len(follow_case1)):  # t2t3t4
        MG, follow_case2 = MTG(follow_case1[i], dynamic)  # t2t5...
        follow1.append(follow_case2)  # [t5t6t7], [t8t9t10],...
        for j in range(len(follow_case2)):  # t5t6t7, t8t9t10,...
            testcase.append(follow_case2[j])  # t1t2t3t4t5t6t7t8...
    return testcase


def statements(argv, mu):
    testcase = argv.copy()
    filename = 'Mutant{}.py'.format(mu)
    Exelines = []
    executable = []
    for i in range(len(testcase)):
        cov = coverage.coverage()
        cov.start()
        result_s_a = eval('Mutant{}()'.format(mu)).Determinant(testcase[i], n)  # oracle
        cov.stop()
        numlist = cov.analysis(filename)
        executable = numlist[1]
        exelist = list(set(numlist[1]) - set(numlist[2]))
        exelist.sort()
        Exelines.append(exelist)
    return Exelines, executable


def riskIndex(argv, dynamic):
    MGS = []  # 原始
    Result = []
    testcase = []  # 原始
    follow1 = []  # 数字代表层数的大小
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case1 = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case1)):
        testcase.append(follow_case1[i])
    MGS.append(MG)
    for i in range(len(follow_case1)):  # t2t3t4
        MG, follow_case2 = MTG(follow_case1[i], dynamic)  # t2t5...
        follow1.append(follow_case2)  # [t5t6t7], [t8t9t10],...
        MGS.append(MG)
        for j in range(len(follow_case2)):  # t5t6t7, t8t9t10,...
            testcase.append(follow_case2[j])  # t1t2t3t4t5t6t7t8...

    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        result_s_a = DeterMinant().Determinant(testcase[i], n)  # oracle
        result_s_m = dynamic.Determinant(testcase[i], n)
        if result_s_a[0] == result_s_m[0]:
            Result.append(0)
        else:
            Result.append(1)

    # 去掉巧合满足性
    for i in range(len(MGS)):
        for j in range(len(MGS[i])):
            if MGS[i][j] == 0 and (Result[i] or Result[i * len(MGS[0]) + j + 1]):  # 如果satisfied
                MGS[i][j] = 3

    return MGS, Result


def classify_second_level_lists(list_of_lists):
    # 创建一个字典，用于存储相同的第二级列表及其对应的索引
    classified_lists = {}

    # 遍历列表，获取第二级列表
    for index, sublist in enumerate(list_of_lists):
        # 将第二级列表转换为元组，以便于作为字典的键
        tuple_sublist = tuple(sublist)

        # 检查字典中是否已经存在该键
        if tuple_sublist in classified_lists:
            # 如果存在，则将当前索引添加到对应的值列表中
            classified_lists[tuple_sublist].append(index)
        else:
            # 如果不存在，则创建一个新的键值对
            classified_lists[tuple_sublist] = [index]

    return classified_lists


if __name__ == '__main__':
    project = 'STVR'
    row = 1
    string = 'DM'
    path = '/Applications/work/data/MT/' + project + '/Result/result25_one.xlsx'  # '+sys.argv[1][:-1]+'
    wb = load_workbook(path)
    # source_case_set, random_input = getOriginalInput()
    if string not in wb.sheetnames:
        ws = wb.create_sheet(string)
    del wb[string]
    ws = wb.create_sheet(string)
    MG_set = []
    Sus_set = []
    Metric_set = []
    Union_set = []
    Unique_set = []
    Flag_set = []
    # with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/'+string+'/OriginalInputNew.json', 'r') as load_f:
    #     data = json.load(load_f)
    # data = json.loads(data)
    # random_input = data['random_input']
    # testcases = []
    # for i in random_input:
    #     testcases.append(getInput(i, DeterMinant()))
    # data = {
    #     'testcases': testcases
    # }
    # json_str = json.dumps(data)
    # with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/Input.json', 'w') as f:
    #     json.dump(json_str, f)
    # with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/Input.json', 'r') as f:
    #     data = json.load(f)
    # data = json.loads(data)
    # testcases = data['testcases']
    # # ExelinesS = []
    # # for i in testcases:
    # #     Exelines, Executable = statements(i)
    # #     ExelinesS.append(Exelines)
    # # data = {
    # #     'Exec': Executable, 'Exel': ExelinesS
    # # }
    # # json_str = json.dumps(data)
    # # with open('/Applications/work/data/MT/'+project+'/' + string + '/statements.json', 'w') as f:
    # #     json.dump(json_str, f)
    # # with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/statements.json', 'r') as f:
    # #     data = json.load(f)
    # # data = json.loads(data)
    # # ExelinesS = data['Exel']
    # # Executable = data['Exec']
    # # inters = set(ExelinesS[0][0])
    # # # 逐个计算交集
    # # for i in range(len(ExelinesS)):
    # #     for j in range(len(ExelinesS[0])):
    # #         inters = inters.intersection(set(ExelinesS[i][j]))
    # #     # print(i, len(list(inters)))
    # # # 将结果转换为列表
    # # inters = list(inters)
    # # inters.sort()
    # # for i in range(len(ExelinesS)):
    # #     for j in range(len(ExelinesS[0])):
    # #         ExelinesS[i][j] = [x for x in ExelinesS[i][j] if x not in inters]
    # # removemg = []
    # # for i in range(len(ExelinesS)):
    # #     a = 0
    # #     for j in range(len(ExelinesS[i])):
    # #         if len(ExelinesS[i][j]) == 0:
    # #             a += 1
    # #     if a == len(ExelinesS[i]):
    # #        removemg.append(i)
    # # for index in sorted(removemg, reverse=True):
    # #     del ExelinesS[index]
    # # removemu = []
    # # for mu in range(1, 8):
    # #     with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
    # #               'r') as load_f:
    # #         data = json.load(load_f)
    # #     data = json.loads(data)
    # #     Flag = data['Flag']
    # #     fault = Executable[Flag.index(1)]
    # #     if fault in inters:
    # #         removemu.append(mu)
    # for mu in range(1,  8):
    #     ExelineS = []
    #     Executable = []
    #     for i in testcases:
    #         Exelines, Executable = statements(i, mu)
    #         ExelineS.append(Exelines)
    #     data = {
    #         'Exec': Executable, 'Exel': ExelineS
    #     }
    #     json_str = json.dumps(data)
    #     with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/statements{}.json'.format(mu), 'w') as f:
    #         json.dump(json_str, f)
    for mu in range(1, 8):  # 1, 8
        # dynamic = eval("Mutant" + str(mu))()
        # Result = []
        # MGS = []
        # for i in range(100):
        #     MG, result = riskIndex(random_input[i], dynamic)
        #     MGS.append(MG)
        #     Result.append(result)
        # Flag = [0] * len(Executable)
        # if mu == 1:
        #     Flag[18] = 1
        # elif mu == 2:
        #     Flag[21] = 1
        # elif mu == 3:
        #     Flag[18] = 1
        # elif mu == 4:
        #     Flag[18] = 1
        # elif mu == 5:
        #     Flag[23] = 1
        # elif mu == 6:
        #     Flag[34] = 1
        # elif mu == 7:
        #     Flag[34] = 1
        with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/statements{}.json'.format(mu), 'r') as f:
            data = json.load(f)
        data = json.loads(data)
        ExelineS = data['Exel']
        Executable = data['Exec']
        # with open('/Applications/work/data/MT/'+project+'/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # SMGS = data['SMGS']
        # Flag = data['Flag']
        # if mu in removemu:
        #     continue
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                  'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        # mSus = data['mSus']
        # mSus2 = data['mSus2']
        # mSus3 = data['mSus3']
        # mSus4 = data['mSus4']
        # metric = data['metric']
        # metric2 = data['metric2']
        # metric3 = data['metric3']
        # metric4 = data['metric4']
        # mSus_nofs = data['mSus_nofs']
        # metric_nofs = data['metric_nofs']
        # mSus4_nofs = data['mSus4_nofs']
        # metric4_nofs = data['metric4_nofs']
        Flag = data['Flag']
        # Flag = [0] * len(Executable)
        # if mu == 1:
        #     Flag[18] = 1
        # elif mu == 2:
        #     Flag[21] = 1
        # elif mu == 3:
        #     Flag[18] = 1
        # elif mu == 4:
        #     Flag[18] = 1
        # elif mu == 5:
        #     Flag[24] = 1
        # elif mu == 6:
        #     Flag[36] = 1
        # elif mu == 7:
        #     Flag[36] = 1
        # data['Flag'] = Flag
        MGS = data['MGS']
        Result = data['Result']
        mSus4 = data['mSus4_one']
        mSus = data['mSus_one']
        Para = data['Para']
        # metric4all = data['metric4_all']
        # metric4 = data['metric4']
        # metric4one = data['metric4_one']
        # metricall = data['metric_all']
        # metricone = data['metric_one']
        # metric = data['metric']
        # Para = data['Para']
        # mSus4_all = data['mSus4_all']
        # mSus_all = data['mSus_all']
        # mSus4_one = data['mSus4_one']
        # mSus_one = data['mSus_one']
        # AllF = data['AllF']
        # Sus = []
        # Metric = []
        # for i in range(len(mSus)):
        #     a = [round(mSus[i], 4)]
        #     b = [metric[i]]
        #     for j in range(len(mSus3)):
        #         a.append(mSus3[j][5][i])
        #         b.append(metric3[j][i])
        #     Sus.append(a)
        #     Metric.append(b)
        # Sus_set.append(Sus)
        # Metric_set.append(Metric)
        # mSus_nofs, metric_nofs = MSlice(MGS, Executable, ExelineS)
        # data['mSus_nofs'] = mSus_nofs
        # data['metric_nofs'] = metric_nofs
        # mSus, metric = MSlice_one(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus_one'] = mSus
        # data['metric_one'] = metric
        # mSus, metric = MSlice_all(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus_all'] = mSus
        # data['metric_all'] = metric
        # mSus4, _, metric4 = MSlice4_one(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus4_one'] = mSus4
        # data['metric4_one'] = metric4
        # data['Para'] = Para
        # mSus4, _, metric4 = MSlice4_all(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus4_all'] = mSus4
        # data['metric4_all'] = metric4
        # mSus3, AllF, _ = MSlice3(MGS, Executable, ExelinesS, Flag)
        # Union_set.append(Union)
        # Unique_set.append(Unique)
        # mSus4_nofs, Para, metric4_nofs = MSlice4(MGS, Executable, ExelineS, Flag, Result)
        # data['Para'] = Para
        # data['mSus4_nofs'] = mSus4_nofs
        # data['metric4_nofs'] = metric4_nofs
        # data = {
        #          'mSus': mSus, 'mSus4': mSus4, 'metric': metric,
        #     'metric4': metric4, 'Para': Para, 'MGS': MGS, 'Flag': Flag,
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'w') as f:
        #     json.dump(json_str, f)
        row = eval('getMetrics_5')(row, ws, mu, MGS, mSus, mSus4, Para[-1], Flag, Para, ExelineS)
    wb.save(path)
