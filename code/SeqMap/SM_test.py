import subprocess
from openpyxl import load_workbook
from publicFun import *
import csv
import json
import os

# 自动切换工作目录为脚本所在目录
os.chdir(os.path.dirname(os.path.abspath(__file__)))

def run_seqmap(mu, e_value: int, probes_file: str, trans_file: str, output_file: str) -> None:
    exec_path = f'Mutants/seqmap_v{mu}/seqmap'
    result = subprocess.run(
        [exec_path, str(e_value), probes_file, trans_file, output_file],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        timeout=30  # 设置超时防止死锁
    )

    if result.returncode != 0:
        with open(f"{output_file}", "w") as log_file:
            log_file.write(f"运行失败：returncode={result.returncode}\n")
            log_file.write(f"stderr:\n{result.stderr}\n")
            log_file.write(f"stdout:\n{result.stdout}\n")
        print(f"seqmap_v{mu}{output_file}运行失败，已记录日志。")

# project = 'EMS'
# string = 'SeqMap'
# source_path = '../../data/MT/'+project+'/' + string + '/source/'
# source_e_path = source_path + "e"
# source_p_path = source_path + "p"
# source_T_path = source_path + "T"
# follow_path = '../../data/MT/'+project+'/' + string + '/follow/'
# follow_e_path = follow_path + "e"
# follow_p_path = follow_path + "p"
# follow_T_path = follow_path + "T"
# num_nu = 5
# for mu in range(num_nu+1):
#     if mu  == 0 or mu == 1 or mu == 2 or mu == 4:
#         continue
#     with open(source_e_path, "r") as fin:
#         for line in fin:
#             line = line.strip()
#             if not line:
#                 continue
#             parts = line.split()
#             if len(parts) != 3:
#                 print(f"格式错误：{line}")
#                 continue
#             _, x_str, id = parts
#             # if os.path.exists(source_path + "output_{}_{}".format(mu, int(id))):
#             #     continue
#             e = int(x_str)
#             run_seqmap(mu, e, source_T_path+"/{}".format(int(id)), source_p_path+"/{}".format(int(id)),
#                        source_path + "output_{}_{}".format(mu, int(id)))
#
# for mu in range(num_nu+1):
#     if mu  == 0:
#         continue
#     with open(follow_e_path, "r") as fin:
#         for line in fin:
#             line = line.strip()
#             if not line:
#                 continue
#             parts = line.split()
#             if len(parts) != 3:
#                 print(f"格式错误：{line}")
#                 continue
#             f, x_str, id = parts
#             mr = f.split('_', 1)[1]
#             # if os.path.exists(follow_path + "output_{}_{}_".format(mu, int(id))+mr):
#             #     continue
#             e = int(x_str)
#             run_seqmap(mu, e, follow_T_path+"/{}_".format(int(id))+mr, follow_p_path+"/{}_".format(int(id))+mr,
#                        follow_path + "output_{}_{}_".format(mu, int(id))+mr)

if __name__ == "__main__":
    num_case = 100
    num_mr = 3
    num_mu = 5
    project = 'EMS'
    string = 'SeqMap'
    # title = [30, 31, 32, 33]
    # N = [1, 3, 5, 10]
    for i in range(1):
        row = 1
        path = '../../data/results/resultEXAM.xlsx'
        # source_path = '../../data/MT/' + project + '/' + string + '/source/'
        # source_e_path = source_path + "e"
        # follow_path = '../../data/MT/' + project + '/' + string + '/follow/'
        wb = load_workbook(path)
        if string in wb.sheetnames:
            del wb[string]
        ws = wb.create_sheet(string)
        # Id = []
        # with open(source_e_path, "r") as fin:
        #     for line in fin:
        #         line = line.strip()
        #         if not line:
        #             continue
        #         parts = line.split()
        #         if len(parts) != 3:
        #             print(f"格式错误：{line}")
        #             continue
        #         _, x_str, id = parts
        #         Id.append(int(id))
        for mu in range(1, num_mu+1):
            if mu == 3:
                continue
            # datadir = source_path + 'statementsource_v{}.csv'.format(mu)
            # csvFile = open(datadir, "r")
            # reader = csv.reader(csvFile)
            # # 建立空字典
            # sourcecov = {}
            # sourceresult = {}
            # for item in reader:
            #     sourcecov[item[0]] = item[1:-1]
            #     if reader.line_num == 1:
            #         continue
            #     sourceresult[item[0]] = item[-1:]
            # csvFile.close()
            # datadir = follow_path + 'statementfollow_v{}.csv'.format(mu)
            # csvFile = open(datadir, "r")
            # reader = csv.reader(csvFile)
            # # 建立空字典
            # followcov = {}
            # followresult = {}
            # for item in reader:
            #     followcov[item[0]] = item[1:-1]
            #     if reader.line_num == 1:
            #         continue
            #     followresult[item[0]] = item[-1:]
            # csvFile.close()
            # ExecutableS = sourcecov.get('inputs')
            # ExecutableS = [int(x) for x in ExecutableS]
            # ExelineS = []
            # Result = []
            # for i in range(len(Id)):
            #     Exelines = []
            #     r = []
            #     candidate = sourcecov.get('input{}'.format(Id[i]))
            #     candidate = [int(x) for x in candidate]
            #     indices = [i for i in range(len(candidate)) if candidate[i] == 1]
            #     e = [ExecutableS[i] for i in indices]
            #     Exelines.append([e])
            #     a = sourceresult.get('input{}'.format(Id[i]))
            #     a = [int(x) for x in a]
            #     r.append([a[0]])
            #     for j in range(num_mr):
            #         candidate = followcov.get('input{}_{}'.format(Id[i], j+1))
            #         candidate = [int(x) for x in candidate]
            #         indices = [i for i in range(len(candidate)) if candidate[i] == 1]
            #         e = [ExecutableS[i] for i in indices]
            #         Exelines.append([e])
            #         a = followresult.get('input{}_{}'.format(Id[i], j+1))
            #         a = [int(x) for x in a]
            #         r.append([a[0]])
            #     Result.append(r)
            #     ExelineS.append(Exelines)
            # data = {
            #     'Exec': ExecutableS, 'Exel': ExelineS, 'Result': Result
            # }
            # json_str = json.dumps(data)
            # with open('../../data/MT/' + project + '/' + string + '/statements{}.json'.format(mu), 'w') as f:
            #     json.dump(json_str, f)
            # with open('../../data/MT/'+project+'/' + string + '/statements{}.json'.format(mu), 'r') as f:
            #     data = json.load(f)
            # data = json.loads(data)
            # ExelineS = data['Exel']
            # ExecutableS = data['Exec']
            # Result = data['Result']
            with open('../../data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                      'r') as load_f:
                data = json.load(load_f)
            data = json.loads(data)
            # MGS = data['MGS']
        #     # ExecutableS = data['Exec']
        #     # ExelinesS = data['Exel']
        #     Flag = [0] * len(ExecutableS)
        #     raw_indices = {1: 572,2: 572,4: 574,5: 574}
        #     mu_to_flag_index = {k: ExecutableS.index(v) for k, v in raw_indices.items()}
        #     if mu in mu_to_flag_index:
        #         Flag[mu_to_flag_index[mu]] = 1
        #     data['Flag'] = Flag
        # #     with open('../../data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        # #               'r') as load_f:
        # #         data = json.load(load_f)
        # #     data = json.loads(data)
        # #     mSus = data['mSus']
        # #     mSus4 = data['mSus4']
        # #     # metric = data['metric']
        # #     # metric2 = data['metric2']
        # #     # metric3 = data['metric3']
        # #     # metric4 = data['metric4']
            Flag = data['Flag']
            # MGS = data['MGS']
            # mSus4 = data['mSus4_nofs']
            # mSus = data['mSus_nofs']
            sbflsus = data['sbflsus']
            # Para = data['Para']
            # if i == 0:
            #     mSus, metric = MSlice_PC_one(MGS, ExecutableS, ExelineS, Flag, Result)
            #     mSus4, _, metric4 = MSlice4_PC_one(MGS, ExecutableS, ExelineS, Flag, Result)
            #     # sbflsus, sbflmetric = SBFL(MGS, Result, Executable, ExelineS)
            #     data['mSus_one'] = mSus
            #     data['metric_one'] = metric
            #     data['mSus4_one'] = mSus4
            #     data['metric4_one'] = metric4
            #     # data['sbflsus'] = sbflsus
            #     # data['sbflmetric'] = sbflmetric
            #     # Union_set.append(Union)
            #     # Unique_set.append(Unique)
            #     # mSus4, _, metric4 = MSlice4_one(MGS, Executable, ExelineS, Flag, Result)
            #     # data['mSus4_one'] = mSus4
            #     # data['metric4_one'] = metric4
            #     # data['Para'] = Para
            #     # mSus4, _, metric4 = MSlice4_all(MGS, Executable, ExelineS, Flag, Result)
            #     # data['mSus4_all'] = mSus4
            #     # data['metric4_all'] = metric4
            #     # data = {
            #     #          'mSus': mSus, 'mSus4': mSus4, 'metric': metric,
            #     #     'metric4': metric4, 'Para': Para, 'MGS': MGS, 'Flag': Flag,
            #     # }
            #     json_str = json.dumps(data)
            #     with open('../../data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
            #               'w') as f:
            #         json.dump(json_str, f)
            # else:
            mSus4 = data['mSus4']
            mSus = data['mSus']
            # row = eval('getMetrics_5')(row, ws, mu, MGS, mSus, mSus4,sbflsus, Para[-1], Flag, Para, ExelineS, N[i])
            row = eval('getMetrics_6')(row, ws, mu, mSus, mSus4, sbflsus, Flag)
        wb.save(path)
