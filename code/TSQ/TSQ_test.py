import os
# import sys
# package_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# sys.path.append(package_path)
from TSQ import *
import coverage
# import numpy as np
# print(coverage.__file__)
import json
from publicFun import *
from Original import *
from Mutant1 import *
from Mutant2 import *
from Mutant3 import *
from Mutant4 import *
from Mutant5 import *
from Mutant6 import *
from openpyxl import load_workbook
# import xlwt
random.seed(1)


# def getOriginalInput():
#     source_case_set = []
#     while 1:
#         # a = random.randint(1, 10)
#         # b = random.randint(1, 10)
#         # c = random.randint(1, 10)
#         a = round(random.uniform(1.0, 10.0), 2)
#         b = round(random.uniform(1.0, 10.0), 2)
#         c = round(random.uniform(1.0, 10.0), 2)
#         if a >= b + c or b >= a + c or c >= a + b:
#             continue
#         else:
#             triangle = [a, b, c]
#         source_case_set.append(triangle)
#         if len(source_case_set) >= 1000:
#             break
#     # 随机取100个测试用例
#     random_input = random.sample(source_case_set, 100)
#     data = {
#             'source_case_set': source_case_set,
#             'random_input': random_input
#     }
#     json_str = json.dumps(data)
#     with open('/Applications/work/data/MT/FS/'+string2+'/OriginalInputNew.json', 'w') as f:
#         json.dump(json_str, f)
#
#     return source_case_set, random_input
#
#
# def FailureRate(dynamic):
#     # 把SourceCases读出来
#     Result = []
#     with open('/Applications/work/data/MT/MFT/'+string+'/OriginalInputNew.json', 'r') as load_f:
#         data = json.load(load_f)
#     data = json.loads(data)
#     source_case_set = data['source_case_set']
#     for i in range(len(source_case_set)):
#         result_s_a = Trisquare().trisquare(source_case_set[i])  # oracle
#         result_s_m = dynamic.trisquare(source_case_set[i])
#         if result_s_a[0] == result_s_m[0]:
#             Result.append(0)
#         else:
#             Result.append(1)
#     FR = round(Result.count(1) / len(Result) * 100, 2)
#     return FR
#
#
# def getResult():
#     Result = []
#     rate = []
#     # while 1:
#     #     a = random.randint(1, 10)
#     #     b = random.randint(1, 10)
#     #     c = random.randint(1, 10)
#     #     if a >= b + c or b >= a + c or c >= a + b:
#     #         continue
#     #     else:
#     #         triangle = [a, b, c]
#     #
#     #         result = Trisquare().trisquare2(triangle)
#     #         Result.append(result)
#     #         if len(Result) >= 1000:
#     #             break
#     with open('/Applications/work/data/MT/MFT/'+string+'/OriginalInputNew.json', 'r') as load_f:
#         data = json.load(load_f)
#     data = json.loads(data)
#     source_case_set = data['source_case_set']
#     for i in range(len(source_case_set)):
#         result = Trisquare().trisquare2(source_case_set[i])  # oracle
#         Result.append(result)
#     rate1 = Result.count(1) / len(Result) * 100
#     rate2 = Result.count(2) / len(Result) * 100
#     rate3 = Result.count(3) / len(Result) * 100
#     rate4 = Result.count(4) / len(Result) * 100
#     rate5 = Result.count(5) / len(Result) * 100
#     rate6 = Result.count(6) / len(Result) * 100
#     rate.append([rate1, rate2, rate3, rate4, rate5, rate6])
#     return rate


def getInput(argv, dynamic):
    testcase = []  # 原始
    follow1 = []  # 数字代表层数的大小
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case1 = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case1)):
        testcase.append(follow_case1[i])
    for i in range(len(follow_case1)):  # t2t3t4
        MG, follow_case2 = MTG(follow_case1[i], dynamic)  # t2t5...
        follow1.append(follow_case2)  # [t5t6t7], [t8t9t10],...
        for j in range(len(follow_case2)):  # t5t6t7, t8t9t10,...
            testcase.append(follow_case2[j])  # t1t2t3t4t5t6t7t8...
    return testcase


def statements(argv, mu):
    testcase = argv.copy()
    filename = 'Mutant{}.py'.format(mu)
    Exelines = []
    executable = []
    for i in range(len(testcase)):
        cov = coverage.coverage()
        cov.start()
        result_s_a = eval('Mutant{}()'.format(mu)).trisquare(testcase[i])
        cov.stop()
        numlist = cov.analysis(filename)
        executable = numlist[1]
        exelist = list(set(numlist[1]) - set(numlist[2]))
        exelist.sort()
        Exelines.append(exelist)
    return Exelines, executable


def riskIndex(argv, dynamic):
    MGS = []  # 原始
    Result = []
    testcase = []  # 原始
    follow1 = []  # 数字代表层数的大小
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case1 = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case1)):
        testcase.append(follow_case1[i])
    MGS.append(MG)
    for i in range(len(follow_case1)):  # t2t3t4
        MG, follow_case2 = MTG(follow_case1[i], dynamic)  # t2t5...
        follow1.append(follow_case2)  # [t5t6t7], [t8t9t10],...
        MGS.append(MG)
        for j in range(len(follow_case2)):  # t5t6t7, t8t9t10,...
            testcase.append(follow_case2[j])  # t1t2t3t4t5t6t7t8...

    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        result_s_a = Trisquare().trisquare(testcase[i])  # oracle
        result_s_m = dynamic.trisquare(testcase[i])
        if result_s_a[0] == result_s_m[0]:
            Result.append(0)
        else:
            Result.append(1)

    # 去掉巧合满足性
    for i in range(len(MGS)):
        for j in range(len(MGS[i])):
            if MGS[i][j] == 0 and (Result[i] or Result[i * len(MGS[0]) + j + 1]):  # 如果satisfied
                MGS[i][j] = 3

    return MGS, Result


if __name__ == '__main__':
    project = 'STVR'
    row = 1
    string = 'TSQ'
    path = '/Applications/work/data/MT/'+project+'/Result/result25_one.xlsx'  # '+sys.argv[1][:-1]+'
    wb = load_workbook(path)
    # source_case_set, random_input = getOriginalInput()
    if string not in wb.sheetnames:
        ws = wb.create_sheet(string)
    del wb[string]
    ws = wb.create_sheet(string)
    MG_set = []
    Sus_set = []
    Metric_set = []
    Union_set = []
    Unique_set = []
    Flag_set = []
    # with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/'+string+'/OriginalInputNew.json', 'r') as load_f:
    #     data = json.load(load_f)
    # data = json.loads(data)
    # random_input = data['random_input']
    # testcases = []
    # for i in random_input:
    #     testcases.append(getInput(i, Trisquare()))
    # data = {
    #     'testcases': testcases
    # }
    # json_str = json.dumps(data)
    # with open('/Applications/work/data/MT/'+project+'/' + string + '/Input.json', 'w') as f:
    #     json.dump(json_str, f)
    # with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/Input.json', 'r') as f:
    #     data = json.load(f)
    # data = json.loads(data)
    # testcases = data['testcases']
    # for mu in range(1, 7):
    #     ExelineS = []
    #     Executable = []
    #     for i in testcases:
    #         Exelines, Executable = statements(i, mu)
    #         ExelineS.append(Exelines)
    #     data = {
    #         'Exec': Executable, 'Exel': ExelineS
    #     }
    #     json_str = json.dumps(data)
    #     with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/statements{}.json'.format(mu), 'w') as f:
    #         json.dump(json_str, f)
    # with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/statements.json', 'r') as f:
    #     data = json.load(f)
    # data = json.loads(data)
    # ExelinesS = data['Exel']
    # Executable = data['Exec']
    # # inters = set(ExelinesS[0][0])
    # # 逐个计算交集
    # for i in range(len(ExelinesS)):
    #     for j in range(len(ExelinesS[0])):
    #         inters = inters.intersection(set(ExelinesS[i][j]))
    # # 将结果转换为列表
    # inters = list(inters)
    # inters.sort()
    # for i in range(len(ExelinesS)):
    #     for j in range(len(ExelinesS[0])):
    #         ExelinesS[i][j] = [x for x in ExelinesS[i][j] if x not in inters]
    # removemg = []
    # for i in range(len(ExelinesS)):
    #     a = 0
    #     for j in range(len(ExelinesS[i])):
    #         if len(ExelinesS[i][j]) == 0:
    #             a += 1
    #     if a == len(ExelinesS[i]):
    #        removemg.append(i)
    # for index in sorted(removemg, reverse=True):
    #     del ExelinesS[index]
    # removemu = []
    # for mu in range(1, 7):  # 1, 7
    #     with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
    #               'r') as load_f:
    #         data = json.load(load_f)
    #     data = json.loads(data)
    #     Flag = data['Flag']
    #     fault = Executable[Flag.index(1)]
    #     if fault in inters:
    #         removemu.append(mu)
    ExelineSet = []
    for mu in range(1, 7):  # 1, 7
        # dynamic = eval("Mutant" + str(mu))()
        # MGS = []
        # Result = []
        # for i in range(100):
            # MG, result = riskIndex(random_input[i], dynamic)
            # MGS.append(MG)
            # Result.append(result)
        # with open('/Applications/work/data/MT/'+project+'/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # MGS = data['MGS']
        # Flag = data['Flag']
        # Flag = [0] * len(Executable)
        with open('/Users/rendaixu/OneDrive/data/MT/'+project+'/' + string + '/statements{}.json'.format(mu), 'r') as f:
            data = json.load(f)
        data = json.loads(data)
        ExelineS = data['Exel']
        Executable = data['Exec']
        ExelineSet.append(ExelineS)
        # if mu == 1:
        #     Flag[16] = 1
        # elif mu == 2:
        #     Flag[19] = 1
        # elif mu == 3:
        #     Flag[21] = 1
        # elif mu == 4:
        #     Flag[24] = 1
        # elif mu == 5:
        #     Flag[27] = 1
        # elif mu == 6:
        #     Flag[28] = 1
        # if mu in removemu:
        #     continue
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                  'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        # mSus = data['mSus']
        # mSus2 = data['mSus2']
        # mSus3 = data['mSus3']
        # mSus4 = data['mSus4']
        # metric = data['metric']
        # metric2 = data['metric2']
        # metric3 = data['metric3']
        # metric4 = data['metric4']
        Flag = data['Flag']
        # Flag = [0] * len(Executable)
        # if mu == 1:
        #     Flag[17] = 1
        # elif mu == 2:
        #     Flag[21] = 1
        # elif mu == 3:
        #     Flag[24] = 1
        # elif mu == 4:
        #     Flag[28] = 1
        # elif mu == 5:
        #     Flag[32] = 1
        # elif mu == 6:
        #     Flag[34] = 1
        # data['Flag'] = Flag
        MGS = data['MGS']
        Result = data['Result']
        mSus4 = data['mSus4_one']
        # mSus_all = data['mSus_all']
        # mSus4 = data['mSus4_all']
        mSus = data['mSus_one']
        Para = data['Para']
        # Sus = []
        # Metric = []
        # for i in range(len(mSus)):
        #     a = [round(mSus[i], 4)]
        #     b = [metric[i]]
        #     for j in range(len(mSus3)):
        #         a.append(mSus3[j][5][i])
        #         b.append(metric3[j][i])
        #     Sus.append(a)
        #     Metric.append(b)
        # Sus_set.append(Sus)
        # Metric_set.append(Metric)
        # mSus, metric = MSlice_one(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus_one'] = mSus
        # data['metric_one'] = metric
        # mSus, metric = MSlice_all(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus_all'] = mSus
        # data['metric_all'] = metric
        # mSus3, AllF, _ = MSlice3(MGS, Executable, ExelinesS, Flag)
        # Union_set.append(Union)
        # Unique_set.append(Unique)
        # mSus4, _, metric4 = MSlice4_one(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus4_one'] = mSus4
        # data['metric4_one'] = metric4
        # data['Para'] = Para
        # mSus4, _, metric4 = MSlice4_all(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus4_all'] = mSus4
        # data['metric4_all'] = metric4
        # data = {
        #          'mSus': mSus, 'mSus4': mSus4, 'metric': metric,
        #     'metric4': metric4, 'Para': Para, 'MGS': MGS, 'Flag': Flag,
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'w') as f:
        #     json.dump(json_str, f)
        row = eval('getMetrics_5')(row, ws, mu, MGS, mSus, mSus4, Para[-1], Flag, Para, ExelineS)
    wb.save(path)


