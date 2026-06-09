from coverage import CoveragePlugin
import os
from Tcas import *
import coverage
import json
from publicFun import *
# from Mutants import *
from Mutant1 import *
from Mutant2 import *
from Mutant3 import *
from Mutant4 import *
from Mutant5 import *
from Mutant6 import *
from Mutant7 import *
from Mutant8 import *
from Mutant9 import *
from Mutant10 import *
from Mutant11 import *
from Mutant12 import *
from Mutant13 import *
from Mutant14 import *
from Mutant15 import *
from Mutant16 import *
from Mutant17 import *
from Mutant18 import *
from Mutant19 import *
from Mutant20 import *
from Original import *
from openpyxl import load_workbook
random.seed(1)


# def getOriginalInput():
#     source_case_set = []
#     path = r"./reference/testplans.alt/universe.txt"
#     with open(path, 'r') as f:
#         list_read = f.readlines()
#     for i in range(len(list_read)):
#         test_case = []
#         list2 = list_read[i].split()
#         for j in range(len(list2)):
#             test_case.append(int(list2[j]))  # 字符串转整型
#         # 获取符合MR要求的源测试用例
#         if test_case[6] <= 3:
#             source_case_set.append(test_case)
#         if len(source_case_set) >= 1000:
#             break
#     # 随机取100个测试用例
#     random_input = random.sample(source_case_set, 100)
#     data = {
#             'source_case_set': source_case_set,
#             'random_input': random_input
#     }
#     json_str = json.dumps(data)
#     with open('../../data/' + string + '/OriginalInput.json', 'w') as f:
#         json.dump(json_str, f)
#     return source_case_set, random_input
#
#
def FailureRate(dynamic):
    # 把SourceCases读出来
    Result = []
    with open('../../data/'+string+'/Input.json', 'r') as load_f:
        data = json.load(load_f)
    data = json.loads(data)
    source_case_set = data['testcases']
    for i in range(len(source_case_set)):
        for j in range(len(source_case_set[i])):
            result_s_a = TCAS().Tcas(source_case_set[i][j])  # oracle
            result_s_m = dynamic.Tcas(source_case_set[i][j])
            if result_s_a == result_s_m:
                Result.append(0)
            else:
                Result.append(1)
    FR = round(Result.count(1) / len(Result) * 100, 2)
    return FR

#

def getInput(argv, dynamic):
    testcase = []  # 原始
    follow1 = []  # 数字代表层数的大小
    MGS = []
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case1 = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case1)):
        testcase.append(follow_case1[i])
    MGS.append(MG)
    for i in range(len(follow_case1)):  # t2t3t4
        MG, follow_case2 = MTG(follow_case1[i], dynamic)  # t2t5...
        follow1.append(follow_case2)  # [t5t6t7], [t8t9t10],...
        MGS.append(MG)
        for j in range(len(follow_case2)):  # t5t6t7, t8t9t10,...
            testcase.append(follow_case2[j])  # t1t2t3t4t5t6t7t8...
    return testcase


def statements(argv, mu):
    testcase = argv.copy()
    filename = 'Mutant{}.py'.format(mu)
    Exelines = []
    executable = []
    for i in range(len(testcase)):
        cov = coverage.coverage()
        cov.start()
        result_s_a = eval('Mutant{}()'.format(mu)).Tcas(testcase[i])  # oracle
        cov.stop()
        numlist = cov.analysis(filename)
        executable = numlist[1]
        exelist = list(set(numlist[1]) - set(numlist[2]))
        exelist.sort()
        Exelines.append(exelist)
    return Exelines, executable


def riskIndex(argv, dynamic):
    MGS = []  # 原始
    Result = []
    testcase = []  # 原始
    follow1 = []  # 数字代表层数的大小
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case1 = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case1)):
        testcase.append(follow_case1[i])
    MGS.append(MG)
    for i in range(len(follow_case1)):  # t2t3t4
        MG, follow_case2 = MTG(follow_case1[i], dynamic)  # t2t5...
        follow1.append(follow_case2)  # [t5t6t7], [t8t9t10],...
        MGS.append(MG)
        for j in range(len(follow_case2)):  # t5t6t7, t8t9t10,...
            testcase.append(follow_case2[j])  # t1t2t3t4t5t6t7t8...

    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        result_s_a = TCAS().Tcas(testcase[i])  # oracle
        result_s_m = dynamic.Tcas(testcase[i])
        if result_s_a == result_s_m:
            Result.append(0)
        else:
            Result.append(1)

    # 去掉巧合满足性
    for i in range(len(MGS)):
        for j in range(len(MGS[i])):
            if MGS[i][j] == 0 and (Result[i] or Result[i * len(MGS[0]) + j + 1]):  # 如果satisfied
                MGS[i][j] = 3

    return MGS, Result


if __name__ == '__main__':
    project = 'EMS'
    string = 'Tcas'
    # title = [30, 31, 32, 33]
    # N = [1, 3, 5, 10]
    for i in range(1):
        row = 1
        path = '../../data/results/resultEXAM.xlsx'  # '+sys.argv[1][:-1]+'
        wb = load_workbook(path)
        # source_case_set, random_input = getOriginalInput()
        if string in wb.sheetnames:
            del wb[string]
        ws = wb.create_sheet(string)
        # with open('../../data/MT/'+project+'/'+string+'/OriginalInput.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # random_input = data['random_input']
        # testcases = []
        # for i in random_input:
        #     testcases.append(getInput(i, TCAS()))
        # data = {
        #     'testcases': testcases
        # }
        # json_str = json.dumps(data)
        # with open('../../data/'+string+'/Input.json', 'w') as f:
        #     json.dump(json_str, f)
        # with open('../../data/MT/'+project+'/' + string + '/Input.json', 'r') as f:
        #     data = json.load(f)
        # data = json.loads(data)
        # testcases = data['testcases']
        # ExelinesS = []
        # for i in testcases:
        #     Exelines, Executable = statements(i)
        #     ExelinesS.append(Exelines)
        # data = {
        #     'Exec': Executable, 'Exel': ExelinesS
        # }
        # json_str = json.dumps(data)
        # with open('../../data/'+string+'/statements.json', 'w') as f:
        #     json.dump(json_str, f)
        #
        # with open('../../data/MT/'+project+'/' + string + '/statements.json', 'r') as f:
        #     data = json.load(f)
        # data = json.loads(data)
        # ExelinesS = data['Exel']
        # Executable = data['Exec']
        # inters = set(ExelinesS[0][0])
        # # 逐个计算交集
        # for i in range(len(ExelinesS)):
        #     for j in range(len(ExelinesS[0])):
        #         inters = inters.intersection(set(ExelinesS[i][j]))
        # # 将结果转换为列表
        # inters = list(inters)
        # inters.sort()
        # for i in range(len(ExelinesS)):
        #     for j in range(len(ExelinesS[0])):
        #         ExelinesS[i][j] = [x for x in ExelinesS[i][j] if x not in inters]
        # removemg = []
        # for i in range(len(ExelinesS)):
        #     a = 0
        #     for j in range(len(ExelinesS[i])):
        #         if len(ExelinesS[i][j]) == 0:
        #             a += 1
        #     if a == len(ExelinesS[i]):
        #        removemg.append(i)
        # for index in sorted(removemg, reverse=True):
        #     del ExelinesS[index]
        # removemu = []
        # for mu in range(1, 21):
        #     with open('../../data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #               'r') as load_f:
        #         data = json.load(load_f)
        #     data = json.loads(data)
        #     Flag = data['Flag']
        #     fault = Executable[Flag.index(1)]
        #     if fault in inters:
        #         removemu.append(mu)
        # for mu in range(1, 21):
        #     ExelineS = []
        #     Executable = []
        #     for i in testcases:
        #         Exelines, Executable = statements(i, mu)
        #         ExelineS.append(Exelines)
        #     data = {
        #         'Exec': Executable, 'Exel': ExelineS
        #     }
        #     json_str = json.dumps(data)
        #     with open('../../data/MT/'+project+'/' + string + '/statements{}.json'.format(mu), 'w') as f:
        #         json.dump(json_str, f)
        for mu in range(1, 21):  # 1, 21
            # dynamic = TCASFactory("Mutant" + str(mu)).getTCAS()
            # #     a = FailureRate(dynamic)
            # #     FR.append(a)
            # MGS = []
            # Result = []
            # for i in range(100):
            #     MG, result = riskIndex(random_input[i], dynamic)
            #     MGS.append(MG)
            #     Result.append(result)
            #     Executable_set.append(ExecutableS)
            #     print(mu)
            #     # with open('../../data/'+project2+'/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
            #     #     data = json.load(load_f)
            #     # data = json.loads(data)
            #     # MGS = data['MGS']
            #     # SMGS = data['SMGS']
            #     # ExecutableS = data['Exec']
            #     # ExelinesS = data['Exel']
            #     # Flag = data['flag']
            #     # ExecutableS_set.append(ExecutableS)
            #     # Flag_set.append(Flag)
            #     Flag = [0] * len(Executable)
            #     if mu == 1:
            #         Flag[13] = 1
            #     elif mu == 2:
            #         Flag[11] = 1
            #     elif mu == 3:
            #         Flag[44] = 1
            #     elif mu == 4:
            #         Flag[48] = 1
            #     elif mu == 5:
            #         Flag[5] = 1
            #     elif mu == 6:
            #         Flag[46] = 1
            #     elif mu == 7:
            #         Flag[16] = 1
            #     elif mu == 8:
            #         Flag[44] = 1
            #     elif mu == 9:
            #         Flag[5] = 1
            #     elif mu == 10:
            #         Flag[45] = 1
            #     elif mu == 11:
            #         Flag[15] = 1
            #     elif mu == 12:
            #         Flag[22] = 1
            #     elif mu == 13:
            #         Flag[13] = 1
            #     elif mu == 14:
            #         Flag[19] = 1
            #     elif mu == 15:
            #         Flag[44] = 1
            #     elif mu == 16:
            #         Flag[44] = 1
            #     elif mu == 17:
            #         Flag[9] = 1
            #     elif mu == 18:
            #         Flag[11] = 1
            #     elif mu == 19:
            #         Flag[9] = 1
            #     elif mu == 20:
            #         Flag[11] = 1
            # if mu in removemu:
            #     continue
            # with open('../../data/MT/'+project+'/' + string + '/statements{}.json'.format(mu), 'r') as f:
            #     data = json.load(f)
            # data = json.loads(data)
            # ExelineS = data['Exel']
            # Executable = data['Exec']
            with open('../../data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                      'r') as load_f:
                data = json.load(load_f)
            data = json.loads(data)
            # Flag = [0] * len(Executable)
            # if mu == 1:
            #     Flag[13] = 1
            # elif mu == 2:
            #     Flag[11] = 1
            # elif mu == 3:
            #     Flag[44] = 1
            # elif mu == 4:
            #     Flag[48] = 1
            # elif mu == 5:
            #     Flag[5] = 1
            # elif mu == 6:
            #     Flag[46] = 1
            # elif mu == 7:
            #     Flag[16] = 1
            # elif mu == 8:
            #     Flag[44] = 1
            # elif mu == 9:
            #     Flag[5] = 1
            # elif mu == 10:
            #     Flag[45] = 1
            # elif mu == 11:
            #     Flag[15] = 1
            # elif mu == 12:
            #     Flag[22] = 1
            # elif mu == 13:
            #     Flag[13] = 1
            # elif mu == 14:
            #     Flag[19] = 1
            # elif mu == 15:
            #     Flag[44] = 1
            # elif mu == 16:
            #     Flag[44] = 1
            # elif mu == 17:
            #     Flag[9] = 1
            # elif mu == 18:
            #     Flag[11] = 1
            # elif mu == 19:
            #     Flag[9] = 1
            # elif mu == 20:
            #     Flag[11] = 1
            # data['Flag'] = Flag
            # mSus = data['mSus']
            # mSus2 = data['mSus2']
            # mSus3 = data['mSus3']
            # mSus4 = data['mSus4']
            # metric = data['metric']
            # metric2 = data['metric2']
            # metric3 = data['metric3']
            # metric4 = data['metric4']
            Flag = data['Flag']
            # data['MGS'] = MGS
            # data['Result'] = Result
            # MGS = data['MGS']
            # Result = data['Result']
            # mSus4 = data['mSus4']
            # mSus = data['mSus']
            sbflsus = data['sbflsus']
            # Para = data['Para']
            # if i == 0:
            #     mSus, metric = MSlice_all(MGS, Executable, ExelineS, Flag, Result)
            #     mSus4, _, metric4 = MSlice4_all(MGS, Executable, ExelineS, Flag, Result)
            #     # sbflsus, sbflmetric = SBFL(MGS, Result, Executable, ExelineS)
            #     data['mSus_all'] = mSus
            #     data['metric_all'] = metric
            #     data['mSus4_all'] = mSus4
            #     data['metric4_all'] = metric4
            #     # data['sbflsus'] = sbflsus
            #     # data['sbflmetric'] = sbflmetric
            #     # Union_set.append(Union)
            #     # Unique_set.append(Unique)
            #     # mSus4, _, metric4 = MSlice4_one(MGS, Executable, ExelineS, Flag, Result)
            #     # data['mSus4_one'] = mSus4
            #     # data['metric4_one'] = metric4
            #     # data['Para'] = Para
            #     # mSus4, _, metric4 = MSlice4_all(MGS, Executable, ExelineS, Flag, Result)
            #     # data['mSus4_all'] = mSus4
            #     # data['metric4_all'] = metric4
            #     # data = {
            #     #          'mSus': mSus, 'mSus4': mSus4, 'metric': metric,
            #     #     'metric4': metric4, 'Para': Para, 'MGS': MGS, 'Flag': Flag,
            #     # }
            #     json_str = json.dumps(data)
            #     with open('../../data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
            #               'w') as f:
            #         json.dump(json_str, f)
            # else:
            mSus4 = data['mSus4']
            mSus = data['mSus']
            # row = eval('getMetrics_5')(row, ws, mu, MGS, mSus, mSus4,sbflsus, Para[-1], Flag, Para, ExelineS, N[i])
            row = eval('getMetrics_6')(row, ws, mu, mSus, mSus4, sbflsus, Flag)
        wb.save(path)
