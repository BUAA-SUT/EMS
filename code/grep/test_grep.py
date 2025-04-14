# 生成follow
# 获取正则表达式
from myutl.Utl import Utl
import linecache
from constant import constantNumber as constant
import os
import random
import subprocess
import csv
from MRs.MR import *
import time
import threading
from openpyxl import load_workbook
from publicFun import *
import json


random.seed(1)


def random_select_MR(MR_list):
    """
    randomly select a MR from the MR list
    :param MR_list:
    :return: the name of selected MR
    """
    index = random.randint(0, len(MR_list) - 1)

    return MR_list[index]


def generate_source_test_case():
    file_path = os.path.join(os.path.abspath('.'), 'files', 'partition_scheme_testcases_1.2')
    file_path2 = os.path.join(os.path.abspath('.'), 'files', 'executed_correct_test_cases')
    alltestcases = []
    source_test_case = {}
    with open(file_path, 'r') as file:
        for aline in file:
            alltestcases.append(aline.strip())
    source_test_case_name = []
    with open(file_path2, 'r') as file:
        for aline in file:
            source_test_case_name.append(int(aline))
    for i in source_test_case_name:
        dictionary = {"input" + str(i): alltestcases[i-1]}
        source_test_case.update(dictionary)

    return source_test_case


def generate_follow_test_case(newdata):
    """
    根据选择的蜕变关系以及原始测试用例生成衍生测试用例
    :param MR_name: 　选择的蜕变关系的名称
    :param source_test_case: 原始测试用例
    :return: 衍生测试用例
    """
    # 读源测试用例
    inputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomInput.csv'
    # csvFile = open(inputdir, "r")
    # reader = csv.reader(csvFile)
    # # 建立空字典
    # sourcedata = {}
    # inputdata = []
    followdata = {}
    factory = MR_factory()
    # a = 0
    # b = 0
    # for item in reader:
    #     # 忽略第一行
    #     if reader.line_num == 1:
    #         inputdata.append([item[0], item[1]])
    #         continue
    #     try:
    #         a += 1
    #         sourcedata[item[0]] = item[1]
    #         inputdata.append([item[0], item[1]])
    #         # 确定测试用例的apply MR
    #         # index =
    #         MRs = linecache.getline(constant.test_cases_2_mrs_path, a). \
    #             replace('\'', '').replace('\'', '').strip().split(
    #                 ':')[1].replace('[', '').replace(']', '')
    #         MRs_list = MRs.split(', ')
    #         source_pattern = newdata.get(item[0])
    #         for i in MRs_list:
    #             b += 1
    #             MR_obj = factory.choose_MR(i)
    #             follow_test_case = MR_obj.generate_follow_test_case(source_pattern, a)
    #             followdata[item[0]+"_"+i[2:]] = follow_test_case
    #             inputdata.append([item[0]+"_"+i[2:], follow_test_case])
    #
    #             MRs = linecache.getline(constant.test_cases_2_mrs_path, b). \
    #                 replace('\'', '').replace('\'', '').strip().split(
    #                 ':')[1].replace('[', '').replace(']', '')
    #             MRs_list = MRs.split(', ')
    #             source_pattern = newdata.get(item[0] + "_" + i[2:])
    #             for j in MRs_list:
    #                 MR_obj = factory.choose_MR(j)
    #                 follow_test_case = MR_obj.generate_follow_test_case(source_pattern, b)
    #                 followdata[item[0] + "_" + i[2:] + "_" + j[2:]] = follow_test_case
    #                 inputdata.append([item[0] + "_" + i[2:] + "_" + j[2:], follow_test_case])
    #     except:
    #         print(item)
    #         print(reader.line_num)
    #         break
    # csvFile.close()
    inputdata = [["name", "value"]]
    file_path = os.path.join(os.path.abspath('.'), 'files', 'executed_correct_test_cases')
    with open(file_path, 'r') as file:
        for aline in file:
            # 确定测试用例的apply MR
            index1 = list(newdata.keys()).index('input' + aline.strip()) + 1
            MRs = linecache.getline(constant.test_cases_2_mrs_path, index1). \
                replace('\'', '').replace('\'', '').strip().split(
                ':')[1].replace('[', '').replace(']', '')
            MRs_list = MRs.split(', ')
            source_pattern = newdata.get('input' + aline.strip())
            inputdata.append(['input' + aline.strip(), source_pattern])
            for i in MRs_list:
                MR_obj = factory.choose_MR(i)
                follow_test_case = MR_obj.generate_follow_test_case(source_pattern, index1)
                followdata['input' + aline.strip() + "_" + i[2:]] = follow_test_case
                inputdata.append(['input' + aline.strip() + "_" + i[2:], follow_test_case])
                index2 = list(newdata.keys()).index('input' + aline.strip() + "_" + i[2:]) + 1
                source_pattern2 = newdata.get('input' + aline.strip() + "_" + i[2:])
                MRs2 = linecache.getline(constant.test_cases_2_mrs_path, index2). \
                    replace('\'', '').replace('\'', '').strip().split(
                    ':')[1].replace('[', '').replace(']', '')
                MRs_list2 = MRs2.split(', ')

                for j in MRs_list2:
                    try:
                        MR_obj = factory.choose_MR(j)
                        follow_test_case = MR_obj.generate_follow_test_case(source_pattern2, index2)
                        followdata['input' + aline.strip() + "_" + i[2:] + "_" + j[2:]] = follow_test_case
                        inputdata.append(['input' + aline.strip() + "_" + i[2:] + "_" + j[2:], follow_test_case])
                    except:
                        print(aline, i, j)
                        exit()

    csvFile = open(inputdir, "w")
    writer = csv.writer(csvFile)
    writer.writerows(inputdata)
    csvFile.close()
    return followdata


def get_input():
    inputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomInput.csv'
    # inputdir = '/home/rdx/data/MT/STVR/grep/RandomInput.csv'
    csvFile = open(inputdir, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    inputcase = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            inputcase[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    return inputcase


def get_output(newdata, mu):
    # outputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomOutput{}.csv'.format(mu)
    outputdir = '/home/rdx/data/MT/STVR/grep/RandomOutput{}.csv'.format(mu)
    outputdata = [["name", "value"]]
    file_path = os.path.join(os.path.abspath('.'), 'files', 'executed_correct_test_cases')
    with open(file_path, 'r') as file:
        a = 0
        for aline in file:
            a += 1
            # 确定测试用例的apply MR
            # aline = '3897'
            MRs = linecache.getline(constant.test_cases_2_mrs_path, a). \
                replace('\'', '').replace('\'', '').strip().split(
                ':')[1].replace('[', '').replace(']', '')
            MRs_list = MRs.split(', ')
            for i in MRs_list:
                a += 1
                source_pattern = newdata.get('input' + aline.strip())
                follow_pattern = newdata.get('input' + aline.strip() + "_" + i[2:])
                try:
                    if i != 'MR11' and i != 'MR9':
                        source_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + source_pattern \
                                         + "\" " + "./targetFiles/file.test"
                        follow_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + follow_pattern \
                                         + "\" " + "./targetFiles/file.test"
                    elif i == 'MR11':
                        source_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + source_pattern \
                                         + "\" " + "./targetFiles/MR11_" + aline.strip()
                        follow_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + follow_pattern \
                                         + "\" " + "./targetFiles/MR11_" + aline.strip()
                    else:
                        source_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + source_pattern \
                                         + "\" " + "./targetFiles/file.test"
                        follow_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + follow_pattern \
                                         + "\" " + "./targetFiles/file.test_MR9_follow"
                except:
                    print(aline, i, a)
                    exit()
                output_source = subprocess.run(source_command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                        shell=True).stdout.decode()
                output_source_list = output_source.split("\n")
                if 'not found' in output_source_list[0]:
                    output_source = 'command not found'
                output_follow = subprocess.run(follow_command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                        shell=True).stdout.decode()
                output_follow_list = output_follow.split("\n")
                if 'not found' in output_follow_list[0]:
                    output_follow = 'command not found'
                outputdata.append(['output_source' + aline.strip()+"_" + i[2:], output_source])
                outputdata.append(['output_follow' + aline.strip()+"_" + i[2:], output_follow])

                # 确定测试用例的apply MR
                # aline = '3897'
                MRs = linecache.getline(constant.test_cases_2_mrs_path, a). \
                    replace('\'', '').replace('\'', '').strip().split(
                    ':')[1].replace('[', '').replace(']', '')
                MRs_list = MRs.split(', ')
                for j in MRs_list:
                    source_pattern = newdata.get('input' + aline.strip() + "_" + i[2:])
                    follow_pattern = newdata.get('input' + aline.strip() + "_" + i[2:] + "_" + j[2:])
                    try:
                        if j != 'MR11' and j != 'MR9':
                            source_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + source_pattern \
                                             + "\" " + "./targetFiles/file.test"
                            follow_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + follow_pattern \
                                             + "\" " + "./targetFiles/file.test"
                        elif j == 'MR11':
                            source_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + source_pattern \
                                             + "\" " + "./targetFiles/MR11_" + aline.strip()
                            follow_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + follow_pattern \
                                             + "\" " + "./targetFiles/MR11_" + aline.strip()
                        else:
                            source_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + source_pattern \
                                             + "\" " + "./targetFiles/file.test"
                            follow_command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + follow_pattern \
                                             + "\" " + "./targetFiles/file.test_MR9_follow"
                    except:
                        print(aline, i, j, a)
                        exit()
                    output_source = subprocess.run(source_command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                                   shell=True).stdout.decode()
                    output_source_list = output_source.split("\n")
                    if 'not found' in output_source_list[0]:
                        output_source = 'command not found'
                    output_follow = subprocess.run(follow_command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                                   shell=True).stdout.decode()
                    output_follow_list = output_follow.split("\n")
                    if 'not found' in output_follow_list[0]:
                        output_follow = 'command not found'
                    outputdata.append(['output_source' + aline.strip() + "_" + i[2:] + "_" + j[2:], output_source])
                    outputdata.append(['output_follow' + aline.strip() + "_" + i[2:] + "_" + j[2:], output_follow])

        csvFile = open(outputdir, "w+")
        writer = csv.writer(csvFile)
        writer.writerows(outputdata)
        csvFile.close()


def getTestcase(newdata, mu):
    inputdata = [["name", "value"]]
    outputdata = [["name", "value"]]
    inputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomInput.csv'
    outputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomOutput{}.csv'.format(mu)

    for i in range(len(newdata)):
        inputdata.append([list(newdata)[i], list(newdata.values())[i]])
        pattern = list(newdata.values())[i].strip()
        command = r"./Mutants/grep_v" + str(mu) + "/grep -E " + "\"" + pattern \
                  + "\" " + "./targetFiles/file.test"
        output = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, shell=True).stdout.decode()
        # with open(outputdir+'output'+list(newdata)[i][5:], 'w+') as file:
        #     file.write(output)
        outputdata.append(['output'+list(newdata)[i][5:], output])
    csvFile1 = open(inputdir, "w")
    writer = csv.writer(csvFile1)
    writer.writerows(inputdata)
    csvFile1.close()

    csvFile2 = open(outputdir, "w+")
    writer = csv.writer(csvFile2)
    writer.writerows(outputdata)
    csvFile2.close()


def verify_result_not_MR11(MR_name, output_source, output_follow):
    return MR_factory().verify_result_no_MR11(MR_name, output_source, output_follow)


def verify_result_MR11(input_index, output_source, output_follow):
    return MR_factory().verify_MR11_result(input_index, output_source, output_follow)


def get_correct_cases(num):
    """
    返回正确的cases
    :return:
    """
    file_path = os.path.join(os.path.abspath('.'), 'files', 'partition_scheme_testcases_1.2')
    correct_cases = []
    i = 0
    with open(file_path, 'r') as file:
        lines = file.readlines()
        lines2 = lines.copy()
        random.shuffle(lines2)
        for aline in lines2:
            aline2 = aline.strip()
            command = r"./Mutants/grep_v0/grep -E " + "\"" + aline2 \
                      + "\" " + "./targetFiles/grep1.dat"
            try:
                # 尝试执行命令
                output = subprocess.run(command, stdout=subprocess.PIPE, shell=True,
                                        stderr=subprocess.DEVNULL)
                if output.returncode != 0:
                    continue
            except subprocess.CalledProcessError as e:
                # 命令执行失败
                # output = e.output.decode()
                # print(f"Command '{command}' failed with output:\n{output}")
                pass
            else:
                # 命令执行成功
                # print(f"Command '{command}' succeeded with output:\n{output}")
                # pass
                if output.stdout.decode():
                    i += 1
                    correct_cases.append(lines.index(aline)+1)
            if i == num:
                break
    correct_cases.sort()
    with open('files/executed_correct_test_cases', 'w+') as file:
        for i in correct_cases:
            file.write(str(i) + '\n')
    return correct_cases


def verify_MRs(mu):
    Result = []
    outputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomOutput{}.csv'.format(mu)
    file_path = os.path.join(os.path.abspath('.'), 'files', 'executed_correct_test_cases')
    csvFile = open(outputdir, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    outputcase = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    with open(file_path, 'r') as file:
        a = 0
        for aline in file:
            # aline = '75'
            a += 1
            v_r = []
            input_index = int(aline)
            MRs = linecache.getline(constant.test_cases_2_mrs_path, a). \
                replace('\'', '').replace('\'', '').strip().split(
                    ':')[1].replace('[', '').replace(']', '')
            MRs_list = MRs.split(', ')
            for i in MRs_list:
                # print(i)
                temp_i = i
                output_source = outputcase.get('output_source' + aline.strip() + '_' + temp_i[2:])
                output_follow = outputcase.get('output_follow' + aline.strip() + '_' + temp_i[2:])
                # ver = i == 'MR11'
                # print(ver)
                if i != 'MR11':
                    # print(9)
                    result = verify_result_not_MR11(i, output_source, output_follow)
                else:
                    # print(11)
                    result = verify_result_MR11(input_index, output_source, output_follow)
                v_r.append(result)
                if result:
                    print(aline.strip(), i)
                a += 1
                MRs2 = linecache.getline(constant.test_cases_2_mrs_path, a). \
                    replace('\'', '').replace('\'', '').strip().split(
                    ':')[1].replace('[', '').replace(']', '')
                MRs_list2 = MRs2.split(', ')
                input_index = int(aline)
                for j in MRs_list2:
                    output_source = outputcase.get('output_source' + aline.strip() + '_' + i[2:] + '_' + j[2:])
                    output_follow = outputcase.get('output_follow' + aline.strip() + '_' + i[2:] + '_' + j[2:])
                    if j != 'MR11':
                        result = verify_result_not_MR11(j, output_source, output_follow)
                    else:
                        result = verify_result_MR11(input_index, output_source, output_follow)
                    v_r.append(result)
                    if result:
                        print(aline.strip(), i, j)
            Result.append(v_r)
    return Result


def getMG(mu, originaldata, ExecutableS):
    MGS = []
    ExelinesS = []
    Result = []
    outputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomOutput{}.csv'.format(mu)
    file_path = os.path.join(os.path.abspath('.'), 'files', 'executed_correct_test_cases')
    csvFile = open(outputdir, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    outputcase = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    with open(file_path, 'r') as file:
        a = 0
        for aline in file:
            a += 1
            MG = []
            v_r1 = []
            Exelines = []
            r = []
            exe1 = []
            r1 = []
            input_index = int(aline)
            MRs = linecache.getline(constant.test_cases_2_mrs_path, a). \
                replace('\'', '').replace('\'', '').strip().split(
                    ':')[1].replace('[', '').replace(']', '')
            MRs_list = MRs.split(', ')
            for i in MRs_list:
                # print(i)
                temp_i = i
                output_source = outputcase.get('output_source' + aline.strip() + '_' + temp_i[2:])
                output_follow = outputcase.get('output_follow' + aline.strip() + '_' + temp_i[2:])
                if i != 'MR11':
                    result = verify_result_not_MR11(i, output_source, output_follow)
                else:
                    result = verify_result_MR11(input_index, output_source, output_follow)
                if result:
                    result = 1
                else:
                    result = 0
                if result == 0 and (int(originaldata.get('output_source' + aline.strip() + '_' + temp_i[2:])[mu-12]) or
                                    int(originaldata.get('output_follow' + aline.strip() + '_' + temp_i[2:])[mu-12])):
                    result = 3
                v_r1.append(result)
                candidate = originaldata.get('output_source' + aline.strip() + '_' + temp_i[2:])[:-11]
                candidate = [int(x) for x in candidate]
                indices1 = [i for i in range(len(candidate)) if candidate[i] == 1]
                candidate = originaldata.get('output_follow' + aline.strip() + '_' + temp_i[2:])[:-11]
                candidate = [int(x) for x in candidate]
                indices2 = [i for i in range(len(candidate)) if candidate[i] == 1]
                exe1.append([[ExecutableS[i] for i in indices1], [ExecutableS[i] for i in indices2]])
                r1.append([int(originaldata.get('output_source' + aline.strip() + '_' + temp_i[2:])[mu-12]),
                           int(originaldata.get('output_follow' + aline.strip() + '_' + temp_i[2:])[mu-12])])
                a += 1
                MRs2 = linecache.getline(constant.test_cases_2_mrs_path, a). \
                    replace('\'', '').replace('\'', '').strip().split(
                    ':')[1].replace('[', '').replace(']', '')
                MRs_list2 = MRs2.split(', ')
                input_index = int(aline)
                v_r2 = []
                exe2 = []
                r2 = []
                for j in MRs_list2:
                    output_source = outputcase.get('output_source' + aline.strip() + '_' + i[2:] + '_' + j[2:])
                    output_follow = outputcase.get('output_follow' + aline.strip() + '_' + i[2:] + '_' + j[2:])
                    if j != 'MR11':
                        result = verify_result_not_MR11(j, output_source, output_follow)
                    else:
                        result = verify_result_MR11(input_index, output_source, output_follow)
                    if result:
                        result = 1  # 违反了
                    else:
                        result = 0
                    if result == 0 and (int(originaldata.get('output_source' + aline.strip() + '_' + i[2:] + '_' + j[2:])[mu - 12]) or
                                        int(originaldata.get('output_follow' + aline.strip() + '_' + i[2:] + '_' + j[2:])[mu - 12])):
                        result = 3
                    v_r2.append(result)
                    candidate = originaldata.get('output_source' + aline.strip() + '_' + i[2:] + '_' + j[2:])[:-11]
                    candidate = [int(x) for x in candidate]
                    indices1 = [i for i in range(len(candidate)) if candidate[i] == 1]
                    candidate = originaldata.get('output_follow' + aline.strip() + '_' + i[2:] + '_' + j[2:])[:-11]
                    candidate = [int(x) for x in candidate]
                    indices2 = [i for i in range(len(candidate)) if candidate[i] == 1]
                    exe2.append([[ExecutableS[i] for i in indices1], [ExecutableS[i] for i in indices2]])
                    r2.append([int(originaldata.get('output_source' + aline.strip() + '_' + i[2:] + '_' + j[2:])[mu - 12]),
                               int(originaldata.get('output_follow' + aline.strip() + '_' + i[2:] + '_' + j[2:])[mu - 12])])
                MG.append(v_r2)
                Exelines.append(exe2)
                r.append(r2)
            MG.insert(0, v_r1)
            Exelines.insert(0, exe1)
            r.insert(0, r1)
            MGS.append(MG)
            ExelinesS.append(Exelines)
            Result.append(r)
    return MGS, ExelinesS, Result


def dataTrans():
    """
    转换数据
    :return:
    """
    inputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomInput.csv'
    outputdir = './mapping relation/input'
    if os.path.exists(outputdir):
        os.remove(outputdir)
    csvFile = open(inputdir, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    inputcase = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            inputcase[item[0]] = item[1]
            with open(outputdir, 'a') as file:
                file.write(item[1]+'\n')
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()


def verify(inputdata):
    path = os.path.join('/Users/rendaixu/Library/CloudStorage/OneDrive-个人/data/MT/STVR/grep/备份/一级follow/testcase_2_MRs')
    tool = Utl()
    file = tool.get_file_object(path)
    all_sets = {}
    row = 0
    for item in file:
        row += 1
        MRs = linecache.getline(path, row). \
            replace('\'', '').replace('\'', '').strip().split(
            ':')[1].replace('[', '').replace(']', '')
        MRs_list = MRs.split(', ')
        data = {
            item.split(':', 1)[0].strip(): MRs_list
        }
        all_sets.update(data)
    file_path = os.path.join(os.path.abspath('.'), 'files', 'executed_correct_test_cases')
    with open(file_path, 'r') as file:
        for aline in file:
            index1 = list(inputdata.keys()).index('input' + aline.strip()) + 1
            num = len(all_sets.get(str(index1)))
            a = 0
            while 1:
                try:
                    if 'input' + aline.strip() in list(inputdata.keys())[index1]:
                        a += 1
                        index1 += 1
                        if index1 >= len(list(inputdata.keys())):
                            break
                    else:
                        break
                except:
                    print('出错：'+aline)
                    exit()
                    break
            if not num == a:
                print(aline, all_sets.get(str(list(inputdata.keys()).index('input' + aline.strip()) + 1)))


def mutaterate(mu):
    outputdir1 = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomOutput0.csv'
    outputdir2 = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomOutput{}.csv'.format(mu)
    csvFile = open(outputdir1, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    outputcase1 = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase1[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    csvFile = open(outputdir2, "r")
    reader = csv.reader(csvFile)
    outputcase2 = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase2[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    num = 0
    outputcase1 = list(outputcase1.values())
    outputcase2 = list(outputcase2.values())
    index = []
    for i in range(len(outputcase1)):
        if not outputcase1[i] == outputcase2[i]:
            num += 1
            index.append(i)
    rate = num / (len(outputcase1))
    print(round(rate*100, 2))
    return index


def run_with_timeout(func, args, timeout):
    thread = threading.Thread(target=func, args=args)
    thread.start()
    thread.join(timeout)

    if thread.is_alive():
        print(f"Mutant {args[1]} 超时，跳过")
        return False
    return True


if __name__ == "__main__":
    # correct = get_correct_cases(1000)
    # source_test_case = generate_source_test_case()
    # inputdata = get_input()
    # timeout = 600  # 设定阈值为**秒
    # for i in range(13):
    #     run_with_timeout(get_output, (inputdata, i), timeout)
    # result = verify_MRs(0)
    # for i in range(0, 12):
    #     print(i)
    #     row = mutaterate(i)
    #     result = verify_MRs(i)

    # getTestcase(source_test_case, 0)  # 0代表original version
    # followdata = generate_follow_test_case(inputdata)
    # verify(inputdata)

    # source_results = []
    # source_result_path = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomOutput0/output13'
    # with open(source_result_path, 'r') as source_file:
    #     for aline in source_file:
    #         source_results.append(aline)
    # source_file.close()
    #
    # outputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/grep/RandomOutput0.csv'
    # csvFile = open(outputdir, "r")
    # reader = csv.reader(csvFile)
    # # 建立空字典
    # outputcase = {}
    # for item in reader:
    #     # 忽略第一行
    #     if reader.line_num == 1:
    #         continue
    #     try:
    #         outputcase[item[0]] = item[1]
    #     except:
    #         print(item)
    #         print(reader.line_num)
    #     # break
    #
    # output_source = outputcase.get('output114')
    # output_source2 = output_source.split('\n')
    # dataTrans()
    project = 'STVR'
    row = 1
    string = 'grep'
    path = '/Applications/work/data/MT/' + project + '/Result/result25_one.xlsx'  # '+sys.argv[1][:-1]+'
    # wb = load_workbook(path)
    # # if string not in wb.sheetnames:
    # #     ws = wb.create_sheet(string)
    # # del wb[string]
    # # ws = wb.create_sheet(string)
    # if string not in wb.sheetnames:
    #     ws = wb.create_sheet(string)
    # del wb[string]
    # ws = wb.create_sheet(string)
    MG_set = []
    M_set = []
    Sus_set = []
    Flag_set = []
    ExelineSet = []
    ExecutableSet = []
    removemu = [3]
    for mu in range(1, 12):  # 1, 12
        if mu in removemu:
            continue
        originaldata = {}
        for i in range(1, 46):
            datadir = '/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/statementResult{}_{}.csv'.format(mu, i)
            csvFile = open(datadir, "r")
            reader = csv.reader(csvFile)
            # 建立空字典
            for item in reader:
                # if i != 1 and reader.line_num == 1:
                #     continue
                originaldata[item[0]] = item[1:]
            csvFile.close()
        ExecutableS = originaldata.get('test cases')[:-11]
        ExecutableS = [int(x) for x in ExecutableS]
        _, ExelinesS, _ = getMG(mu, originaldata, ExecutableS)
    #     # MG_set.append(MGS)
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        MGS = data['MGS']
        # data['Exec'] = ExecutableS
        # data['Exel'] = ExelinesS
        Executable = data['Exec']
        ExelineS = data['Exel']
        # ExelineSet.append(ExelineS)
        # ExecutableSet.append(Executable)
    # #
    # #     # inters = set(ExelinesS[0][0][0][0])
    # #     # # 逐个计算交集
    # #     # for i in range(len(ExelinesS)):
    # #     #     for j in range(len(ExelinesS[i])):
    # #     #         for m in range(len(ExelinesS[i][j])):
    # #     #             for n in range(len(ExelinesS[i][j][m])):
    # #     #                 inters = inters.intersection(set(ExelinesS[i][j][m][n]))
    # #     # # 将结果转换为列表
    # #     # inters = list(inters)
    # #     # inters.sort()
    # #     # for i in range(len(ExelinesS)):
    # #     #     for j in range(len(ExelinesS[i])):
    # #     #         for m in range(len(ExelinesS[i][j])):
    # #     #             for n in range(len(ExelinesS[i][j][m])):
    # #     #                 ExelinesS[i][j][m][n] = [x for x in ExelinesS[i][j][m][n] if x not in inters]
    # #
    # #     # AllF = data['AllF']
    # #     # Para = data['Para']
    # #     mSus = data['mSus']
    # #     mSus4 = data['mSus4']
    # #     Para = data['Para']
        Flag = data['Flag']
        Flag_set.append(Flag)
        # mSus = data['mSus']
        # mSus4 = data['mSus4']
        Result = data['Result']
        mSus4 = data['mSus4_one']
        mSus = data['mSus_one']
        Para = data['Para']
        # Para = data['Para']
        # mSus4_all = data['mSus4_all']
        # mSus_all = data['mSus_all']
        # mSus4_one = data['mSus4_one']
        # mSus_one = data['mSus_one']
        # mSus4_nofs = data['mSus4_nofs']
        # mSus_nofs = data['mSus_nofs']
    #     # fault = Executable[Flag.index(1)]
    #     # if fault in inters:
    #     #     removemu.append(mu)
    #     # metric = data['metric']
    #     # metric2 = data['metric2']
    #     # metric3 = data['metric3']
    #     # metric4 = data['metric4']
    #     # data['Exec'] = ExecutableS
    #     # data['Exel'] = ExelinesS
    #     # data['Result'] = Result
    #     # Sus = []
    #     # Metric = []
    #     # for i in range(len(mSus)):
    #     #     a = [round(mSus[i], 4)]
    #     #     b = [metric[i]]
    #     #     for j in range(len(mSus3)):
    #     #         a.append(mSus3[j][5][i])
    #     #         b.append(metric3[j][i])
    #     #     Sus.append(a)
    #     #     Metric.append(b)
    #     # Sus_set.append(Sus)
    #     # Metric_set.append(Metric)
    #     mSus_nofs, metric_nofs = MSlice_grep(MGS, ExecutableS, ExelinesS)
    #     data['mSus_nofs'] = mSus_nofs
    #     data['metric_nofs'] = metric_nofs
    #     # mSus3, AllF, _ = MSlice3(MGS, Executable, ExelinesS, Flag)
    #     # Union_set.append(Union)
    #     # Unique_set.append(Unique)
    #     mSus4_nofs, Para, metric4_nofs = MSlice_grep4(MGS, ExecutableS, ExelinesS, Flag, Result)
        # data['Para'] = Para
        # data['mSus4_nofs'] = mSus4_nofs
        # data['metric4_nofs'] = metric4_nofs
        # mSus, metric = MSlice_grep_one(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus_one'] = mSus
        # data['metric_one'] = metric
        # mSus, metric = MSlice_grep_all(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus_all'] = mSus
        # data['metric_all'] = metric
        # mSus4, _, metric4 = MSlice_grep4_one(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus4_one'] = mSus4
        # data['metric4_one'] = metric4
        # data['Para'] = Para
        # mSus4, _, metric4 = MSlice_grep4_all(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus4_all'] = mSus4
        # data['metric4_all'] = metric4
        # data = {
        #          'mSus': mSus, 'mSus4': mSus4, 'metric': metric,
        #     'metric4': metric4, 'Para': Para, 'MGS': MGS, 'Flag': Flag,
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'w') as f:
        #     json.dump(json_str, f)
        row = eval('getMetrics_5_grep')(row, ws, mu, MGS, mSus, mSus4, Para[-1], Flag, Para, ExelineS)
        print(mu)
    wb.save(path)

    # for mu in range(1, 2):  # 1, 12
    #     with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
    #         data = json.load(load_f)
    #     data = json.loads(data)
    #     ExelinesS = data['Exel']
    #     inters = set(ExelinesS[0][0][0][0])
    #     # 逐个计算交集
    #     for i in range(len(ExelinesS)):
    #         for j in range(len(ExelinesS[i])):
    #             for m in range(len(ExelinesS[i][j])):
    #                 for n in range(len(ExelinesS[i][j][m])):
    #                     inters = inters.intersection(set(ExelinesS[i][j][m][n]))
    #         print(i, len(list(inters)))
    #     # 将结果转换为列表
    #     inters = list(inters)
    #     inters.sort()

    # for i in range(len(ExelineSet[7])):
    #     for j in range(len(ExelineSet[7][i])):
    #         for m in range(len(ExelineSet[7][i][j])):
    #             for n in range(len(ExelineSet[7][i][j][m])):
    #                 if 7142 in ExelineSet[7][i][j][m][n]:
    #                     print(i, j, m, n)
