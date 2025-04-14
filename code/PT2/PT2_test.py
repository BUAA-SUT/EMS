import shutil
from typing import Dict

from PT2 import *
from openpyxl import load_workbook
from publicFun import *
import json
import sys
import os
import time
import zipfile
import csv
import subprocess
import threading
import multiprocessing
import chardet
import codecs
import io
random.seed(1)


def copyFile(fileDir, tarDir, picknumber):
    pathDir = os.listdir(fileDir)  # 取图片的原始路径
    sample = random.sample(pathDir, picknumber)  # 随机选取picknumber数量的样本图片
    for i in range(len(sample)):
        shutil.copy(fileDir + sample[i], tarDir + "input{}.txt".format(i))
    return


def movefile(tarpath, tardir, oridir):
    if os.path.exists(tardir):
        if os.path.exists(tardir + '/' + tarpath):
            # os.remove(tarpath)
            pass
        else:
            shutil.move(oridir + tarpath, tardir)
    else:
        os.makedirs(tardir)
        shutil.move(oridir + tarpath, tardir)


def zip_file(src_dir):

    zip_name = src_dir + '.zip'

    z = zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED)

    for dirpath, dirnames, filenames in os.walk(src_dir):

        fpath = dirpath.replace(src_dir, '')

        fpath = fpath and fpath + os.sep or ''

        for filename in filenames:

            z.write(os.path.join(dirpath, filename), fpath+filename)

    z.close()


def un_zip(file_name):
    """unzip zip file"""
    zip_file = zipfile.ZipFile(file_name)
    tardir = '/Applications/work/data/MT/FS/PT/'
    if os.path.exists(file_name[:-4]):
        return
    for names in zip_file.namelist():
        zip_file.extract(names, tardir)
    zip_file.close()


# def getOriginalInput(num):
#     source_folder = '/Applications/work/data/MT/MFT/PT/input/'  # 源文件夹路径
#     target_file = '/Users/rendaixu/OneDrive/data/MT/STVR/PT2/RandomInput.csv'  # 移动到新的文件夹路径
#     txt_files = [f for f in os.listdir(source_folder) if f.endswith('.txt')]
#     random.shuffle(txt_files)  # 随机选择 100 个文件
#     with open(target_file, 'w', newline='', encoding='utf-8') as f:
#         writer = csv.writer(f)
#         writer.writerow(["name", "value"])
#         n = 1
#         for file_name in txt_files:
#             source_file = os.path.join(source_folder, file_name)  # 原始文件路径
#             with open(source_file, 'r', encoding='utf-8') as f2:
#                 file_content = f2.read().strip()  # 读取文件内容并去除空格
#                 file_content = file_content.replace('\x00', '')
#                 if not file_content:
#                     continue  # 如果文件内容为空，则跳过该文件
#                 writer.writerow(['input{}'.format(n-1), file_content])  # 将文件名和内容写入 CSV 文件中
#                 n += 1
#             if n == num+1:
#                 break


def getOriginalInput(num):
    source_folder = '/Applications/work/data/MT/MFT/PT/input/'  # 源文件夹路径
    target_folder = '/Users/rendaixu/OneDrive/data/MT/STVR/PT2/RandomInput'  # 移动到新的文件夹路径
    txt_files = [f for f in os.listdir(source_folder) if f.endswith('.txt')]
    random.shuffle(txt_files)
    n = 0
    for file_name in txt_files:
        source_file = os.path.join(source_folder, file_name)  # 原始文件路径
        target_file = os.path.join(target_folder, 'input{}.txt'.format(n))  # 目标文件路径
        with open(source_file, 'r') as f:
            file_content = f.read().strip()  # 读取文件内容并去除空格
            file_content = file_content.replace('\x00', '')
            if not file_content:
                continue  # 如果文件内容为空，则跳过该文件
        shutil.copy(source_file, target_file)  # 复制文件到目标文件夹中
        n += 1
        if n == num:
            break


def FailureRate(mu):
    # 把SourceCases读出来
    Result = []
    for i in range(1000):
        ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(0, i))
        original_output = MR().getResults(ts)
        program_ts = ts
        program_ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(mu, i))
        program_output = MR().getResults(program_ts)
        isViolate = MR().assertViolation(original_output, program_output)
        if isViolate:
            Result.append(1)
        else:
            Result.append(0)
    FR = round(Result.count(1) / len(Result) * 100, 2)
    return FR


def getMG(mu, mr_list, num_of_samples, originaloutput, ts, removecase):
    outputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/PT2/RandomOutput{}.csv'.format(mu)
    csvFile = open(outputdir, "r")
    reader = csv.reader(csvFile)
    mutateoutput = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            mutateoutput[item[0]] = item[1].split("\n")
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    MGS = []
    for i in range(num_of_samples):
        if i in removecase:
            continue
        MGs = []
        MG = [0] * len(mr_list)
        ts.setInputOutput("input{}".format(i), "output{}".format(i))
        original_output = MR().getResults(ts, mutateoutput)
        followup_ts = copy.copy(ts)
        for j in range(len(mr_list)):
            followup_ts.setInputOutput("input{}_{}".format(i, j), "output{}_{}".format(i, j))
            mr = mr_list[j]
            followup_output = mr.getResults(followup_ts, mutateoutput)
            expected_output = mr.getExpectedOutput(ts, original_output)
            isViolate = mr.assertViolation(expected_output, followup_output)
            if isViolate:
                MG[j] = 1
            elif (mutateoutput["output{}".format(i)] != originaloutput["output{}".format(i)] or
                  mutateoutput["output{}_{}".format(i, j)] != originaloutput["output{}_{}".format(i, j)]):
                MG[j] = 3
        MGs.append(MG)
        for m in range(len(mr_list)):
            MG = [0] * len(mr_list)
            ts.setInputOutput("input{}_{}".format(i, m), "output{}_{}".format(i, m))
            original_output = MR().getResults(ts, mutateoutput)
            followup_ts = copy.copy(ts)
            for n in range(len(mr_list)):
                followup_ts.setInputOutput("input{}_{}_{}".format(i, m, n), "output{}_{}_{}".format(i, m, n))
                mr = mr_list[n]
                followup_output = mr.getResults(followup_ts, mutateoutput)
                expected_output = mr.getExpectedOutput(ts, original_output)
                isViolate = mr.assertViolation(expected_output, followup_output)
                if isViolate:
                    MG[n] = 1
                elif (mutateoutput["output{}_{}".format(i, m)] != originaloutput["output{}_{}".format(i, m)] or
                      mutateoutput["output{}_{}_{}".format(i, m, n)] != originaloutput["output{}_{}_{}".format(i, m, n)]):
                    MG[n] = 3
            MGs.append(MG)
        MGS.append(MGs)
    return MGS


def get_input():
    # inputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/PT/RandomInput.csv'
    inputdir = '/home/rdx/data/MT/STVR/PT2/RandomInput.csv'
    # 建立空字典
    csvFile = open(inputdir, "r")
    # reader = csv.reader(csvFile)
    reader = csv.reader((line.replace('\0', '') for line in csvFile))
    # 建立空字典
    inputcase = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            inputcase[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()
    return inputcase


def dataTrans(mr, num_of_samples):
    inputs_write = '/Users/rendaixu/OneDrive/data/MT/STVR/PT2/RandomInput.csv'
    # outputs_write = '/Users/rendaixu/OneDrive/data/MT/FS/PT/RandomOutput'
    inputs_read = '/Users/rendaixu/OneDrive/data/MT/STVR/PT2/RandomInput'
    # outputs_read = '/Applications/work/data/MT/FS/PT/RandomOutput'
    fileHeader = ["name", "value"]
    inputdata = [fileHeader]

    for i in range(num_of_samples):
        f = open(inputs_read+'/input'+str(i)+'.txt')
        data = ['input{}'.format(i), f.read()]
        inputdata.append(data)
        for j in range(mr):
            f = open(inputs_read + '/input{}_{}.txt'.format(i, j))
            data = ['input{}_{}'.format(i, j), f.read()]
            inputdata.append(data)
            for k in range(mr):
                f = open(inputs_read + '/input{}_{}_{}.txt'.format(i, j, k))
                data = ['input{}_{}_{}'.format(i, j, k), f.read()]
                inputdata.append(data)
    # 将数据存下来
    with open(inputs_write, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(inputdata)


def verify(mr, num_of_samples):
    inputs_read = '/Applications/work/data/MT/MFT/PT/RandomOutput'
    fileHeader = ["name", "value"]
    inputdata = [fileHeader]
    for i in range(num_of_samples):
        f = open(inputs_read+'/output0_'+str(i)+'.txt')
        data = ['output0_{}'.format(i), f.read()]
        inputdata.append(data)
        for j in range(mr):
            f = open(inputs_read + '/output0_{}_{}.txt'.format(i, j))
            data = ['output0_{}_{}'.format(i, j), f.read()]
            inputdata.append(data)
            for k in range(mr):
                f = open(inputs_read + '/output0_{}_{}_{}.txt'.format(i, j, k))
                data = ['output0_{}_{}_{}'.format(i, j, k), f.read()]
                inputdata.append(data)
    inputcase = get_input()
    lst1 = list(inputcase.items())
    lst2 = list(inputcase.items())
    for i in range(len(lst1)):
        if not lst1[i][1] == lst2[i][1]:
            print(i)


def getTestcase(mr_list, test_case, num_of_samples):
    # for i in range(num_of_samples):
    #     test_case.setInputOutput("infile_{}".format(i), "outfile_{}".format(i), "outtree_{}".format(i))
    #     test_case.generateRandomTestcase()
    # for i in range(num_of_samples):
    #     for j in range(len(mr_list)):
    #         mr = mr_list[j]
    #         mr.setTestCase(test_case)
    #         mr.original_ts.setInputOutput("input{}.fa".format(i), "reference{}.fa".format(i), "e{}.fa".format(i), "output{}.txt".format(i))
    #         mr.getFollowInput(j)
    for i in range(num_of_samples):
        # test_case.setInputOutput("input{}.txt".format(i), "output{}.txt".format(i))
        # test_case.generateInput()  # 生成原始测试用例
        for j in range(len(mr_list)):
            mr = mr_list[j]
            mr.setTestCase(test_case)
            mr.original_ts.setInputOutput("input{}.txt".format(i), "output{}.txt".format(i))
            mr.getFollowInput(j)
            for k in range(len(mr_list)):
                mr = mr_list[k]
                mr.setTestCase(test_case)
                mr.original_ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}.txt".format(i, j))
                mr.getFollowInput(k)


def getOutput(mu, inputcase):
    fileHeader = ["name", "value"]
    outputdata = [fileHeader]
    command = './Mutants/printtokens2_v{}/print_tokens2'.format(mu)
    lst = list(inputcase.items())
    for i in range(len(lst)):
        # 将输入字符串写入临时文件
        try:
            output = subprocess.run(command, input=lst[i][1].encode(), stdout=subprocess.PIPE,
                                    stderr=subprocess.STDOUT, shell=True).stdout.decode()
        except UnicodeDecodeError:
            print("Mutant{} 编码格式问题，跳过".format(mu))
            return 0
        data = ['output'+lst[i][0][5:], output]
        outputdata.append(data)
    # outputs_write = '/Users/rendaixu/OneDrive/data/MT/STVR/PT/RandomOutput{}.csv'.format(mu)
    outputs_write = '/home/rdx/data/MT/STVR/PT2/RandomOutput{}.csv'.format(mu)
    with open(outputs_write, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(outputdata)
    print('第{}个已完成'.format(mu))


def mutaterate(mu):
    outputdir1 = '/Users/rendaixu/OneDrive/data/MT/STVR/PT2/RandomOutput0.csv'
    outputdir2 = '/Users/rendaixu/OneDrive/data/MT/STVR/PT2/RandomOutput{}.csv'.format(mu)
    if not os.path.exists(outputdir2):
        return 0
    csvFile = open(outputdir1, "r")
    reader = csv.reader(csvFile)
    # 建立空字典
    outputcase1 = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase1[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    csvFile = open(outputdir2, "r")
    reader = csv.reader(csvFile)
    outputcase2 = {}
    for item in reader:
        # 忽略第一行
        if reader.line_num == 1:
            continue
        try:
            outputcase2[item[0]] = item[1]
        except:
            print(item)
            print(reader.line_num)
            break
    csvFile.close()

    num = 0
    outputcase1 = list(outputcase1.values())
    outputcase2 = list(outputcase2.values())
    index = []
    for i in range(len(outputcase1)):
        if not outputcase1[i] == outputcase2[i]:
            num += 1
            index.append(i)
    rate = num / (len(outputcase1))
    print(mu, round(rate*100, 2))
    return index


def run_getOutput(i, inputcase, timeout):
    p = multiprocessing.Process(target=getOutput, args=(i, inputcase))
    p.start()
    p.join(timeout)
    if p.is_alive():
        p.terminate()
        print("Mutant{} 超时，跳过".format(i))


if __name__ == "__main__":
    num_case = 100
    removecase = [3, 41, 68]
    num_mr = 11
    num_mu = 14
    # getOriginalInput(num_case)
    # dataTrans(num_mr, num_case)
    # inputcase = get_input()
    # timeout = 120  # 设定阈值为**秒
    # for i in range(5, 50):
    #     run_getOutput(i, inputcase, timeout)
    # for mu in range(1, 50):
    #     mutaterate(mu)
    # outputdir = '/Users/rendaixu/OneDrive/data/MT/STVR/PT2/RandomOutput0.csv'
    # csvFile = open(outputdir, "r")
    # reader = csv.reader(csvFile)
    # # 建立空字典
    # originaloutput = {}
    # for item in reader:
    #     # 忽略第一行
    #     if reader.line_num == 1:
    #         continue
    #     try:
    #         originaloutput[item[0]] = item[1].split("\n")
    #     except:
    #         print(item)
    #         print(reader.line_num)
    #         break
    # csvFile.close()
    # mr_list = [MR1(), MR2(), MR3(), MR4(), MR5(), MR6(), MR7(), MR8(), MR9(), MR10(), MR11()]
    # ts = TestCase()
    # getTestcase(mr_list, ts, num_case)


    # MGS_set = []
    # SMGS_set = []
    # for mu in range(1, 24):
    #     MGS, SMGS = getMG(mu, mr_list, num_of_samples, originaloutput, ts)
    #     MGS_set.append(MGS)
    #     SMGS_set.append(SMGS)

    # verify(mr, num_of_samples)
    # # myenv = MyEnv()
    # # myenv.CreateWorkingDirs()
    project = 'STVR'
    row = 1
    string = 'PT2'
    path = '/Applications/work/data/MT/' + project + '/Result/result25_one.xlsx'  # '+sys.argv[1][:-1]+'
    wb = load_workbook(path)
    if string not in wb.sheetnames:
        ws = wb.create_sheet(string)
    del wb[string]
    ws = wb.create_sheet(string)
    MG_set = []
    M_set = []
    ExecutableS_set = []
    Sus_set = []
    Flag_set = []
    # inters = set(ExelinesS[0][0])
    # # 逐个计算交集
    # for i in range(len(ExelinesS)):
    #     for j in range(len(ExelinesS[0])):
    #         # if len(ExelinesS[i][j]) < 50:
    #         #     print(i,j)
    #         #     break
    #         inters = inters.intersection(set(ExelinesS[i][j]))
    #     # print(i, len(list(inters)))
    # # 将结果转换为列表
    # inters = list(inters)
    # inters.sort()
    # for i in range(len(ExelinesS)):
    #     for j in range(len(ExelinesS[0])):
    #         ExelinesS[i][j] = [x for x in ExelinesS[i][j] if x not in inters]
    # removemg = []
    # for i in range(len(ExelinesS)):
    #     a = 0
    #     for j in range(len(ExelinesS[i])):
    #         if len(ExelinesS[i][j]) == 0:
    #             a += 1
    #     if a == len(ExelinesS[i]):
    #        removemg.append(i)
    # for index in sorted(removemg, reverse=True):
    #     del ExelinesS[index]
    removemu = [1, 2, 3]
    # for mu in range(1, num_mu+1):
    #     with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
    #               'r') as load_f:
    #         data = json.load(load_f)
    #     data = json.loads(data)
    #     Flag = data['Flag']
    #     fault = Executable[Flag.index(1)]
    #     if fault in inters:
    #         removemu.append(mu)
    for mu in range(1, num_mu+1):  # 1, num_mu+1
        if mu in removemu:
            continue
        # datadir = '/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/statementResult{}.csv'.format(mu)
        # csvFile = open(datadir, "r")
        # reader = csv.reader(csvFile)
        # # # 建立空字典
        # originaldata = {}
        # result = {}
        # for item in reader:
        #     originaldata[item[0]] = item[1:-1]
        #     if reader.line_num == 1:
        #         continue
        #     result[item[0]] = item[-1:]
        # csvFile.close()
        # ExecutableS = originaldata.get('inputs')
        # ExecutableS = [int(x) for x in ExecutableS]
        # ExelinesS = []
        # Result = []
        # for i in range(num_case):
        #     if i in removecase:
        #         continue
        #     Exelines = []
        #     r = []
        #     candidate = originaldata.get('input{}'.format(i))
        #     candidate = [int(x) for x in candidate]
        #     indices = [i for i in range(len(candidate)) if candidate[i] == 1]
        #     Exelines.append([ExecutableS[i] for i in indices])
        #     a = result.get('input{}'.format(i))
        #     a = [int(x) for x in a]
        #     r.append(a)
        #     for j in range(num_mr):
        #         candidate = originaldata.get('input{}_{}'.format(i, j))
        #         candidate = [int(x) for x in candidate]
        #         indices = [i for i in range(len(candidate)) if candidate[i] == 1]
        #         Exelines.append([ExecutableS[i] for i in indices])
        #         a = result.get('input{}_{}'.format(i, j))
        #         a = [int(x) for x in a]
        #         r.append(a)
        #     for m in range(num_mr):
        #         for n in range(num_mr):
        #             candidate = originaldata.get('input{}_{}_{}'.format(i, m, n))
        #             candidate = [int(x) for x in candidate]
        #             indices = [i for i in range(len(candidate)) if candidate[i] == 1]
        #             Exelines.append([ExecutableS[i] for i in indices])
        #             a = result.get('input{}_{}_{}'.format(i, m, n))
        #             a = [int(x) for x in a]
        #             r.append(a)
        #     Result.append(r)
        #     ExelinesS.append(Exelines)
        # data = {
        #     'Exec': ExecutableS, 'Exel': ExelinesS, 'Result': Result
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/statements{}.json'.format(mu), 'w') as f:
        #     json.dump(json_str, f)
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/statements{}.json'.format(mu), 'r') as f:
            data = json.load(f)
        data = json.loads(data)
        ExelineS = data['Exel']
        Executable = data['Exec']
        # Result = data['Result']
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/statements.json', 'r') as f:
        #     data = json.load(f)
        # data = json.loads(data)
        # ExelinesS1 = data['Exel']
        # Executable1 = data['Exec']
        # MGS = getMG(mu, mr_list, num_case, originaloutput, ts, removecase)
        # data['MGS'] = MGS
        with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
                  'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        # SMGS = data['SMGS']
        MGS = data['MGS']
        # ExecutableS = data['Exec']
        # ExelinesS = data['Exel']
        Flag = data['Flag']
    #     # Sus = MSlice(SMGS, ExecutableS, ExelinesS)
    #     # FaSus, percent = FAILTIMSlice(SMGS, ExecutableS, ExelinesS, Flag)
    #     # mSus = data['Sus']
    #     # mSus2, AllF = MSlice2(SMGS, ExecutableS, ExelinesS, Flag)
    #     # mSus = data['Sus']
    #     # Flag = [0] * len(Executable)
    #     # if mu == 1:
    #     #     Flag[49] = 1
    #     # elif mu == 2:
    #     #     Flag[50] = 1
    #     # elif mu == 3:
    #     #     Flag[51] = 1
    #     # elif mu == 4:
    #     #     Flag[66] = 1
    #     # elif mu == 5:
    #     #     Flag[70] = 1
    #     # elif mu == 6:
    #     #     Flag[70] = 1
    #     # elif mu == 7:
    #     #     Flag[80] = 1
    #     # elif mu == 8:
    #     #     Flag[84] = 1
    #     # elif mu == 9:
    #     #     Flag[85] = 1
    #     # elif mu == 10:
    #     #     Flag[89] = 1
    #     # elif mu == 11:
    #     #     Flag[89] = 1
    #     # elif mu == 12:
    #     #     Flag[145] = 1
    #     # elif mu == 13:
    #     #     Flag[161] = 1
    #     # elif mu == 14:
    #     #     Flag[40] = 1
    #     # data['Flag'] = Flag
    #     # mSus, metric = MSlice(MGS, ExecutableS, ExelinesS)
    #     # mSus2, AllF, metric2 = MSlice2(MGS, ExecutableS, ExelinesS, Flag)
    #     # if mu in removemu:
    #     #     continue
    # #     with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
    # #               'r') as load_f:
    # #         data = json.load(load_f)
    # #     data = json.loads(data)
    #     mSus = data['mSus']
    # #     # mSus2 = data['mSus2']
    # #     # mSus3 = data['mSus3']
    #     mSus4 = data['mSus4']
    # #     # metric = data['metric']
    # #     # metric2 = data['metric2']
    # #     # metric3 = data['metric3']
    # #     # metric4 = data['metric4']
    # #     Flag = data['Flag']
    # #     MGS = data['MGS']
        Result = data['result']
        mSus4 = data['mSus4_one']
        mSus = data['mSus_one']
        # mSus4 = data['mSus4_all']
        # mSus = data['mSus_all']
        Para = data['Para']
        # Para = data['Para']
        # mSus4 = data['mSus4_all']
        # mSus = data['mSus_all']
        # metric4 = data['metric4_all']
        # metric = data['metric_all']
        # mSus4_one = data['mSus4_one']
        # mSus_one = data['mSus_one']
        # mSus4_nofs = data['mSus4_nofs']
        # mSus_nofs = data['mSus_nofs']
    # #     # AllF = data['AllF']
    # #     # Para = data['Para']
    # #     # Sus = []
    # #     # Metric = []
    # #     # for i in range(len(mSus)):
    # #     #     a = [round(mSus[i], 4)]
    # #     #     b = [metric[i]]
    # #     #     for j in range(len(mSus3)):
    # #     #         a.append(mSus3[j][5][i])
    # #     #         b.append(metric3[j][i])
    # #     #     Sus.append(a)
    # #     #     Metric.append(b)
    # #     # Sus_set.append(Sus)
    # #     # Metric_set.append(Metric)
    #     mSus_nofs, metric_nofs = MSlice(MGS, Executable, ExelineS)
    #     data['mSus_nofs'] = mSus_nofs
    #     data['metric_nofs'] = metric_nofs
    #     # mSus3, AllF, _ = MSlice3(MGS, Executable, ExelinesS, Flag)
    #     # Union_set.append(Union)
    #     # Unique_set.append(Unique)
    #     result = []
    #     for i in range(len(Result)):
    #         a = []
    #         for j in range(len(Result[i])):
    #             a.append(Result[i][j][0])
    #         result.append(a)
    #     data['result'] = result
    #     mSus4_nofs, Para, metric4_nofs = MSlice4(MGS, Executable, ExelineS, Flag, result)
        # data['Para'] = Para
        # data['mSus4_nofs'] = mSus4_nofs
        # data['metric4_nofs'] = metric4_nofs
        # mSus, metric = MSlice_one(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus_one'] = mSus
        # data['metric_one'] = metric
        # mSus, metric = MSlice_all(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus_all'] = mSus
        # data['metric_all'] = metric
        # mSus4, _, metric4 = MSlice4_one(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus4_one'] = mSus4
        # data['metric4_one'] = metric4
        # data['Para'] = Para
        # mSus4, _, metric4 = MSlice4_all(MGS, Executable, ExelineS, Flag, Result)
        # data['mSus4_all'] = mSus4
        # data['metric4_all'] = metric4
        # data = {
        #          'mSus': mSus, 'mSus4': mSus4, 'metric': metric,
        #     'metric4': metric4, 'Para': Para, 'MGS': MGS, 'Flag': Flag,
        # }
        # json_str = json.dumps(data)
        # with open('/Users/rendaixu/OneDrive/data/MT/' + project + '/' + string + '/mutant' + str(mu) + '.json',
        #           'w') as f:
        #     json.dump(json_str, f)
        row = eval('getMetrics_5')(row, ws, mu, MGS, mSus, mSus4, Para[-1], Flag, Para, ExelineS)
    wb.save(path)

